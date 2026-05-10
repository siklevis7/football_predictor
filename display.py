"""
display.py — print_ticket, print_upcoming, and ExcelTracker.
"""

import json
import numpy as np
import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import config
from betting import (
    implied, edge, ev, kelly_stake,
    bet_quality, quality_icon, should_bet,
)


class ExcelTracker:
    """
    Creates and appends to predictions_tracker.xlsx.
    One row per prediction appended automatically after each prediction.
    Manual columns (actual result, bet placed, etc.) left blank for you to fill.
    """

    PRED_HEADERS = [
        # Auto-filled
        "ID", "Date", "League", "Home Team", "Away Team",
        "Home xG", "Away xG", "Total xG",
        "Pred 1X2", "Home Win%", "Draw%", "Away Win%",
        "Over 2.5%", "Under 2.5%", "Over 1.5%", "Over 3.5%",
        "BTTS Yes%", "BTTS No%",
        "Top Correct Score", "Top CS Prob%",
        "HT Home Win%", "HT Draw%", "HT Away Win%",
        "Half Most Goals", "HT/FT Most Likely",
        "Win To Nil Home%", "Win To Nil Away%",
        "Draw No Bet Home%", "Draw No Bet Away%",
        "Exp Corners Total", "Exp Cards Total", "Exp Booking Pts",
        "Best Market", "Edge%", "EV", "Kelly Stake (RWF)", "Bankroll (RWF)",
        "Odds Entered",
        "H2H Factor", "Inj Factor Home", "Inj Factor Away",
        "Momentum Home", "Momentum Away",
        "Injury Notes Home", "Injury Notes Away",
        # Manual columns
        "ACTUAL Score", "ACTUAL 1X2 (H/D/A)", "ACTUAL Goals",
        "ACTUAL BTTS (Y/N)", "ACTUAL HT Score",
        "ACTUAL Corners", "ACTUAL Cards",
        "Bet Placed? (Y/N)", "Market Bet On",
        "Stake Placed (RWF)", "Odds Taken", "Result (W/L/P)",
        "Profit/Loss (RWF)",
        # Formula columns
        "1X2 Correct?", "O/U 2.5 Correct?", "BTTS Correct?",
        "Corners Correct?", "Cards Correct?",
        "Closing Odds (fill after)", "CLV (Closing Line Value)",
        "Bankroll After (RWF)", "Running ROI%",
    ]

    PERF_HEADERS = [
        "Metric", "Value"
    ]

    # Column indices (1-based) for formula references
    COL_PRED_1X2      = 8   # H
    COL_HOME_WIN_PCT  = 9   # I
    COL_ACTUAL_1X2    = 45  # AS (ACTUAL 1X2)
    COL_ACTUAL_GOALS  = 46  # AT
    COL_OVER25_PCT    = 12  # L
    COL_BET_PLACED    = 51  # AY
    COL_STAKE         = 53  # BA
    COL_RESULT        = 55  # BC
    COL_PL            = 56  # BD  Profit/Loss
    COL_1X2_CORRECT   = 57  # BE
    COL_OU_CORRECT    = 58  # BF
    COL_BTTS_CORRECT  = 59  # BG
    COL_BANKROLL_AFT  = 60  # BH
    COL_ROI           = 61  # BI

    # Styling
    HDR_FILL_AUTO   = PatternFill("solid", fgColor="1F3864")   # dark blue
    HDR_FILL_MANUAL = PatternFill("solid", fgColor="833C00")   # dark orange
    HDR_FILL_CALC   = PatternFill("solid", fgColor="375623")   # dark green
    HDR_FONT        = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    DATA_FONT       = Font(name="Arial", size=9)
    PERF_HDR_FILL   = PatternFill("solid", fgColor="2E4057")
    BORDER          = Border(
        left  = Side(style="thin", color="CCCCCC"),
        right = Side(style="thin", color="CCCCCC"),
        top   = Side(style="thin", color="CCCCCC"),
        bottom= Side(style="thin", color="CCCCCC"),
    )

    def __init__(self, path: Path):
        self.path = path
        if not path.exists():
            self._create()

    def _create(self):
        wb = openpyxl.Workbook()

        # ── Predictions sheet ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Predictions"
        ws.freeze_panes = "E2"

        n_auto   = 44   # columns A–AR  (auto-filled by script)
        n_manual = 13   # columns AS–BE (you fill in)
        n_calc   = 5    # columns BF–BJ (Excel formulas)

        for col_idx, header in enumerate(self.PRED_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = self.HDR_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = self.BORDER

            if col_idx <= n_auto:
                cell.fill = self.HDR_FILL_AUTO
            elif col_idx <= n_auto + n_manual:
                cell.fill = self.HDR_FILL_MANUAL
            else:
                cell.fill = self.HDR_FILL_CALC

        # Column widths
        col_widths = {
            1: 6, 2: 12, 3: 22, 4: 22,
            5: 9, 6: 9, 7: 9,
            8: 12, 9: 9, 10: 9, 11: 9,
            12: 10, 13: 10, 14: 10, 15: 10,
            16: 10, 17: 10,
            18: 14, 19: 11,
            20: 13, 21: 10, 22: 13,
            23: 14, 24: 14,
            25: 14, 26: 14, 27: 14, 28: 14,
            29: 12, 30: 12, 31: 12,
            32: 16, 33: 8, 34: 8, 35: 15, 36: 15,
            37: 16,
            38: 10, 39: 12, 40: 12, 41: 11, 42: 11,
            43: 22, 44: 22,
        }
        for c, w in col_widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w
        # Manual + calc columns
        for c in range(45, len(self.PRED_HEADERS)+2):
            ws.column_dimensions[get_column_letter(c)].width = 14
        ws.row_dimensions[1].height = 36

        # ── Performance sheet ──────────────────────────────────────────────────
        wp = wb.create_sheet("Performance")
        wp.freeze_panes = "A2"

        perf_metrics = [
            ("Total Predictions",   f"=COUNTA(Predictions!B2:B10000)"),
            ("Bets Placed",         f"=COUNTIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\")"),
            ("Bets Won",            f"=COUNTIF(Predictions!{get_column_letter(56)}2:{get_column_letter(56)}10000,\"W\")"),
            ("Bets Lost",           f"=COUNTIF(Predictions!{get_column_letter(56)}2:{get_column_letter(56)}10000,\"L\")"),
            ("Strike Rate %",       f"=IFERROR(B4/B3*100,0)"),
            ("Total Staked (RWF)",  f"=IFERROR(SUMIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!{get_column_letter(54)}2:{get_column_letter(54)}10000),0)"),
            ("Total Return (RWF)",  f"=IFERROR(SUMIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!{get_column_letter(57)}2:{get_column_letter(57)}10000)+B7,0)"),
            ("Net Profit (RWF)",    f"=B8-B7"),
            ("ROI %",               f"=IFERROR(B9/B7*100,0)"),
            ("1X2 Accuracy %",      f"=IFERROR(COUNTIF(Predictions!{get_column_letter(58)}2:{get_column_letter(58)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(58)}2:{get_column_letter(58)}10000)*100,0)"),
            ("O/U 2.5 Accuracy %",  f"=IFERROR(COUNTIF(Predictions!{get_column_letter(59)}2:{get_column_letter(59)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(59)}2:{get_column_letter(59)}10000)*100,0)"),
            ("BTTS Accuracy %",     f"=IFERROR(COUNTIF(Predictions!{get_column_letter(60)}2:{get_column_letter(60)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(60)}2:{get_column_letter(60)}10000)*100,0)"),
            ("Corners Accuracy %",  f"=IFERROR(COUNTIF(Predictions!{get_column_letter(61)}2:{get_column_letter(61)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(61)}2:{get_column_letter(61)}10000)*100,0)"),
            ("Cards Accuracy %",    f"=IFERROR(COUNTIF(Predictions!{get_column_letter(62)}2:{get_column_letter(62)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(62)}2:{get_column_letter(62)}10000)*100,0)"),
            ("Avg CLV % (bets)",    f"=IFERROR(AVERAGEIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!{get_column_letter(64)}2:{get_column_letter(64)}10000),0)"),
            ("Avg Edge % (bets)",   f"=IFERROR(AVERAGEIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!AH2:AH10000),0)"),
            ("Avg EV (bets)",       f"=IFERROR(AVERAGEIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!AI2:AI10000),0)"),
            ("Current Bankroll",    f"=IFERROR(INDEX(Predictions!{get_column_letter(65)}2:{get_column_letter(65)}10000,MATCH(9E+307,Predictions!{get_column_letter(65)}2:{get_column_letter(65)}10000)),0)"),
        ]

        wp["A1"] = "Metric"; wp["B1"] = "Value"
        for cell in [wp["A1"], wp["B1"]]:
            cell.fill = self.PERF_HDR_FILL
            cell.font = self.HDR_FONT
            cell.alignment = Alignment(horizontal="center")

        for r, (metric, formula) in enumerate(perf_metrics, start=2):
            wp.cell(row=r, column=1, value=metric).font = Font(bold=True, name="Arial", size=10)
            cell = wp.cell(row=r, column=2, value=formula)
            cell.font = Font(name="Arial", size=10)
            cell.number_format = "#,##0.00"

        wp.column_dimensions["A"].width = 26
        wp.column_dimensions["B"].width = 20

        # ── Legend sheet ───────────────────────────────────────────────────────
        wl = wb.create_sheet("Legend")
        legend_rows = [
            ("COLUMN COLOR GUIDE", ""),
            ("Dark Blue headers", "Auto-filled by prediction script"),
            ("Dark Orange headers", "You fill in manually after the match"),
            ("Dark Green headers", "Calculated automatically by Excel formulas"),
            ("", ""),
            ("MANUAL COLUMN GUIDE", ""),
            ("ACTUAL Score", "e.g. 2-1"),
            ("ACTUAL 1X2", "H = Home win, D = Draw, A = Away win"),
            ("ACTUAL Goals", "Total goals scored"),
            ("ACTUAL BTTS", "Y if both teams scored, N if not"),
            ("ACTUAL HT Score", "e.g. 1-0"),
            ("Bet Placed?", "Y or N"),
            ("Market Bet On", "e.g. over_2.5, home_win, btts_yes"),
            ("Stake Placed (RWF)", "Exact amount you staked"),
            ("Odds Taken", "Decimal odds you got"),
            ("Result (W/L/P)", "W=Win, L=Loss, P=Push/Void"),
            ("Profit/Loss (RWF)", "Positive if won, negative if lost"),
            ("", ""),
            ("POSITION CODES (for lineup entry)", ""),
            ("GK", "Goalkeeper"),
            ("DF", "Defender"),
            ("MF", "Midfielder"),
            ("FW", "Forward"),
        ]
        wl["A1"] = "PREDICTIONS TRACKER — GUIDE"
        wl["A1"].font = Font(bold=True, size=14, name="Arial")
        for r, (col, desc) in enumerate(legend_rows, start=3):
            wl.cell(row=r, column=1, value=col).font  = Font(bold=bool(col and "GUIDE" in col or "CODES" in col), name="Arial")
            wl.cell(row=r, column=2, value=desc).font = Font(name="Arial")
        wl.column_dimensions["A"].width = 35
        wl.column_dimensions["B"].width = 50

        try:
            wb.save(self.path)
            print(f"  [Tracker] Created {self.path.name} with Predictions, Performance, Legend sheets.")
        except PermissionError:
            print(f"  ⚠  Could not create tracker — close {self.path.name} in Excel first.")

    def append_prediction(self, data: dict):
        """Append one prediction row to the Predictions sheet."""
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Predictions"]
        next_row = ws.max_row + 1
        pred_id  = next_row - 1

        # Determine predicted 1X2
        probs_1x2 = {
            "Home Win": data.get("home_win_pct", 0),
            "Draw"    : data.get("draw_pct", 0),
            "Away Win": data.get("away_win_pct", 0),
        }
        pred_1x2 = max(probs_1x2, key=probs_1x2.get)
        half_most = max(
            {"1st Half": data.get("first_half_pct",0),
             "2nd Half": data.get("second_half_pct",0),
             "Equal"   : data.get("equal_pct",0)},
            key=lambda k: {"1st Half": data.get("first_half_pct",0),
                           "2nd Half": data.get("second_half_pct",0),
                           "Equal"   : data.get("equal_pct",0)}[k]
        )

        # Build the HTFT most likely label
        htft = data.get("htft", {})
        htft_best = max(htft, key=htft.get) if htft else "D/H"

        auto_values = [
            pred_id,
            data.get("date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            data.get("league_code", ""),
            data.get("home_team", ""),
            data.get("away_team", ""),
            round(data.get("xg_h", 0), 2),
            round(data.get("xg_a", 0), 2),
            round(data.get("xg_h", 0) + data.get("xg_a", 0), 2),
            pred_1x2,
            round(data.get("home_win_pct", 0)*100, 1),
            round(data.get("draw_pct", 0)*100, 1),
            round(data.get("away_win_pct", 0)*100, 1),
            round(data.get("over25_pct", 0)*100, 1),
            round(data.get("under25_pct", 0)*100, 1),
            round(data.get("over15_pct", 0)*100, 1),
            round(data.get("over35_pct", 0)*100, 1),
            round(data.get("btts_yes_pct", 0)*100, 1),
            round(data.get("btts_no_pct", 0)*100, 1),
            data.get("top_correct_score", ""),
            round(data.get("top_cs_pct", 0)*100, 1),
            round(data.get("ht_hw_pct", 0)*100, 1),
            round(data.get("ht_d_pct", 0)*100, 1),
            round(data.get("ht_aw_pct", 0)*100, 1),
            half_most,
            htft_best,
            round(data.get("home_wtn_pct", 0)*100, 1),
            round(data.get("away_wtn_pct", 0)*100, 1),
            round(data.get("dnb_home_pct", 0)*100, 1),
            round(data.get("dnb_away_pct", 0)*100, 1),
            round(data.get("exp_corners", 0), 1),
            round(data.get("exp_cards", 0), 1),
            round(data.get("exp_booking_pts", 0), 1),
            data.get("best_market", ""),
            round(data.get("edge_pct", 0)*100, 2),
            round(data.get("ev", 0), 3),
            data.get("kelly_stake_rwf", 0),
            data.get("bankroll_rwf", 0),
            data.get("odds_entered", ""),
            round(data.get("h2h_factor", 1.0), 3),
            round(data.get("inj_h_factor", 1.0), 3),
            round(data.get("inj_a_factor", 1.0), 3),
            round(data.get("mom_h", 1.0), 3),
            round(data.get("mom_a", 1.0), 3),
            data.get("injury_notes_home", ""),
            data.get("injury_notes_away", ""),
        ]

        # Write auto values
        for col_idx, val in enumerate(auto_values, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font   = self.DATA_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=False)

        # Skip manual columns (45–57) — leave blank
        n_auto   = len(auto_values)
        n_manual = 13
        for col_idx in range(n_auto+1, n_auto+n_manual+1):
            cell = ws.cell(row=next_row, column=col_idx, value="")
            cell.fill   = PatternFill("solid", fgColor="FFF2CC")
            cell.border = self.BORDER

        # Column layout after League column was added (1-based):
        # Auto  cols 1-45 (45 values)
        # Manual cols 46-58:
        #   46=ACTUAL Score, 47=ACTUAL 1X2, 48=ACTUAL Goals, 49=ACTUAL BTTS
        #   50=ACTUAL HT Score, 51=ACTUAL Corners, 52=ACTUAL Cards
        #   53=Bet Placed?, 54=Market Bet On, 55=Stake Placed
        #   56=Odds Taken, 57=Result, 58=Profit/Loss
        # Formula cols 59-67:
        #   59=1X2 Correct?, 60=O/U Correct?, 61=BTTS Correct?
        #   62=Corners Correct?, 63=Cards Correct?
        #   64=Closing Odds (manual), 65=CLV, 66=Bankroll After, 67=Running ROI%

        r   = next_row
        ac  = get_column_letter

        # Named column refs — all shifted +1 since League is now col 3
        c_pred_1x2    = ac(9)    # I  - Predicted 1X2
        c_over25      = ac(13)   # M  - Over 2.5%
        c_btts_yes    = ac(17)   # Q  - BTTS Yes%
        c_exp_corners = ac(30)   # AD - Expected corners
        c_exp_cards   = ac(31)   # AE - Expected cards
        c_bankroll    = ac(37)   # AK - Bankroll at prediction time
        c_act_score   = ac(46)   # AT - ACTUAL Score
        c_act_1x2     = ac(47)   # AU - ACTUAL 1X2
        c_act_goals   = ac(48)   # AV - ACTUAL Goals
        c_act_btts    = ac(49)   # AW - ACTUAL BTTS
        c_act_ht      = ac(50)   # AX - ACTUAL HT Score
        c_act_corners = ac(51)   # AY - ACTUAL Corners
        c_act_cards   = ac(52)   # AZ - ACTUAL Cards
        c_bet_placed  = ac(53)   # BA - Bet Placed?
        c_stake       = ac(55)   # BC - Stake Placed
        c_odds_taken  = ac(56)   # BD - Odds Taken
        c_result      = ac(57)   # BE - Result W/L/P
        c_pl          = ac(58)   # BF - Profit/Loss
        c_closing     = ac(64)   # BL - Closing Odds (manual)

        formula_cols = [
            # 59: 1X2 Correct?
            (59, f'=IF({c_act_1x2}{r}="","",'
                 f'IF(AND({c_pred_1x2}{r}="Home Win",{c_act_1x2}{r}="H"),"✓",'
                 f'IF(AND({c_pred_1x2}{r}="Draw",{c_act_1x2}{r}="D"),"✓",'
                 f'IF(AND({c_pred_1x2}{r}="Away Win",{c_act_1x2}{r}="A"),"✓","✗"))))'),

            # 60: O/U 2.5 Correct?
            (60, f'=IF({c_act_goals}{r}="","",IF(AND({c_over25}{r}>50,{c_act_goals}{r}>2),"✓",'
                 f'IF(AND({c_over25}{r}<=50,{c_act_goals}{r}<=2),"✓","✗")))'),

            # 61: BTTS Correct?
            (61, f'=IF({c_act_btts}{r}="","",IF(AND({c_btts_yes}{r}>50,{c_act_btts}{r}="Y"),"✓",'
                 f'IF(AND({c_btts_yes}{r}<=50,{c_act_btts}{r}="N"),"✓","✗")))'),

            # 62: Corners Correct?
            (62, f'=IF({c_act_corners}{r}="","",IF(AND({c_exp_corners}{r}>9.5,{c_act_corners}{r}>9),"✓",'
                 f'IF(AND({c_exp_corners}{r}<=9.5,{c_act_corners}{r}<=9),"✓","✗")))'),

            # 63: Cards Correct?
            (63, f'=IF({c_act_cards}{r}="","",IF(AND({c_exp_cards}{r}>3.5,{c_act_cards}{r}>3),"✓",'
                 f'IF(AND({c_exp_cards}{r}<=3.5,{c_act_cards}{r}<=3),"✓","✗")))'),

            # 64: Closing Odds — MANUAL
            (64, ""),

            # 65: CLV
            (65, f'=IFERROR(({c_odds_taken}{r}/{c_closing}{r}-1)*100,"")'),

            # 66: Bankroll After
            (66, f'=IF({c_pl}{r}="",{c_bankroll}{r},{c_bankroll}{r}+IF({c_pl}{r}="",0,{c_pl}{r}))'),

            # 67: Running ROI%
            (67, f'=IFERROR(SUMIF({c_bet_placed}$2:{c_bet_placed}{r},"Y",'
                 f'{c_pl}$2:{c_pl}{r})/SUMIF({c_bet_placed}$2:{c_bet_placed}{r},"Y",'
                 f'{c_stake}$2:{c_stake}{r})*100,0)'),
        ]

        for col_idx, formula in formula_cols:
            cell = ws.cell(row=next_row, column=col_idx, value=formula)
            if col_idx == 63:
                # Closing Odds — manual, styled like manual columns but green border
                cell.fill   = PatternFill("solid", fgColor="FFF2CC")
            else:
                cell.fill   = PatternFill("solid", fgColor="E2EFDA")
                cell.font   = Font(name="Arial", size=9, color="006100")
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        # Save with retry — Windows locks .xlsx files when open in Excel.
        # Give the user a chance to close it rather than crashing.
        for attempt in range(10):
            try:
                wb.save(self.path)
                print(f"  [Tracker] Row {pred_id} saved → {self.path.name}")
                return
            except PermissionError:
                if attempt == 0:
                    print()
                    print(f"  ⚠  Cannot save tracker — file is open in Excel.")
                    print(f"  ⚠  Close  {self.path.name}  then press Enter to retry …")
                input("  [Press Enter when Excel is closed] ")
        # If all retries exhausted, save to a temp file instead
        import tempfile, shutil
        tmp = self.path.with_suffix(".tmp.xlsx")
        wb.save(tmp)
        print(f"  ⚠  Could not save to original file after 10 attempts.")
        print(f"  ⚠  Saved to {tmp.name} instead — rename manually.")


# ══════════════════════════════════════════════════════════════════════════════
# TICKET PRINTER
# ══════════════════════════════════════════════════════════════════════════════
W = 74

def bar(p, w=24):
    f = int(min(max(float(p),0),1)*w)
    return "█"*f + "░"*(w-f)

def row(text):   return f"║ {str(text):<{W-2}} ║"
def sep(l="╠",c="═",r="╣"): return f"{l}{c*W}{r}"
def hdr(text):   return f"║ {'【 '+text+' 】':^{W-2}} ║"

def pct(v): return f"{v*100:.1f}%"

def print_ticket(home, away, mdl, s, odds, bankroll, injuries_h, injuries_a,
                 players_h, players_a, h2h_df,
                 home_rates=None, away_rates=None, referee_factor=1.0,
                 weather=None, standings=None, fixture_info=None,
                 opening_odds=None):

    res  = s.result();     ou   = s.ou();      bt   = s.btts()
    cs   = s.correct_score(); ht = s.halftime(); sht = s.second_half()
    hmg  = s.half_most_goals(); dc = s.dc();   dnb  = s.draw_no_bet()
    wtn  = s.win_to_nil(); htft = s.htft();    xg   = s.xg()
    # Use team-specific rates for corners and cards
    home_cor = home_rates.get("avg_corners") if home_rates else None
    away_cor = away_rates.get("avg_corners") if away_rates else None
    home_crd = home_rates.get("avg_cards")   if home_rates else None
    away_crd = away_rates.get("avg_cards")   if away_rates else None
    w_goal   = float((weather or {}).get("goal_factor",   1.0))
    w_corner = float((weather or {}).get("corner_factor", 1.0))
    cor  = s.corners(home_avg=home_cor, away_avg=away_cor,
                     weather_factor=w_corner)
    crd  = s.cards(home_avg=home_crd, away_avg=away_crd,
                   referee_factor=referee_factor)
    fg   = s.first_goal()
    eg   = s.exact_goals(); ms  = s.multiscores()
    sibh = s.score_in_both_halves()
    tfg  = s.time_of_first_goal()
    str_ = mdl.strengths()
    hf   = mdl.form(home); af = mdl.form(away)

    # AH probabilities for common lines
    ah_m05 = s.asian_handicap(-0.5)
    ah_p05 = s.asian_handicap(+0.5)
    ah_m15 = s.asian_handicap(-1.5)
    ah_p15 = s.asian_handicap(+1.5)

    # Full market probability map — covers ALL odds inputs
    mkt_probs = {
        # 1X2
        "home_win"  : res["home_win"],
        "draw"      : res["draw"],
        "away_win"  : res["away_win"],
        # Over/Under
        "over_2.5"  : ou["over_2.5"],
        "under_2.5" : ou["under_2.5"],
        "over_1.5"  : ou["over_1.5"],
        "over_3.5"  : ou["over_3.5"],
        # BTTS
        "btts_yes"  : bt["btts_yes"],
        "btts_no"   : bt["btts_no"],
        # Double Chance
        "dc_1x"     : dc["1X"],
        "dc_x2"     : dc["X2"],
        # Draw No Bet
        "dnb_home"  : dnb["dnb_home"],
        "dnb_away"  : dnb["dnb_away"],
        # Win to Nil
        "wtn_home"  : wtn["home_wtn"],
        "wtn_away"  : wtn["away_wtn"],
        # Asian Handicap
        "ah_-0.5"   : ah_m05["home"],
        "ah_+0.5"   : ah_p05["away"],
        "ah_-1.0"   : s.asian_handicap(-1.0)["home"],
        "ah_+1.0"   : s.asian_handicap(+1.0)["away"],
        "ah_-1.5"   : ah_m15["home"],
        "ah_+1.5"   : ah_p15["away"],
        "dc_12"     : dc["12"],
        "ht_hw"     : ht["ht_hw"],
        "ht_d"      : ht["ht_d"],
        "ht_aw"     : ht["ht_aw"],
        "ht_ov_0.5" : ht["ht_ou_05"],
        "ht_ov_1.5" : ht["ht_ou_15"],
        "sh_ov_0.5" : sht["sh_ou_05"],
        "sh_ov_1.5" : sht["sh_ou_15"],
        "htft_hh"   : htft.get("H/H", 0),
        "htft_hd"   : htft.get("H/D", 0),
        "htft_dh"   : htft.get("D/H", 0),
        "htft_dd"   : htft.get("D/D", 0),
        "htft_ah"   : htft.get("A/H", 0),
        "htft_aa"   : htft.get("A/A", 0),
        "under_3.5" : ou["under_3.5"],
        "over_4.5"  : ou.get("over_4.5", 0),
        # Corners
        "cor_ov_7.5" : cor.get("ov_7.5", 0),
        "cor_ov_8.5" : cor.get("ov_8.5", 0),
        "cor_ov_9.5" : cor.get("ov_9.5", 0),
        "cor_ov_10.5": cor.get("ov_10.5", 0),
        "cor_un_7.5" : cor.get("un_7.5", 0),
        "cor_un_8.5" : cor.get("un_8.5", 0),
        "cor_un_9.5" : cor.get("un_9.5", 0),
        "cor_un_10.5": cor.get("un_10.5", 0),
        # Cards
        "crd_ov_1.5" : crd.get("ov_1.5", 0),
        "crd_ov_2.5" : crd.get("ov_2.5", 0),
        "crd_ov_3.5" : crd.get("ov_3.5", 0),
        "crd_ov_4.5" : crd.get("ov_4.5", 0),
        "crd_un_1.5" : crd.get("un_1.5", 0),
        "crd_un_2.5" : crd.get("un_2.5", 0),
        "crd_un_3.5" : crd.get("un_3.5", 0),
        # Booking Points
        "bp_ov_20.5" : crd.get("bp_ov_20.5", 0),
        "bp_ov_30.5" : crd.get("bp_ov_30.5", 0),
        "bp_ov_40.5" : crd.get("bp_ov_40.5", 0),
        # Extended O/U lines
        "over_0.5"   : ou.get("over_0.5", 0),
        "under_0.5"  : ou.get("under_0.5", 0),
        "under_1.5"  : ou.get("under_1.5", 0),
        "over_4.5"   : ou.get("over_4.5", 0),
        "under_4.5"  : ou.get("under_4.5", 0),
        "over_5.5"   : ou.get("over_5.5", 0),
        "under_5.5"  : ou.get("under_5.5", 0),
        # Home/Away team to score
        "home_score_yes": float((s.sh > 0).mean()),
        "home_score_no" : float((s.sh == 0).mean()),
        "away_score_yes": float((s.sa > 0).mean()),
        "away_score_no" : float((s.sa == 0).mean()),
        # HT Under lines
        "ht_un_0.5"  : 1 - ht["ht_ou_05"],
        "ht_un_1.5"  : 1 - ht["ht_ou_15"],
        "ht_ov_2.5"  : float((np.random.poisson(s.mh*0.45) +
                              np.random.poisson(s.ma*0.45) > 2.5).mean()),
        "ht_un_2.5"  : float((np.random.poisson(s.mh*0.45) +
                              np.random.poisson(s.ma*0.45) <= 2.5).mean()),
        # 2H Under lines
        "sh_un_0.5"  : 1 - sht["sh_ou_05"],
        "sh_un_1.5"  : 1 - sht["sh_ou_15"],
        "sh_ov_2.5"  : float((np.random.poisson(s.mh*0.55) +
                              np.random.poisson(s.ma*0.55) > 2.5).mean()),
        "sh_un_2.5"  : float((np.random.poisson(s.mh*0.55) +
                              np.random.poisson(s.ma*0.55) <= 2.5).mean()),
        # Home/Away goal lines
        "home_ov_0.5": float((s.sh > 0).mean()),
        "home_ov_1.5": float((s.sh > 1).mean()),
        "home_ov_2.5": float((s.sh > 2).mean()),
        "home_un_0.5": float((s.sh < 1).mean()),
        "home_un_1.5": float((s.sh <= 1).mean()),
        "home_un_2.5": float((s.sh <= 2).mean()),
        "away_ov_0.5": float((s.sa > 0).mean()),
        "away_ov_1.5": float((s.sa > 1).mean()),
        "away_ov_2.5": float((s.sa > 2).mean()),
        "away_un_0.5": float((s.sa < 1).mean()),
        "away_un_1.5": float((s.sa <= 1).mean()),
        "away_un_2.5": float((s.sa <= 2).mean()),
        # FT Result + Goals combos
        "res_goals_h1_ov": float(((s.sh > s.sa) & (s.tt > 2.5)).mean()),
        "res_goals_d_ov" : float(((s.sh == s.sa) & (s.tt > 2.5)).mean()),
        "res_goals_a_ov" : float(((s.sh < s.sa) & (s.tt > 2.5)).mean()),
        "res_goals_h_un" : float(((s.sh > s.sa) & (s.tt <= 2.5)).mean()),
        "res_goals_d_un" : float(((s.sh == s.sa) & (s.tt <= 2.5)).mean()),
        "res_goals_a_un" : float(((s.sh < s.sa) & (s.tt <= 2.5)).mean()),
        # FT Result + BTTS combos
        "res_btts_h_y"   : float(((s.sh > s.sa) & (s.sh > 0) & (s.sa > 0)).mean()),
        "res_btts_d_y"   : float(((s.sh == s.sa) & (s.sh > 0)).mean()),
        "res_btts_a_y"   : float(((s.sh < s.sa) & (s.sh > 0) & (s.sa > 0)).mean()),
        # Goals + BTTS combos
        "goals_btts_ov_y": float(((s.tt > 2.5) & (s.sh > 0) & (s.sa > 0)).mean()),
        "goals_btts_ov_n": float(((s.tt > 2.5) & ~((s.sh > 0) & (s.sa > 0))).mean()),
        "goals_btts_un_y": float(((s.tt <= 2.5) & (s.sh > 0) & (s.sa > 0)).mean()),
        "goals_btts_un_n": float(((s.tt <= 2.5) & ~((s.sh > 0) & (s.sa > 0))).mean()),
        # Highest scoring half
        "high_scoring_half_1st": hmg["first_half"],
        "high_scoring_half_2nd": hmg["second_half"],
        # Corners extended
        "cor_ov_11.5": cor.get("ov_11.5", 0),
        "cor_un_11.5": cor.get("un_11.5", 0),
        # Red card (NegBin approximation: ~15% chance in any PL match)
        "red_card_yes": float(min(0.15 * (1 + crd.get("mean_t", 3.5) / 4.0), 0.35)),
        "red_card_no" : float(max(0.65, 1 - min(0.15 * (1 + crd.get("mean_t", 3.5) / 4.0), 0.35))),
        # Correct scores
        **{f"cs_{h}_{a}": float(((s.sh == h) & (s.sa == a)).mean())
           for h in range(6) for a in range(6)},
        "cs_5_5": float(((s.sh == 5) & (s.sa == 5)).mean()),
    }

    print(); print(f"╔{'═'*W}╗")
    league_tag = config.ACTIVE_LEAGUE.get("name","").replace("  "," ").strip()[:50]
    print(row("")); print(row("  ⚽  BAYESIAN FOOTBALL INTELLIGENCE  v10.0  ⚽"))
    print(row(f"  {league_tag}"))
    print(row("")); print(sep())
    print(row(f"  🏟   {home}  vs  {away}"))
    # Fixture info line
    if fixture_info and fixture_info.get("fixture_id"):
        fi     = fixture_info
        dt_str = fi.get("date", "")[:16].replace("T", " ") if fi.get("date") else "TBD"
        venue  = fi.get("venue", "Unknown venue")
        ref    = fi.get("referee", "") or "TBD"
        rnd    = fi.get("round", "")
        print(row(f"  📅   {dt_str}   {rnd}"))
        print(row(f"  🏟   {venue}   Referee: {ref}"))
    else:
        print(row(f"  📅   {datetime.now().strftime('%A, %d %B %Y  %H:%M')}"))
    print(row(f"  💰   Bankroll: {bankroll:,.0f} RWF   Sims: {config.N_SIM:,}"))
    # Weather line
    if weather and weather.get("description"):
        wf_g = weather.get("goal_factor", 1.0)
        wf_c = weather.get("corner_factor", 1.0)
        effect = ""
        if wf_g < 0.99:
            effect += f"  goals ×{wf_g:.2f}"
        if wf_c != 1.0:
            effect += f"  corners ×{wf_c:.2f}"
        print(row(f"  🌤   {weather['description']}{effect}"))
    # Standings line
    if standings and "home" in standings and "away" in standings:
        hs = standings["home"]; as_ = standings["away"]
        print(row(f"  📊  {hs['team'][:20]}: #{hs['pos']} {hs['pts']}pts  |  "
                  f"{as_['team'][:20]}: #{as_['pos']} {as_['pts']}pts"))
    print(sep())

    # Strengths
    print(hdr("TEAM STRENGTH  (posterior MAP)"))
    print(sep("╠","─","╣"))
    print(row(f"  {'Team':<26}  {'Attack':>8}  {'Defense':>8}  {'Net':>8}  {'xG Avg':>7}"))
    print(sep("╠","─","╣"))
    for team, f_ in [(home, hf), (away, af)]:
        atk, dfn = str_[team]
        net = atk-dfn
        print(row(f"  {team[:26]:<26}  {atk:>+8.3f}  {dfn:>+8.3f}  {net:>+8.3f}  {f_['avg_xg']:>7.2f}"))
    print(sep())

    # Form
    print(hdr("RECENT FORM  (last 5)"))
    print(sep("╠","─","╣"))
    for team, f_ in [(home, hf), (away, af)]:
        print(row(f"  {team[:26]:<26}  {f_['form']}  "
                  f"Pts:{f_['pts']}  GF:{f_['gf']}  GA:{f_['ga']}  "
                  f"GD:{f_['gd']:+d}  xG:{f_['avg_xg']:.2f}"))
    print(sep())

    # Adjustment factors
    print(hdr("MODEL ADJUSTMENTS  (v8.0)"))
    print(sep("╠","─","╣"))
    print(row(f"  H2H Factor: {mdl.h2h_factor:.3f}   "
              f"Momentum Home: ×{mdl.mom_h:.3f}   Away: ×{mdl.mom_a:.3f}"))
    print(row(f"  Injury — Home: ×{mdl.inj_h_factor:.3f}   Away: ×{mdl.inj_a_factor:.3f}   "
              f"Fatigue — Home: ×{mdl.fatigue_h:.3f}   Away: ×{mdl.fatigue_a:.3f}"))
    print(row(f"  Dixon-Coles ρ: {mdl.rho:+.3f}   "
              f"({'lowers' if mdl.rho < 0 else 'raises'} low-score probabilities)"))
    print(row(f"  Weather goal ×{mdl.weather_goal_factor:.3f}   "
              f"corner ×{mdl.weather_corner_factor:.3f}"))
    cor_src = "team data" if cor.get("using_team_data") else "PL averages"
    crd_src = "team data" if crd.get("using_team_data") else "PL averages"
    ref_str = f"×{referee_factor:.2f}" if referee_factor != 1.0 else "average"
    print(row(f"  Corners source: {cor_src}   Cards source: {crd_src}   "
              f"Referee: {ref_str}"))
    print(sep())

    # Injuries shown only as summary count — details hidden, factor applied silently
    n_inj_h = len([i for i in injuries_h if i.get("player_name")])
    n_inj_a = len([i for i in injuries_a if i.get("player_name")])
    if n_inj_h or n_inj_a:
        print(row(f"  🏥 Injuries: {home[:20]} {n_inj_h} absent  |  "
                  f"{away[:20]} {n_inj_a} absent  (applied to model silently)"))

    # xG
    print(hdr("EXPECTED GOALS  (xG)"))
    print(sep("╠","─","╣"))
    print(row(f"  {home[:22]:<22}  xG {xg['xg_h']:.2f}   {bar(xg['xg_h']/4)}"))
    print(row(f"  {away[:22]:<22}  xG {xg['xg_a']:.2f}   {bar(xg['xg_a']/4)}"))
    tone = "High-scoring" if xg["tot"]>3 else "Moderate" if xg["tot"]>2 else "Low-scoring"
    print(row(f"  Total xG: {xg['tot']:.2f}   [{tone}]"))
    print(sep())

    # H2H
    if not h2h_df.empty and len(h2h_df) >= 3:
        print(hdr("HEAD-TO-HEAD  (last meetings)"))
        print(sep("╠","─","╣"))
        for _, hr in h2h_df.head(6).iterrows():
            result_tag = ("H" if hr["home_goals"]>hr["away_goals"]
                          else "D" if hr["home_goals"]==hr["away_goals"] else "A")
            date_str = str(hr.get("match_date",""))[:10]
            print(row(f"  {date_str}  {hr['home_team'][:18]:<18} "
                      f"{hr['home_goals']}-{hr['away_goals']}  "
                      f"{hr['away_team'][:18]:<18}  [{result_tag}]"))
        print(sep())

    # 1X2
    print(hdr("MATCH RESULT  (1X2)"))
    print(sep("╠","─","╣"))
    for lbl, p in [(f"  1  Home  ({home[:18]})", res["home_win"]),
                   (f"  X  Draw               ", res["draw"]),
                   (f"  2  Away  ({away[:18]})", res["away_win"])]:
        arrow = "▶" if p==max(res.values()) else " "
        print(row(f"{arrow} {lbl[:42]:<42} {pct(p):>7}  {bar(p)}"))
    print(sep("╠","─","╣"))
    print(row(f"  Double Chance  1X:{pct(dc['1X'])}  X2:{pct(dc['X2'])}  12:{pct(dc['12'])}"))
    print(row(f"  Draw No Bet    Home:{pct(dnb['dnb_home'])}   Away:{pct(dnb['dnb_away'])}"))
    print(row(f"  Win To Nil     Home:{pct(wtn['home_wtn'])}   Away:{pct(wtn['away_wtn'])}"))
    print(sep())

    # Goals
    print(hdr("GOALS MARKETS"))
    print(sep("╠","─","╣"))
    for l in (0.5,1.5,2.5,3.5,4.5,5.5):
        ov,un = ou[f"over_{l}"],ou[f"under_{l}"]
        fo = "←" if ov>=un else " "; fu = "←" if un>ov else " "
        print(row(f"  O/U {l}   Over:{pct(ov)} {fo}   Under:{pct(un)} {fu}"))
    print(sep("╠","─","╣"))
    print(row(f"  BTTS Yes:{pct(bt['btts_yes'])}   BTTS No:{pct(bt['btts_no'])}"))
    print(sep("╠","─","╣"))
    print(row(f"  Exact Goals: " + "  ".join(f"{n}:{pct(eg[f'exactly_{n}'])}"
                                              for n in range(6))))
    print(row(f"  Exactly 6+: {pct(eg['exactly_7plus'])}"))
    print(sep())

    # Half analysis
    print(hdr("HALF-TIME / SECOND-HALF ANALYSIS"))
    print(sep("╠","─","╣"))
    print(row(f"  HT Result   Home:{pct(ht['ht_hw'])}  Draw:{pct(ht['ht_d'])}  Away:{pct(ht['ht_aw'])}"))
    print(row(f"  2H Result   Home:{pct(sht['sh_hw'])}  Draw:{pct(sht['sh_d'])}  Away:{pct(sht['sh_aw'])}"))
    print(row(f"  HT O/U 0.5: Over {pct(ht['ht_ou_05'])}   HT O/U 1.5: Over {pct(ht['ht_ou_15'])}"))
    print(row(f"  2H O/U 0.5: Over {pct(sht['sh_ou_05'])}   2H O/U 1.5: Over {pct(sht['sh_ou_15'])}"))
    print(row(f"  Half Most Goals: 1st {pct(hmg['first_half'])}  "
              f"2nd {pct(hmg['second_half'])}  Equal {pct(hmg['equal'])}"))
    print(sep())

    # Score in both halves
    print(hdr("SCORE IN BOTH HALVES"))
    print(sep("╠","─","╣"))
    print(row(f"  {home[:28]:<28} scores in both halves: {pct(sibh['home_score_both_halves'])}"))
    print(row(f"  {away[:28]:<28} scores in both halves: {pct(sibh['away_score_both_halves'])}"))
    print(sep())

    # 3-way handicap
    print(hdr("3-WAY HANDICAP  (home team perspective)"))
    print(sep("╠","─","╣"))
    print(row(f"  {'HCP':<8}  {'Half':<5}  {'Home':>7}  {'Draw':>7}  {'Away':>7}"))
    print(sep("╠","─","╣"))
    for hcp in (-2,-1,0,1,2):
        for half in ("FT","HT","2H"):
            d = s.three_way_handicap(hcp, half)
            print(row(f"  {hcp:>+3}      {half:<5}  {pct(d['home']):>7}  "
                      f"{pct(d['draw']):>7}  {pct(d['away']):>7}"))
    print(sep())

    # Asian handicap
    print(hdr("ASIAN HANDICAP  (home team perspective)"))
    print(sep("╠","─","╣"))
    for hcp in (-1.5,-1.0,-0.75,-0.5,-0.25,0.0,0.25,0.5,1.0,1.5):
        d = s.asian_handicap(hcp)
        push = f"  Push:{pct(d['push'])}" if d["push"]>0.01 else ""
        print(row(f"  AH {hcp:+.2f}   Home:{pct(d['home'])}   Away:{pct(d['away'])}{push}"))
    print(sep())

    # HT/FT
    print(hdr("HT / FT COMBINED"))
    print(sep("╠","─","╣"))
    htft_sorted = sorted(htft.items(), key=lambda x: x[1], reverse=True)
    for label, prob in htft_sorted:
        arrow = "▶" if label==htft_sorted[0][0] else " "
        print(row(f"{arrow} {label:<8}  {pct(prob):>7}  {bar(prob, 20)}"))
    print(sep())

    # Correct Score
    print(hdr("CORRECT SCORE  (Top 15)"))
    print(sep("╠","─","╣"))
    for _, r2 in cs[cs["score"]!="other"].head(15).iterrows():
        h_s,a_s = r2["score"].split("-")
        tag = "H" if int(h_s)>int(a_s) else "D" if int(h_s)==int(a_s) else "A"
        print(row(f"  [{tag}]  {r2['score']:>5}   {pct(r2['prob']):>7}   {bar(r2['prob']*5, 20)}"))
    oth = cs[cs["score"]=="other"]["prob"].values
    if len(oth): print(row(f"  [other combined]: {pct(oth[0])}"))
    print(sep())

    # Multiscores
    print(hdr("MULTISCORES  (bundled)"))
    print(sep("╠","─","╣"))
    for label, prob in sorted(ms.items(), key=lambda x: x[1], reverse=True):
        print(row(f"  {label:<40}  {pct(prob):>7}"))
    print(sep())

    # First goal
    print(hdr("FIRST GOAL & TIMING"))
    print(sep("╠","─","╣"))
    print(row(f"  {home[:30]:<30} scores first:  {pct(fg['home'])}"))
    print(row(f"  {away[:30]:<30} scores first:  {pct(fg['away'])}"))
    print(row(f"  No goals (0-0):                                {pct(fg['no_goal'])}"))
    print(sep("╠","─","╣"))
    print(row(f"  First goal time:  0-30 min:{pct(tfg['0_30'])}  "
              f"31-60:{pct(tfg['31_60'])}  61-90:{pct(tfg['61_90'])}"))
    print(sep())

    # Anytime goalscorer
    if players_h or players_a:
        print(hdr("ANYTIME GOALSCORER"))
        print(sep("╠","─","╣"))
        gs = s.anytime_goalscorer(players_h, players_a)
        print(row(f"  {home[:26]:<26}  {'Pos':<4}  {'Prob':>7}  {'xG/90':>7}  Src"))
        print(sep("╠","─","╣"))
        for p in gs["home"][:8]:
            xg_str = f"{p['xg90']:.3f}" if p.get("xg90") else "  —  "
            src    = "★" if p.get("source") == "actual" else "~"
            print(row(f"  {p['name'][:26]:<26}  {p['position']:<4}  "
                      f"{pct(p['prob']):>7}  {xg_str:>7}  {src}"))
        print(sep("╠","─","╣"))
        print(row(f"  {away[:26]:<26}  {'Pos':<4}  {'Prob':>7}  {'xG/90':>7}  Src"))
        print(sep("╠","─","╣"))
        for p in gs["away"][:8]:
            xg_str = f"{p['xg90']:.3f}" if p.get("xg90") else "  —  "
            src    = "★" if p.get("source") == "actual" else "~"
            print(row(f"  {p['name'][:26]:<26}  {p['position']:<4}  "
                      f"{pct(p['prob']):>7}  {xg_str:>7}  {src}"))
        print(sep("╠","─","╣"))
        print(row("  ★ = actual player data   ~ = position-weight estimate"))
        print(sep())

    # Corners
    print(hdr("CORNERS  (Negative Binomial)"))
    print(sep("╠","─","╣"))
    print(row(f"  Expected → Home:{cor['mean_h']:.1f}  Away:{cor['mean_a']:.1f}  "
              f"Total:{cor['mean_t']:.1f}  HT:{cor['mean_ht_corners']:.1f}"))
    print(sep("╠","─","╣"))
    for l in (7.5,8.5,9.5,10.5,11.5,12.5):
        ov,un = cor[f"ov_{l}"],cor[f"un_{l}"]
        fo = "←" if ov>=un else " "
        print(row(f"  Corners O/U {l}   Over:{pct(ov)} {fo}   Under:{pct(un)}"))
    print(sep())

    # Cards + Bookings
    print(hdr("CARDS & BOOKING POINTS  (Negative Binomial)"))
    print(sep("╠","─","╣"))
    print(row(f"  Expected → Home:{crd['mean_h']:.1f}  Away:{crd['mean_a']:.1f}  "
              f"Total:{crd['mean_t']:.1f}  Booking Pts:{crd['mean_booking_pts']:.0f}"))
    print(sep("╠","─","╣"))
    for l in (1.5,2.5,3.5,4.5,5.5):
        ov,un = crd[f"ov_{l}"],crd[f"un_{l}"]
        fo = "←" if ov>=un else " "
        print(row(f"  Cards O/U {l}   Over:{pct(ov)} {fo}   Under:{pct(un)}"))
    print(sep("╠","─","╣"))
    for l in (20.5,30.5,40.5,50.5):
        print(row(f"  Booking Points Over {l}:  {pct(crd[f'bp_ov_{l}'])}"))
    print(sep())

    # Value analysis
    print(hdr("💎  VALUE ANALYSIS  |  EV & QUARTER-KELLY (RWF)"))
    print(sep("╠","─","╣"))

    value_bets = []
    opening_odds = opening_odds or {}
    for mkt, mp in mkt_probs.items():
        if mkt in odds:
            e_val    = edge(mp, odds[mkt])
            ev_val   = ev(mp, odds[mkt])
            quality  = bet_quality(mkt, mp, e_val, ev_val=ev_val)

            # Line movement: opening vs current odds
            op_odds  = opening_odds.get(mkt)
            if op_odds and op_odds > 1.0:
                line_move     = odds[mkt] - op_odds        # +ve = drifted (market moved away)
                line_move_pct = (odds[mkt] / op_odds - 1)  # percentage change
            else:
                line_move = line_move_pct = None

            if e_val > 0:
                value_bets.append({
                    "market"    : mkt,
                    "model_p"   : mp,
                    "implied"   : implied(odds[mkt]),
                    "edge"      : e_val,
                    "ev"        : ev_val,
                    "odds"      : odds[mkt],
                    "op_odds"   : op_odds,
                    "line_move" : line_move,
                    "line_pct"  : line_move_pct,
                    "stake"     : kelly_stake(mp, odds[mkt], bankroll)
                                 if should_bet(quality) else 0.0,
                    "quality"   : quality,
                })

    # Sort: bettable bets first (by EV), then observe-only (by edge).
    # EV is the correct criterion for long-term profitability.
    # Edge alone ignores the odds level — +4% at 1.50 is far worse than +4% at 4.00.
    value_bets.sort(
        key=lambda x: (should_bet(x["quality"]), x["ev"]),
        reverse=True
    )

    best_market = ""; best_edge = 0.0; best_ev = 0.0; best_stake = 0.0; best_odds = 0.0

    if not odds:
        print(row("  No odds entered — EV analysis skipped."))
    elif not value_bets:
        print(row("  ⚠  No positive edge found. Recommendation: SKIP."))
    else:
        print(row(f"  {'Market':<13} {'Model%':>7} {'Implied%':>9} "
                  f"{'Edge':>7} {'EV@1u':>8} {'Stake RWF':>12}  Quality"))
        print(sep("╠","─","╣"))

        bettable = [b for b in value_bets if should_bet(b["quality"])]
        observe  = [b for b in value_bets if not should_bet(b["quality"])
                    and b["quality"] not in ("never", "ev_cap", "observe_league")]
        never    = [b for b in value_bets if b["quality"] in
                    ("never", "ev_cap", "observe_league")]

        for a in bettable:
            icon = quality_icon(a["quality"])
            # Line movement indicator
            lm = a.get("line_pct")
            if lm is not None:
                if lm > 0.03:
                    lm_tag = f"  📈+{lm:.1%}"   # drifted — market moved toward you
                elif lm < -0.03:
                    lm_tag = f"  📉{lm:.1%}"    # shortened — sharp money against you
                else:
                    lm_tag = f"  ↔ flat"
            else:
                lm_tag = ""
            print(row(f"  {icon} {a['market']:<12} {pct(a['model_p']):>7} "
                      f"{pct(a['implied']):>9} {a['edge']:>+7.1%} "
                      f"{a['ev']:>+8.3f} {a['stake']:>12,.0f}  "
                      f"{'PREMIUM' if a['quality']=='premium' else 'BET'}{lm_tag}"))

        if observe:
            print(sep("╠","─","╣"))
            print(row("  ── OBSERVE ONLY (fails probability or edge floor) ──────"))
            for a in observe:
                min_p = config.MIN_PROB_BY_MARKET.get(a["market"], config.MIN_PROB_DEFAULT)
                reason = (
                    f"prob {pct(a['model_p'])} < floor {pct(min_p)}"
                    if a["quality"] == "low_prob"
                    else f"edge {a['edge']:+.1%} < {pct(config.MIN_EDGE)} min"
                    if a["quality"] == "low_edge"
                    else f"fails both conditions"
                )
                print(row(f"  ⚠️  {a['market']:<12} {pct(a['model_p']):>7} "
                          f"{pct(a['implied']):>9} {a['edge']:>+7.1%}  "
                          f"[{reason}]"))

        if never:
            print(sep("╠","─","╣"))
            print(row("  ── DISABLED / CAPPED MARKETS ────────────────────────────"))
            for a in never:
                q = a["quality"]
                if q == "never":
                    reason = "banned — unreliable pre-calibration"
                elif q == "ev_cap":
                    reason = f"EV {a['ev']:.3f} > cap {config.MAX_EV_THRESHOLD} — model overconfidence"
                elif q == "observe_league":
                    reason = "league observe-only (negative CLV history)"
                else:
                    reason = q
                print(row(f"  {quality_icon(q)} {a['market']:<14} {pct(a['model_p']):>7} "
                          f"{pct(a['implied']):>9} {a['edge']:>+7.1%}  [{reason}]"))

        if config.AFL_LEAGUE_ID in config.OBSERVE_ONLY_LEAGUES:
            print(sep("╠","═","╣"))
            lg_name = config.ACTIVE_LEAGUE.get("name", "This league")
            print(row(f"  👁️  {lg_name} is in OBSERVE-ONLY mode"))
            print(row(f"     Negative CLV history. All markets shown but no bets placed."))
            print(row(f"     Remove league from OBSERVE_ONLY_LEAGUES in config.py"))
            print(row(f"     once it shows positive avg CLV over 30+ tracked bets."))

        if bettable:
            best = bettable[0]
            best_market = best["market"]; best_edge = best["edge"]
            best_ev = best["ev"]; best_stake = best["stake"]
            best_odds = best["odds"]

            print(sep("╠","═","╣"))
            print(hdr("🎯  RECOMMENDED BET"))
            print(sep("╠","─","╣"))
            pct_of_br    = best["stake"] / bankroll * 100 if bankroll else 0
            min_p        = config.MIN_PROB_BY_MARKET.get(best["market"], config.MIN_PROB_DEFAULT)
            min_e        = config.UNDER_MARKET_MIN_EDGE if best["market"] in config.UNDER_MARKETS else config.MIN_EDGE
            under_tag    = "  ⚠ Under market — raised edge floor applied" if best["market"] in config.UNDER_MARKETS else ""
            print(row(f"  Market  : {best['market']}{under_tag}"))
            print(row(f"  Odds    : {best['odds']:.2f}"))
            # Show opening line and line movement when available
            if best.get("op_odds"):
                lm = best["line_pct"]
                if lm is not None and abs(lm) > 0.01:
                    direction = "📈 drifted (market moved your way)" if lm > 0 else "📉 shortened (smart money against)"
                    print(row(f"  Opening : {best['op_odds']:.2f}  →  {best['odds']:.2f}  "
                              f"({lm:+.1%})  {direction}"))
                    # Pre-bet CLV: how much better is current vs opening?
                    pre_clv = (1/best["op_odds"] - 1/best["odds"]) * 100
                    clv_label = "✅ positive pre-bet CLV" if pre_clv > 0 else "⚠ negative pre-bet CLV"
                    print(row(f"  Pre-bet CLV: {pre_clv:+.2f}pp  ({clv_label})"))
                else:
                    print(row(f"  Opening : {best['op_odds']:.2f}  (line unchanged)"))
            print(row(f"  Model % : {pct(best['model_p'])}  ≥  floor {pct(min_p)}  ✓"))
            print(row(f"  Edge    : {best['edge']:+.1%}  ≥  min {pct(min_e)}  ✓"))
            print(row(f"  EV per 1 RWF: {best['ev']:+.4f}"))
            print(row(f"  ▶ STAKE : {best['stake']:,.0f} RWF  ({pct_of_br:.2f}% of bankroll)"))
            print(row(f"  ▶ RETURN: {best['stake'] * best['odds']:,.0f} RWF  if win"))
            if best["quality"] == "premium":
                print(row(f"  🔥 PREMIUM — model prob ≥ {pct(config.PREMIUM_PROB)} "
                          f"AND edge ≥ {pct(config.PREMIUM_EDGE)}  (both conditions met)"))
        else:
            print(sep("╠","═","╣"))
            print(hdr("⚠️  NO BETTABLE EDGE FOUND"))
            print(sep("╠","─","╣"))
            print(row("  All positive edges fail the probability floor."))
            print(row("  These are high-edge longshots — variance too high to bet."))
            print(row("  Recommendation: SKIP this game or observe only."))

    print(sep("╠","═","╣"))
    print(hdr("🛡  BANKROLL RULES  (Dual-Condition Filter)"))
    print(sep("╠","─","╣"))
    print(row(f"  • BET only when BOTH conditions are met simultaneously:"))
    print(row(f"      Edge  ≥ {pct(config.MIN_EDGE)} above implied  AND  Model prob ≥ market floor"))
    print(row(f"  • Under markets (U0.5–U6.5): raised edge floor ≥ {pct(config.UNDER_MARKET_MIN_EDGE)}"))
    print(row(f"      (NB model may underestimate variance; Under lines are bookmaker-sharp)"))
    print(row(f"  • PREMIUM (🔥): edge ≥ {pct(config.PREMIUM_EDGE)} AND model prob ≥ {pct(config.PREMIUM_PROB)}"))
    print(row(f"  • Best market = highest EV (not edge). EV accounts for odds level."))
    print(row(f"  • ⚠️ entries show edge but fail probability or edge floor — OBSERVE ONLY."))
    print(row(f"  • Quarter-Kelly (×0.25). Hard cap: {config.MAX_KELLY_PCT*100:.0f}% of bankroll per bet."))
    print(row(f"  • Drawdown 30% from peak → halve all stakes immediately."))
    print(sep())
    print(row("  ⚠  Statistical model. Gamble responsibly. 18+."))
    print(f"╚{'═'*W}╝\n")

    return {
        "best_market": best_market, "best_edge": best_edge,
        "best_ev": best_ev, "best_stake": best_stake,
        "best_odds": best_odds,
    }


# ══════════════════════════════════════════════════════════════════════════════
# UPCOMING FIXTURES DISPLAY
# ══════════════════════════════════════════════════════════════════════════════
def print_upcoming(fixtures: list):
    if not fixtures:
        return
    print(f"\n{'─'*74}")
    print(f"  📅  UPCOMING PREMIER LEAGUE FIXTURES")
    print(f"{'─'*74}")
    for f in fixtures:
        print(f"  {f['date']} {f['time']}   "
              f"{f['home']:<25} vs  {f['away']:<25}  {f['matchday']}")
    print(f"{'─'*74}\n")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# LEAGUE SELECTOR
# ══════════════════════════════════════════════════════════════════════════════
