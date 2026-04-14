#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║        BAYESIAN FOOTBALL PREDICTION ENGINE  v8.0                           ║
║        Auto Odds · Auto Lineups · Weather · Fixture Lookup · Dixon-Coles   ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import os, sys, json, time, warnings, getpass
from datetime import datetime, timedelta
from difflib import get_close_matches
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from scipy.optimize import minimize
from scipy.stats import norm
from dotenv import load_dotenv

import mysql.connector
from mysql.connector import Error as MySQLError

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

warnings.filterwarnings("ignore")
np.random.seed(42)

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
ENV_FILE     = SCRIPT_DIR / ".env"
TRACKER_FILE = SCRIPT_DIR / "predictions_tracker.xlsx"

# ── Model constants ────────────────────────────────────────────────────────────
N_POSTERIOR     = 2_000
N_SIM           = 10_000
DECAY_HALF_LIFE = 60
RECENT_DAYS     = 35
QUARTER_KELLY   = 0.25
MAX_KELLY_PCT   = 0.05

# ── Bet quality thresholds — BOTH conditions must be true simultaneously ───────
# MIN_EDGE: minimum gap between model probability and bookmaker implied probability
# MIN_PROB: minimum model probability — filters out high-edge longshots where
#           variance is too large to confirm the edge before ruin
#
# A bet is only worth placing when BOTH are satisfied:
#   edge >= MIN_EDGE  AND  model_prob >= MIN_PROB_BY_MARKET (or MIN_PROB_DEFAULT)
#
# Premium (🔥) requires BOTH:
#   edge >= PREMIUM_EDGE  AND  model_prob >= PREMIUM_PROB
#
# Acceptable (✅) requires BOTH:
#   edge >= MIN_EDGE  AND  model_prob >= MIN_PROB_DEFAULT
#
# Warning (⚠️): positive edge but fails one or both conditions — observe only

MIN_EDGE         = 0.04     # minimum edge to consider betting (raised from 3% to 4%)
MIN_PROB_DEFAULT = 0.40     # minimum model probability for any bet
PREMIUM_EDGE     = 0.07     # edge threshold for premium classification
PREMIUM_PROB     = 0.50     # model must favour the outcome to call it premium

# Per-market minimum probability overrides
# Markets where lower confidence is acceptable (inherently less predictable)
MIN_PROB_BY_MARKET = {
    "home_win"  : 0.35,
    "draw"      : 0.25,   # draws are hard — lower floor but needs higher edge
    "away_win"  : 0.30,
    "over_2.5"  : 0.45,
    "under_2.5" : 0.45,
    "over_1.5"  : 0.45,
    "over_3.5"  : 0.40,
    "btts_yes"  : 0.40,
    "btts_no"   : 0.40,
    "dc_1x"     : 0.55,
    "dc_x2"     : 0.55,
    "dnb_home"  : 0.40,
    "dnb_away"  : 0.35,
    "wtn_home"  : 0.30,
    "wtn_away"  : 0.25,
    "ah_-0.5"   : 0.45,
    "ah_+0.5"   : 0.45,
    "ah_-1.5"   : 0.40,
    "ah_+1.5"   : 0.40,
    # Corners and cards — model confidence must be reasonable
    "cor_ov_7.5" : 0.45, "cor_ov_8.5" : 0.45,
    "cor_ov_9.5" : 0.45, "cor_ov_10.5": 0.45,
    "cor_un_7.5" : 0.45, "cor_un_8.5" : 0.45,
    "cor_un_9.5" : 0.45, "cor_un_10.5": 0.45,
    "crd_ov_1.5" : 0.45, "crd_ov_2.5" : 0.45,
    "crd_ov_3.5" : 0.45, "crd_ov_4.5" : 0.40,
    "crd_un_1.5" : 0.45, "crd_un_2.5" : 0.45,
    "crd_un_3.5" : 0.45,
    "bp_ov_20.5" : 0.45, "bp_ov_30.5" : 0.45,
    "bp_ov_40.5" : 0.40,
}

# API-Football — sole data source for all match data
# Free plan explicitly covers seasons 2022, 2023, 2024 only.
# Season 2025 (2025/26 current) requires a paid plan.
# 2022 = 2022/23  |  2023 = 2023/24  |  2024 = 2024/25 (most recent complete)
AFL_BASE        = "https://v3.football.api-sports.io"
AFL_LEAGUE_ID   = 39
AFL_SEASONS     = [2022, 2023, 2024]   # confirmed free plan range
AFL_MAX_DAILY   = 100

# football-data.org is no longer used as primary source.
# Kept only as a fallback if API-Football fixture fetch fails.
FD_BASE         = "https://api.football-data.org/v4"
FD_COMP         = "PL"
FD_SEASONS      = [2023, 2024]

# PL averages for non-goal markets
PL_CORNERS_H    = 5.4
PL_CORNERS_A    = 4.7
PL_CARDS_H      = 1.75
PL_CARDS_A      = 2.10

# Weather API (Open-Meteo — free, no key required)
WEATHER_BASE    = "https://api.open-meteo.com/v1/forecast"

# Premier League venue coordinates for weather lookup
# (latitude, longitude)
VENUE_COORDS = {
    "Emirates Stadium"             : (51.5550, -0.1084),
    "Anfield"                      : (53.4308, -2.9608),
    "Old Trafford"                 : (53.4631, -2.2913),
    "Etihad Stadium"               : (53.4831, -2.2004),
    "Stamford Bridge"              : (51.4816, -0.1910),
    "Tottenham Hotspur Stadium"    : (51.6042, -0.0666),
    "St. James' Park"              : (54.9754, -1.6218),
    "Villa Park"                   : (52.5092, -1.8847),
    "Goodison Park"                : (53.4388, -2.9666),
    "Amex Stadium"                 : (50.8609, -0.0832),
    "Molineux Stadium"             : (52.5902, -2.1302),
    "London Stadium"               : (51.5386, -0.0162),
    "Selhurst Park"                : (51.3983, -0.0855),
    "Brentford Community Stadium"  : (51.4882, -0.2866),
    "Gtech Community Stadium"      : (51.4882, -0.2866),
    "Craven Cottage"               : (51.4749, -0.2218),
    "City Ground"                  : (52.9399, -1.1323),
    "Vitality Stadium"             : (50.7352, -1.8382),
    "Bramall Lane"                 : (53.3703, -1.4706),
    "Kenilworth Road"              : (51.8839, -0.4317),
    "Turf Moor"                    : (53.7889, -2.2302),
    "King Power Stadium"           : (52.6204, -1.1424),
    "Portman Road"                 : (52.0545, 1.1446),
    "St Mary's Stadium"            : (50.9058, -1.3914),
    "Stadium of Light"             : (54.9147, -1.3883),
}

# Bet365 bookmaker ID in API-Football
BET365_ID = 8

ALIASES = {
    "man city": "Manchester City", "man utd": "Manchester United",
    "man united": "Manchester United", "arsenal": "Arsenal",
    "chelsea": "Chelsea", "liverpool": "Liverpool",
    "tottenham": "Tottenham", "spurs": "Tottenham",
    "newcastle": "Newcastle United", "villa": "Aston Villa",
    "aston villa": "Aston Villa", "wolves": "Wolverhampton Wanderers",
    "west ham": "West Ham United", "brighton": "Brighton",
    "everton": "Everton", "brentford": "Brentford",
    "fulham": "Fulham", "crystal palace": "Crystal Palace",
    "bournemouth": "Bournemouth", "nottm forest": "Nottingham Forest",
    "forest": "Nottingham Forest", "leicester": "Leicester City",
    "ipswich": "Ipswich", "southampton": "Southampton",
    "luton": "Luton", "burnley": "Burnley",
    "sheffield utd": "Sheffield Utd",
}

# ══════════════════════════════════════════════════════════════════════════════
# ENVIRONMENT SETUP  —  first-run wizard
# ══════════════════════════════════════════════════════════════════════════════
def setup_env():
    """Ask for credentials on first run, save to .env file."""
    load_dotenv(ENV_FILE)
    changed = False

    keys = {
        "AFL_API_KEY"   : ("API-Football key (from dashboard.api-football.com)", False),
        "FD_API_KEY"    : ("football-data.org API key (free)", False),
        "MYSQL_HOST"    : ("MySQL host", False),
        "MYSQL_PORT"    : ("MySQL port", False),
        "MYSQL_USER"    : ("MySQL username", False),
        "MYSQL_PASSWORD": ("MySQL password", True),
        "MYSQL_DB"      : ("MySQL database name", False),
        "BANKROLL_RWF"  : ("Starting bankroll in RWF", False),
    }

    defaults = {
        "MYSQL_HOST": "localhost",
        "MYSQL_PORT": "3306",
        "MYSQL_USER": "root",
        "MYSQL_DB"  : "football_predictor",
        "BANKROLL_RWF": "20000",
    }

    env_lines = {}
    if ENV_FILE.exists():
        with open(ENV_FILE) as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    env_lines[k.strip()] = v.strip()

    for key, (label, is_secret) in keys.items():
        current = env_lines.get(key) or os.getenv(key) or defaults.get(key, "")
        if not current:
            print(f"\n  [{label}]")
            if is_secret:
                val = getpass.getpass(f"  Enter {label}: ").strip()
            else:
                val = input(f"  Enter {label}: ").strip()
            env_lines[key] = val
            changed = True

    if changed:
        with open(ENV_FILE, "w") as f:
            for k, v in env_lines.items():
                f.write(f"{k}={v}\n")
        load_dotenv(ENV_FILE, override=True)
        print("\n  [Setup] Credentials saved to .env\n")


def get_env(key, default=""):
    load_dotenv(ENV_FILE)
    return os.getenv(key, default)


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE LAYER
# ══════════════════════════════════════════════════════════════════════════════
class DB:
    """MySQL connection wrapper with auto-reconnect."""

    def __init__(self):
        self.conn = None
        self.connect()

    def connect(self):
        try:
            self.conn = mysql.connector.connect(
                host     = get_env("MYSQL_HOST", "localhost"),
                port     = int(get_env("MYSQL_PORT", "3306")),
                user     = get_env("MYSQL_USER", "root"),
                password = get_env("MYSQL_PASSWORD", ""),
                database = get_env("MYSQL_DB", "football_predictor"),
                charset  = "utf8mb4",
                autocommit = True,
            )
        except MySQLError as e:
            print(f"\n  [DB ERROR] Cannot connect to MySQL: {e}")
            print("  Make sure MySQL is running and credentials in .env are correct.")
            sys.exit(1)

    def execute(self, sql, params=None, fetch=False):
        try:
            if not self.conn.is_connected():
                self.connect()
            cur = self.conn.cursor(dictionary=True)
            cur.execute(sql, params or ())
            if fetch:
                result = cur.fetchall()
                cur.close()
                return result
            self.conn.commit()
            last_id = cur.lastrowid
            cur.close()
            return last_id
        except MySQLError as e:
            print(f"  [DB] Query error: {e}")
            return [] if fetch else None

    def executemany(self, sql, data):
        try:
            if not self.conn.is_connected():
                self.connect()
            cur = self.conn.cursor()
            cur.executemany(sql, data)
            self.conn.commit()
            cur.close()
        except MySQLError as e:
            print(f"  [DB] Batch error: {e}")

    def fetchall(self, sql, params=None):
        return self.execute(sql, params, fetch=True)

    def fetchone(self, sql, params=None):
        rows = self.fetchall(sql, params)
        return rows[0] if rows else None

    def api_requests_today(self, api="afl"):
        row = self.fetchone(
            "SELECT COUNT(*) as n FROM api_request_log "
            "WHERE api=%s AND DATE(requested_at)=CURDATE()", (api,)
        )
        return row["n"] if row else 0

    def log_api_request(self, api, endpoint):
        self.execute(
            "INSERT INTO api_request_log (api, endpoint) VALUES (%s,%s)",
            (api, endpoint)
        )

    def close(self):
        if self.conn and self.conn.is_connected():
            self.conn.close()


# ══════════════════════════════════════════════════════════════════════════════
# API CLIENTS
# ══════════════════════════════════════════════════════════════════════════════
class APIFootball:
    """Client for api-football.com with daily budget tracking."""

    def __init__(self, db: DB):
        self.db  = db
        self.key = get_env("AFL_API_KEY", "")
        self.session = requests.Session()
        self.session.headers.update({
            "x-apisports-key": self.key,
            "x-rapidapi-key" : self.key,
        })

    def _remaining(self):
        used = self.db.api_requests_today("afl")
        return AFL_MAX_DAILY - used

    def _get(self, endpoint, params=None):
        remaining = self._remaining()
        if remaining <= 0:
            print(f"  [API-Football] Daily limit reached (100/day). Try tomorrow.")
            return None
        url = f"{AFL_BASE}/{endpoint}"
        try:
            r = self.session.get(url, params=params, timeout=30)
            r.raise_for_status()
            self.db.log_api_request("afl", endpoint)
            data = r.json()
            if data.get("errors"):
                print(f"  [API-Football] Error: {data['errors']}")
                return None
            return data.get("response", [])
        except requests.RequestException as e:
            print(f"  [API-Football] Request failed: {e}")
            return None

    def fetch_fixtures(self, season):
        print(f"    Fetching fixtures season {season} …", end=" ", flush=True)
        data = self._get("fixtures", {"league": AFL_LEAGUE_ID,
                                      "season": season, "status": "FT"})
        if data is None:
            print("failed")
            return []
        print(f"{len(data)} matches")
        return data

    def fetch_match_stats(self, fixture_id):
        return self._get("fixtures/statistics", {"fixture": fixture_id})

    def fetch_lineups(self, fixture_id):
        return self._get("fixtures/lineups", {"fixture": fixture_id})

    def fetch_events(self, fixture_id):
        return self._get("fixtures/events", {"fixture": fixture_id})

    def fetch_injuries(self, team_id, season=2024):
        return self._get("injuries", {"league": AFL_LEAGUE_ID,
                                      "season": season, "team": team_id})

    def fetch_h2h(self, team_a_id, team_b_id, last=20):
        return self._get("fixtures/headtohead",
                         {"h2h": f"{team_a_id}-{team_b_id}", "last": last})

    def fetch_next_fixtures(self, next_n=10):
        """
        Fetch upcoming fixtures using date range (free plan compatible).
        The 'next' parameter requires a paid plan — use 'from/to' instead.
        """
        today = datetime.now().strftime("%Y-%m-%d")
        future = (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d")
        return self._get("fixtures", {
            "league" : AFL_LEAGUE_ID,
            "season" : 2024,
            "from"   : today,
            "to"     : future,
        })

    def fetch_teams(self, season=2024):
        return self._get("teams", {"league": AFL_LEAGUE_ID, "season": season})

    def find_upcoming_fixture(self, home_team_id: int, away_team_id: int,
                              season: int = 2024) -> dict:
        """
        Find the next scheduled fixture between two teams.
        Returns fixture dict with id, date, venue, referee.
        Searches upcoming 60 days to cover full fixture schedule.
        """
        today  = datetime.now().strftime("%Y-%m-%d")
        future = (datetime.now() + timedelta(days=60)).strftime("%Y-%m-%d")
        data   = self._get("fixtures", {
            "league" : AFL_LEAGUE_ID,
            "season" : season,
            "team"   : home_team_id,
            "from"   : today,
            "to"     : future,
        })
        if not data:
            return {}
        # Find match against the away team
        for m in data:
            teams = m.get("teams", {})
            h_id  = teams.get("home", {}).get("id")
            a_id  = teams.get("away", {}).get("id")
            if h_id == home_team_id and a_id == away_team_id:
                fix = m.get("fixture", {})
                return {
                    "fixture_id" : fix.get("id"),
                    "date"       : fix.get("date", ""),
                    "venue"      : fix.get("venue", {}).get("name", ""),
                    "city"       : fix.get("venue", {}).get("city", ""),
                    "referee"    : fix.get("referee", "") or "",
                    "status"     : fix.get("status", {}).get("short", ""),
                    "round"      : m.get("league", {}).get("round", ""),
                }
        return {}

    def fetch_prematch_odds(self, fixture_id: int) -> dict:
        """
        Fetch pre-match odds from Bet365 (bookmaker_id=8) for a fixture.
        Returns raw bets list from API response.
        """
        data = self._get("odds", {
            "fixture"    : fixture_id,
            "bookmaker"  : BET365_ID,
        })
        if not data:
            return {}
        # Response is list of bookmaker objects
        for bm in data:
            for bookmaker in bm.get("bookmakers", []):
                if bookmaker.get("id") == BET365_ID:
                    return {b["name"]: b["values"]
                            for b in bookmaker.get("bets", [])}
        return {}

    def fetch_live_lineups(self, fixture_id: int) -> list:
        """
        Fetch confirmed lineups for a fixture.
        Returns empty list if lineups not yet released.
        """
        return self._get("fixtures/lineups", {"fixture": fixture_id}) or []

    def fetch_player_stats(self, team_id: int, season: int = 2024) -> list:
        """
        Fetch player statistics for a team/season.
        Returns list of player dicts with goals, shots, xG per game.
        Uses league filter to get PL-specific stats.
        """
        data = self._get("players", {
            "team"   : team_id,
            "season" : season,
            "league" : AFL_LEAGUE_ID,
        })
        return data or []

    def fetch_standings(self, season: int = 2024) -> list:
        """Fetch current PL standings table."""
        data = self._get("standings", {
            "league" : AFL_LEAGUE_ID,
            "season" : season,
        })
        if not data:
            return []
        try:
            return data[0]["league"]["standings"][0]
        except (IndexError, KeyError):
            return []


class FootballDataOrg:
    """Fallback client for football-data.org (goals only)."""

    def __init__(self, db: DB):
        self.db  = db
        self.key = get_env("FD_API_KEY", "")
        self.session = requests.Session()
        self.session.headers.update({"X-Auth-Token": self.key})

    def _get(self, endpoint, params=None, retries=3):
        url = f"{FD_BASE}/{endpoint}"
        for attempt in range(retries):
            try:
                r = self.session.get(url, params=params, timeout=30)
                if r.status_code == 429:
                    wait = int(r.headers.get("X-RequestCounter-Reset", 61))
                    print(f"\n  [FD] Rate limit — sleeping {wait}s …")
                    time.sleep(wait)
                    continue
                r.raise_for_status()
                return r.json()
            except requests.RequestException as e:
                if attempt == retries - 1:
                    print(f"  [FD] Failed: {e}")
                    return {}
                time.sleep(5)
        return {}

    def fetch_season(self, season):
        print(f"    Fetching FD season {season} …", end=" ", flush=True)
        data = self._get(f"competitions/{FD_COMP}/matches",
                         {"season": season, "status": "FINISHED"})
        matches = data.get("matches", [])
        print(f"{len(matches)} matches")
        return matches


# ══════════════════════════════════════════════════════════════════════════════
# DATA MANAGER  —  orchestrates fetching, caching, backfill
# ══════════════════════════════════════════════════════════════════════════════

class WeatherFetcher:
    """
    Fetches match-day weather from Open-Meteo (completely free, no API key).
    Used to adjust goal lambda for rain and high wind.

    Weather effects on goals (from literature):
      Heavy rain (>5mm/h)  : −8% to lambda (slippery ball, tired legs)
      Strong wind (>40km/h): −5% to lambda (long-ball chaos, less possession)
      Combined             : up to −12% combined

    Effects on corners:
      Heavy rain  : −6% (fewer crosses, shorter play)
      Strong wind : +3% (more misplaced passes → more corner situations)

    All effects are multiplicative adjustments to the base lambda/rates.
    """

    def __init__(self):
        self.session = requests.Session()

    def get_match_weather(self, venue_name: str, match_datetime_str: str) -> dict:
        """
        Returns weather adjustment factors for a given venue and match time.
        match_datetime_str: ISO format e.g. "2025-04-19T15:00:00+01:00"

        Returns dict:
          goal_factor    : float multiplier for lambda (default 1.0)
          corner_factor  : float multiplier for corners (default 1.0)
          description    : human-readable weather summary
        """
        default = {"goal_factor": 1.0, "corner_factor": 1.0,
                   "description": "Weather unavailable"}

        coords = VENUE_COORDS.get(venue_name)
        if not coords:
            # Try partial match
            for k, v in VENUE_COORDS.items():
                if any(w in venue_name for w in k.split()[:2]):
                    coords = v
                    break
        if not coords:
            return default

        try:
            match_dt = pd.to_datetime(match_datetime_str).tz_localize(None)
            date_str = match_dt.strftime("%Y-%m-%d")

            resp = self.session.get(
                WEATHER_BASE,
                params={
                    "latitude"      : coords[0],
                    "longitude"     : coords[1],
                    "hourly"        : "precipitation,wind_speed_10m",
                    "daily"         : "wind_speed_10m_max,precipitation_sum",
                    "forecast_days" : 7,
                    "timezone"      : "Europe/London",
                },
                timeout=10
            )
            resp.raise_for_status()
            data = resp.json()

            # Find the match hour index
            hourly_times = data.get("hourly", {}).get("time", [])
            match_hour   = match_dt.strftime("%Y-%m-%dT%H:00")
            try:
                idx = hourly_times.index(match_hour)
            except ValueError:
                # Fall back to closest hour
                idx = 0

            precip    = data["hourly"]["precipitation"][idx]    # mm/h
            wind      = data["hourly"]["wind_speed_10m"][idx]   # km/h

            # Compute factors
            goal_factor   = 1.0
            corner_factor = 1.0
            parts         = []

            if precip >= 5.0:
                goal_factor   *= 0.92
                corner_factor *= 0.94
                parts.append(f"Heavy rain {precip:.1f}mm/h → goals −8%")
            elif precip >= 2.0:
                goal_factor   *= 0.96
                corner_factor *= 0.97
                parts.append(f"Light rain {precip:.1f}mm/h → goals −4%")

            if wind >= 40.0:
                goal_factor   *= 0.95
                corner_factor *= 1.03
                parts.append(f"Strong wind {wind:.0f}km/h → goals −5%")
            elif wind >= 25.0:
                goal_factor   *= 0.98
                parts.append(f"Moderate wind {wind:.0f}km/h → goals −2%")

            description = "  ·  ".join(parts) if parts else \
                          f"Clear  rain:{precip:.1f}mm  wind:{wind:.0f}km/h"

            return {
                "goal_factor"  : round(goal_factor, 4),
                "corner_factor": round(corner_factor, 4),
                "description"  : description,
                "precip_mm"    : precip,
                "wind_kmh"     : wind,
            }

        except Exception as e:
            return default


class DataManager:
    """
    Manages all data operations:
    - Loads match history from MySQL
    - Runs background backfill of rich stats from API-Football
    - Fetches live data (injuries, H2H) for each prediction
    - Shows progress messages as requested
    """

    def __init__(self, db: DB, afl: APIFootball, fd: FootballDataOrg):
        self.db      = db
        self.afl     = afl
        self.fd      = fd
        self.weather = WeatherFetcher()

    # ── Startup: sync new matches + show backfill status ──────────────────────
    def startup_sync(self):
        print("\n[Data] Checking for new matches …")
        self._sync_basic_matches()
        self._show_backfill_status()
        # Backfill now runs per-fixture (after you enter the teams)
        # so your daily budget goes to the teams you actually care about.

    def _sync_basic_matches(self):
        """
        Sync basic match data (goals, dates, teams) using API-Football
        as the sole primary source for all 3 seasons.

        Budget logic:
        - Seasons 2022 and 2023 are complete. Fetch once, never again.
        - Season 2024 (current): re-fetch only if not fetched in last 6 hours.
        - Each season costs exactly 1 API request.
        - If budget is too low, skip silently — cached data still works.
        """
        for season in AFL_SEASONS:
            existing = self.db.fetchone(
                "SELECT COUNT(*) as n FROM matches_basic WHERE season=%s",
                (season,)
            )
            n_existing = existing["n"] if existing else 0

            # Completed seasons — fetch once, never again
            if season in (2022, 2023) and n_existing >= 370:
                print(f"  [Sync] Season {season}: complete ({n_existing} matches) ✓")
                continue

            # Most recent complete season (2024/25) — skip if fetched in last 6 hours
            if season == 2024 and n_existing > 0:
                last = self.db.fetchone(
                    "SELECT MAX(created_at) as last FROM matches_basic "
                    "WHERE season=%s", (season,)
                )
                if last and last["last"]:
                    age_h = (datetime.now() - last["last"]).total_seconds() / 3600
                    if age_h < 6:
                        print(f"  [Sync] Season {season} (2024/25): fetched "
                              f"{age_h:.1f}h ago — skipping.")
                        continue

            # Check budget before fetching
            remaining = self.afl._remaining()
            if remaining < 1:
                print(f"  [Sync] Season {season}: no API budget — "
                      f"using {n_existing} cached matches.")
                continue

            # Fetch from API-Football
            print(f"  [Sync] Season {season} …", end=" ", flush=True)
            raw = self.afl.fetch_fixtures(season)
            if not raw:
                print(f"failed — using {n_existing} cached matches.")
                continue

            inserted = 0
            for m in raw:
                ft = m.get("score", {}).get("fulltime", {})
                ht = m.get("score", {}).get("halftime", {})
                if not ft or ft.get("home") is None:
                    continue

                fix    = m.get("fixture", {})
                teams  = m.get("teams", {})
                mid    = fix.get("id")
                if not mid:
                    continue

                exists = self.db.fetchone(
                    "SELECT match_id FROM matches_basic WHERE match_id=%s",
                    (mid,)
                )
                if exists:
                    continue

                match_date_raw = fix.get("date", "")
                try:
                    match_date = pd.to_datetime(match_date_raw).to_pydatetime().replace(tzinfo=None)
                except Exception:
                    continue

                self.db.execute(
                    """INSERT IGNORE INTO matches_basic
                       (match_id, source, season, match_date, matchday,
                        home_team_name, away_team_name,
                        home_goals, away_goals,
                        ht_home_goals, ht_away_goals)
                       VALUES (%s,'api-football',%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (mid, season, match_date,
                     m.get("league", {}).get("round", "").replace("Regular Season - ", ""),
                     teams.get("home", {}).get("name", ""),
                     teams.get("away", {}).get("name", ""),
                     int(ft["home"]), int(ft["away"]),
                     int(ht["home"]) if ht and ht.get("home") is not None else None,
                     int(ht["away"]) if ht and ht.get("away") is not None else None)
                )
                inserted += 1

            total_now = n_existing + inserted
            if inserted > 0:
                print(f"+{inserted} new  (total: {total_now})")
            else:
                print(f"up to date ({total_now} matches) ✓")

            # Small pause between season requests
            if season != AFL_SEASONS[-1]:
                time.sleep(1.5)

    def _show_backfill_status(self):
        total    = self.db.fetchone("SELECT COUNT(*) as n FROM matches_basic")
        enriched = self.db.fetchone(
            "SELECT COUNT(DISTINCT match_id) as n FROM matches_stats"
        )
        n_total    = total["n"]    if total    else 0
        n_enriched = enriched["n"] if enriched else 0
        n_pending  = n_total - n_enriched
        remaining  = self.afl._remaining()

        # Per-season breakdown
        season_rows = self.db.fetchall(
            "SELECT season, COUNT(*) as n FROM matches_basic GROUP BY season ORDER BY season"
        )
        season_str = "  ".join(
            f"{r['season']}:{r['n']}" for r in season_rows
        ) if season_rows else "none"

        print(f"\n[Data] Matches in DB: {n_total} total  "
              f"({season_str})")
        print(f"[Data] Enriched with xG/shots/corners: {n_enriched}  |  "
              f"Pending: {n_pending}")
        print(f"[API]  Budget today: {remaining}/{AFL_MAX_DAILY} requests remaining")

        if n_pending > 0 and remaining > 6:
            # Each match costs 3 requests (stats + lineups + events)
            can_do = (remaining - 6) // 3
            print(f"[Backfill] Can enrich ~{can_do} more matches today "
                  f"(fixture teams go first)")

    def _run_backfill_batch(self, priority_teams: list = None):
        """
        Fetch stats for matches using today's remaining budget.
        If priority_teams is given, those teams' matches are enriched first.
        This ensures the two teams you are predicting always have rich data.
        """
        remaining = self.afl._remaining()
        # Reserve 6 requests for live prediction use (injuries x2, H2H, upcoming)
        budget = max(0, remaining - 6)
        if budget < 3:
            return

        # Build query — prioritise fixture teams if given
        if priority_teams and len(priority_teams) == 2:
            pending = self.db.fetchall(
                """SELECT b.match_id FROM matches_basic b
                   LEFT JOIN matches_stats s ON b.match_id = s.match_id
                   WHERE s.match_id IS NULL
                     AND (b.home_team_name IN (%s,%s)
                          OR b.away_team_name IN (%s,%s))
                   ORDER BY b.match_date DESC
                   LIMIT %s""",
                (*priority_teams, *priority_teams, budget // 3)
            )
            label = f"{priority_teams[0]} & {priority_teams[1]}"
        else:
            pending = self.db.fetchall(
                """SELECT b.match_id FROM matches_basic b
                   LEFT JOIN matches_stats s ON b.match_id = s.match_id
                   WHERE s.match_id IS NULL
                   ORDER BY b.match_date DESC
                   LIMIT %s""",
                (budget // 3,)
            )
            label = "all teams"

        if not pending:
            print(f"[Backfill] All matches for {label} fully enriched ✓")
            return

        n = len(pending)
        print(f"\n[Backfill] Enriching {n} matches for {label} "
              f"({n*3} requests) …")

        for row in pending:
            self._enrich_match(row["match_id"])
            time.sleep(1.2)

        enriched_now = self.db.fetchone(
            "SELECT COUNT(DISTINCT match_id) as n FROM matches_stats"
        )
        print(f"[Backfill] Done. Total enriched: "
              f"{enriched_now['n'] if enriched_now else 0}")

    def _enrich_match(self, match_id):
        """Fetch and store stats, lineups, events for one match."""
        # Stats
        stats = self.afl.fetch_match_stats(match_id)
        if stats:
            for team_data in stats:
                team_name = team_data.get("team", {}).get("name", "")
                is_home = None
                home_row = self.db.fetchone(
                    "SELECT home_team_name FROM matches_basic WHERE match_id=%s",
                    (match_id,)
                )
                if home_row:
                    is_home = 1 if team_name == home_row["home_team_name"] else 0

                s = {item["type"]: item.get("value")
                     for item in team_data.get("statistics", [])}

                def parse_stat(val):
                    if val is None: return None
                    if isinstance(val, str):
                        val = val.replace("%", "").strip()
                        try: return float(val)
                        except: return None
                    return float(val) if val is not None else None

                self.db.execute(
                    """INSERT IGNORE INTO matches_stats
                       (match_id, team_name, is_home,
                        shots_total, shots_on_target, shots_off_target, shots_blocked,
                        possession_pct, passes_total, passes_accurate, pass_accuracy_pct,
                        fouls, yellow_cards, red_cards, corners, offsides, saves, xg)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (match_id, team_name, is_home,
                     parse_stat(s.get("Total Shots")),
                     parse_stat(s.get("Shots on Goal")),
                     parse_stat(s.get("Shots off Goal")),
                     parse_stat(s.get("Blocked Shots")),
                     parse_stat(s.get("Ball Possession")),
                     parse_stat(s.get("Total passes")),
                     parse_stat(s.get("Passes accurate")),
                     parse_stat(s.get("Passes %")),
                     parse_stat(s.get("Fouls")),
                     parse_stat(s.get("Yellow Cards")),
                     parse_stat(s.get("Red Cards")),
                     parse_stat(s.get("Corner Kicks")),
                     parse_stat(s.get("Offsides")),
                     parse_stat(s.get("Goalkeeper Saves")),
                     parse_stat(s.get("expected_goals")))
                )

        # Lineups
        lineups = self.afl.fetch_lineups(match_id)
        if lineups:
            home_row = self.db.fetchone(
                "SELECT home_team_name FROM matches_basic WHERE match_id=%s",
                (match_id,)
            )
            for team_data in lineups:
                team_name = team_data.get("team", {}).get("name", "")
                is_home = 1 if home_row and team_name == home_row["home_team_name"] else 0
                for p in team_data.get("startXI", []):
                    pi = p.get("player", {})
                    self.db.execute(
                        """INSERT IGNORE INTO matches_lineups
                           (match_id, team_name, is_home, player_name, position,
                            is_starter, shirt_number)
                           VALUES (%s,%s,%s,%s,%s,1,%s)""",
                        (match_id, team_name, is_home,
                         pi.get("name", ""), pi.get("pos", ""),
                         pi.get("number"))
                    )
                for p in team_data.get("substitutes", []):
                    pi = p.get("player", {})
                    self.db.execute(
                        """INSERT IGNORE INTO matches_lineups
                           (match_id, team_name, is_home, player_name, position,
                            is_starter, shirt_number)
                           VALUES (%s,%s,%s,%s,%s,0,%s)""",
                        (match_id, team_name, is_home,
                         pi.get("name", ""), pi.get("pos", ""),
                         pi.get("number"))
                    )

        # Events
        events = self.afl.fetch_events(match_id)
        if events:
            for ev in events:
                self.db.execute(
                    """INSERT IGNORE INTO match_events
                       (match_id, minute, extra_time, team_name, player_name,
                        event_type, detail)
                       VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                    (match_id,
                     ev.get("time", {}).get("elapsed"),
                     ev.get("time", {}).get("extra", 0),
                     ev.get("team", {}).get("name", ""),
                     ev.get("player", {}).get("name", ""),
                     ev.get("type", ""),
                     ev.get("detail", ""))
                )


    def team_non_goal_rates(self, team: str) -> dict:
        """
        Compute team-specific corners and cards rates from DB.
        Returns per-game averages. Falls back to PL averages if < 5 matches enriched.
        """
        rows = self.db.fetchall(
            """SELECT s.corners, s.yellow_cards
               FROM matches_stats s
               JOIN matches_basic b ON s.match_id = b.match_id
               WHERE (b.home_team_name = %s AND s.is_home = 1)
                  OR (b.away_team_name = %s AND s.is_home = 0)
               ORDER BY b.match_date DESC
               LIMIT 40""",
            (team, team)
        )
        if not rows or len(rows) < 5:
            return {}
        corners = [r["corners"] for r in rows if r["corners"] is not None]
        cards   = [r["yellow_cards"] for r in rows if r["yellow_cards"] is not None]
        return {
            "avg_corners": float(np.mean(corners)) if corners else None,
            "avg_cards"  : float(np.mean(cards))   if cards   else None,
            "n_matches"  : len(rows),
        }

    def referee_card_rate(self, referee_name: str) -> float:
        """
        Compute referee-specific card rate as a multiplier vs PL average.
        1.0 = average. >1.0 = card-happy referee. Falls back to 1.0 if unknown.
        """
        if not referee_name:
            return 1.0
        PL_AVG_CARDS_PER_GAME = 3.85
        rows2 = self.db.fetchall(
            """SELECT AVG(card_count) as avg_cards FROM (
                   SELECT b.match_id, COUNT(e.id) as card_count
                   FROM matches_basic b
                   JOIN match_events e ON b.match_id = e.match_id
                   WHERE e.event_type IN ('Yellow Card','Red Card')
                     AND b.match_id IN (
                         SELECT DISTINCT match_id FROM match_events
                         WHERE player_name LIKE %s
                     )
                   GROUP BY b.match_id
               ) sub""",
            (f"%{referee_name.split()[-1]}%",)
        )
        if rows2 and rows2[0]["avg_cards"]:
            ref_avg = float(rows2[0]["avg_cards"])
            return float(np.clip(ref_avg / PL_AVG_CARDS_PER_GAME, 0.6, 1.8))
        return 1.0

    def load_for_fixture(self, home: str, away: str) -> pd.DataFrame:
        """
        Pull all matches involving either team.
        Joins basic + stats tables for a richer feature set.
        """
        rows = self.db.fetchall(
            """SELECT b.match_id, b.match_date, b.season,
                      b.home_team_name, b.away_team_name,
                      b.home_goals, b.away_goals,
                      b.ht_home_goals, b.ht_away_goals,
                      sh.shots_on_target  as home_sot,
                      sh.xg               as home_xg,
                      sh.corners          as home_corners,
                      sh.yellow_cards     as home_yellows,
                      sh.possession_pct   as home_possession,
                      sa.shots_on_target  as away_sot,
                      sa.xg               as away_xg,
                      sa.corners          as away_corners,
                      sa.yellow_cards     as away_yellows,
                      sa.possession_pct   as away_possession
               FROM matches_basic b
               LEFT JOIN matches_stats sh
                      ON b.match_id = sh.match_id AND sh.is_home = 1
               LEFT JOIN matches_stats sa
                      ON b.match_id = sa.match_id AND sa.is_home = 0
               WHERE b.home_team_name IN (%s,%s)
                  OR b.away_team_name IN (%s,%s)
               ORDER BY b.match_date""",
            (home, away, home, away)
        )
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows)
        df["match_date"] = pd.to_datetime(df["match_date"])
        return df

    # ── Fetch injuries for prediction ──────────────────────────────────────────
    def fetch_injuries_for_team(self, team_name: str) -> list:
        """Fetch current injuries. Uses cache if fetched today."""
        cached = self.db.fetchall(
            """SELECT player_name, injury_type, reason, expected_return
               FROM injuries
               WHERE team_name=%s AND DATE(fetched_at)=CURDATE()""",
            (team_name,)
        )
        if cached:
            return cached

        # Need team_id from teams table
        team_row = self.db.fetchone(
            "SELECT team_id FROM teams WHERE name LIKE %s LIMIT 1",
            (f"%{team_name.split()[0]}%",)
        )
        if not team_row:
            return []

        remaining = self.afl._remaining()
        if remaining <= 0:
            print(f"  [API] No budget for injury fetch.")
            return []

        print(f"  [API] Fetching injuries for {team_name} …", end=" ", flush=True)
        data = self.afl.fetch_injuries(team_row["team_id"])
        if not data:
            print("none found")
            return []

        # Clear old cached injuries for this team
        self.db.execute("DELETE FROM injuries WHERE team_name=%s", (team_name,))

        results = []
        for item in data:
            p = item.get("player", {})
            inj = item.get("injury", {})
            self.db.execute(
                """INSERT INTO injuries
                   (team_name, player_name, injury_type, reason, expected_return)
                   VALUES (%s,%s,%s,%s,%s)""",
                (team_name, p.get("name", ""), inj.get("type", ""),
                 inj.get("reason", ""), p.get("missing_estimated_date", "TBD"))
            )
            results.append({
                "player_name"    : p.get("name", ""),
                "injury_type"    : inj.get("type", ""),
                "reason"         : inj.get("reason", ""),
                "expected_return": p.get("missing_estimated_date", "TBD"),
            })
        print(f"{len(results)} injuries")
        return results

    # ── Fetch H2H ──────────────────────────────────────────────────────────────
    def fetch_h2h(self, home: str, away: str) -> pd.DataFrame:
        """Get head-to-head history from DB or API."""
        cached = self.db.fetchall(
            """SELECT match_date, home_team, away_team, home_goals, away_goals
               FROM head_to_head
               WHERE (team_a=%s AND team_b=%s)
                  OR (team_a=%s AND team_b=%s)
               ORDER BY match_date DESC LIMIT 20""",
            (home, away, away, home)
        )
        if cached and len(cached) >= 5:
            return pd.DataFrame(cached)

        # Try to get team IDs
        th = self.db.fetchone("SELECT team_id FROM teams WHERE name LIKE %s LIMIT 1",
                              (f"%{home.split()[0]}%",))
        ta = self.db.fetchone("SELECT team_id FROM teams WHERE name LIKE %s LIMIT 1",
                              (f"%{away.split()[0]}%",))

        if not th or not ta:
            return pd.DataFrame()

        remaining = self.afl._remaining()
        if remaining <= 0:
            return pd.DataFrame()

        print(f"  [API] Fetching H2H {home} vs {away} …", end=" ", flush=True)
        data = self.afl.fetch_h2h(th["team_id"], ta["team_id"])
        if not data:
            print("none")
            return pd.DataFrame()

        rows = []
        for m in data:
            ft = m.get("score", {}).get("fulltime", {})
            if not ft or ft.get("home") is None:
                continue
            self.db.execute(
                """INSERT INTO head_to_head
                   (team_a, team_b, match_date, home_team, away_team,
                    home_goals, away_goals)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                (home, away,
                 pd.to_datetime(m.get("fixture", {}).get("date")).replace(tzinfo=None),
                 m.get("teams", {}).get("home", {}).get("name", ""),
                 m.get("teams", {}).get("away", {}).get("name", ""),
                 int(ft["home"]), int(ft["away"]))
            )
            rows.append({
                "match_date": m.get("fixture", {}).get("date"),
                "home_team" : m.get("teams", {}).get("home", {}).get("name", ""),
                "away_team" : m.get("teams", {}).get("away", {}).get("name", ""),
                "home_goals": int(ft["home"]),
                "away_goals": int(ft["away"]),
            })
        print(f"{len(rows)} historical meetings")
        return pd.DataFrame(rows) if rows else pd.DataFrame()

    # ── Next fixtures ──────────────────────────────────────────────────────────
    def fetch_upcoming(self) -> list:
        print("  [API] Fetching upcoming PL fixtures …", end=" ", flush=True)
        data = self.afl.fetch_next_fixtures()
        if not data:
            print("unavailable")
            return []
        fixtures = []
        for m in data:
            status = m.get("fixture", {}).get("status", {}).get("short", "")
            if status in ("FT", "AET", "PEN"):
                continue   # skip already finished
            fixtures.append({
                "date"    : m.get("fixture", {}).get("date", "")[:10],
                "time"    : m.get("fixture", {}).get("date", "")[11:16],
                "home"    : m.get("teams", {}).get("home", {}).get("name", ""),
                "away"    : m.get("teams", {}).get("away", {}).get("name", ""),
                "venue"   : m.get("fixture", {}).get("venue", {}).get("name", ""),
                "matchday": m.get("league", {}).get("round", ""),
            })
        print(f"{len(fixtures)} upcoming fixtures")
        return fixtures

    # ── Load teams registry ────────────────────────────────────────────────────
    def ensure_teams(self):
        """Fetch and cache team registry only if not done yet."""
        count = self.db.fetchone("SELECT COUNT(*) as n FROM teams")
        if count and count["n"] > 0:
            return  # already cached — never fetch again
        remaining = self.afl._remaining()
        if remaining <= 2:
            print("  [API] Not enough budget to fetch team registry today.")
            return
        print("  [API] Fetching team registry (one-time) …", end=" ", flush=True)
        data = self.afl.fetch_teams()
        if not data:
            print("failed")
            return
        for item in data:
            t = item.get("team", {})
            self.db.execute(
                """INSERT IGNORE INTO teams (team_id, name, short_name, country)
                   VALUES (%s,%s,%s,%s)""",
                (t.get("id"), t.get("name", ""),
                 t.get("code", ""), t.get("country", ""))
            )
        print(f"{len(data)} teams cached ✓")


    # ── Fixture auto-lookup ────────────────────────────────────────────────────
    def find_fixture(self, home: str, away: str) -> dict:
        """
        Find the upcoming scheduled fixture between two teams.
        Looks up team IDs from the teams table, then queries API-Football
        for the next match between them within 60 days.
        Returns fixture info dict or empty dict if not found / no budget.
        """
        if self.afl._remaining() < 1:
            return {}

        # Get team IDs from DB
        th = self.db.fetchone(
            "SELECT team_id FROM teams WHERE name LIKE %s LIMIT 1",
            (f"%{home.split()[0]}%",)
        )
        ta = self.db.fetchone(
            "SELECT team_id FROM teams WHERE name LIKE %s LIMIT 1",
            (f"%{away.split()[0]}%",)
        )
        if not th or not ta:
            return {}

        print(f"  [API] Looking up fixture: {home} vs {away} …", end=" ", flush=True)
        fix = self.afl.find_upcoming_fixture(
            th["team_id"], ta["team_id"], season=2024
        )
        if fix and fix.get("fixture_id"):
            dt_str = fix["date"][:16].replace("T", " ") if fix.get("date") else "TBD"
            print(f"found  (ID:{fix['fixture_id']}  {dt_str}  {fix.get('venue','')})")
        else:
            print("not found in next 60 days")
        return fix

    # ── Automatic odds fetching ────────────────────────────────────────────────
    def fetch_auto_odds(self, fixture_id: int) -> dict:
        """
        Fetch pre-match Bet365 odds and map to our market key format.
        Returns dict compatible with the existing mkt_probs / EV system.
        """
        if not fixture_id or self.afl._remaining() < 1:
            return {}

        print(f"  [API] Fetching Bet365 odds (fixture {fixture_id}) …",
              end=" ", flush=True)
        raw = self.afl.fetch_prematch_odds(fixture_id)
        if not raw:
            print("unavailable")
            return {}

        odds = {}

        # ── 1X2 ───────────────────────────────────────────────────────────────
        mw = raw.get("Match Winner", [])
        for v in mw:
            val = v.get("value", ""); odd = float(v.get("odd", 0))
            if odd <= 1.0: continue
            if val == "Home":  odds["home_win"]  = odd
            elif val == "Draw": odds["draw"]      = odd
            elif val == "Away": odds["away_win"]  = odd

        # ── Over/Under goals ──────────────────────────────────────────────────
        for bet_name, values in raw.items():
            if "Goals Over/Under" in bet_name or "Over/Under" in bet_name:
                for v in values:
                    val = v.get("value", ""); odd = float(v.get("odd", 0))
                    if odd <= 1.0: continue
                    if val == "Over 1.5":  odds["over_1.5"]  = odd
                    elif val == "Under 1.5": odds["under_1.5"] = odd
                    elif val == "Over 2.5":  odds["over_2.5"]  = odd
                    elif val == "Under 2.5": odds["under_2.5"] = odd
                    elif val == "Over 3.5":  odds["over_3.5"]  = odd
                    elif val == "Under 3.5": odds["under_3.5"] = odd
                    elif val == "Over 4.5":  odds["over_4.5"]  = odd
                    elif val == "Under 4.5": odds["under_4.5"] = odd

        # ── BTTS ──────────────────────────────────────────────────────────────
        btts = raw.get("Both Teams Score", [])
        for v in btts:
            val = v.get("value", ""); odd = float(v.get("odd", 0))
            if odd <= 1.0: continue
            if val == "Yes": odds["btts_yes"] = odd
            elif val == "No": odds["btts_no"]  = odd

        # ── Double Chance ─────────────────────────────────────────────────────
        dc = raw.get("Double Chance", [])
        for v in dc:
            val = v.get("value", ""); odd = float(v.get("odd", 0))
            if odd <= 1.0: continue
            if val == "Home/Draw":  odds["dc_1x"] = odd
            elif val == "Draw/Away": odds["dc_x2"] = odd
            elif val == "Home/Away": odds["dc_12"] = odd

        # ── Asian Handicap ────────────────────────────────────────────────────
        for bet_name, values in raw.items():
            if "Asian Handicap" in bet_name:
                for v in values:
                    val = v.get("value", ""); odd = float(v.get("odd", 0))
                    if odd <= 1.0: continue
                    if val == "Home -0.5":  odds["ah_-0.5"] = odd
                    elif val == "Away +0.5": odds["ah_+0.5"] = odd
                    elif val == "Home -1.5": odds["ah_-1.5"] = odd
                    elif val == "Away +1.5": odds["ah_+1.5"] = odd
                    elif val == "Home +0.5": odds["ah_h+0.5"] = odd
                    elif val == "Away -0.5": odds["ah_a-0.5"] = odd

        # ── Corners ───────────────────────────────────────────────────────────
        for bet_name, values in raw.items():
            if "Corner" in bet_name and "Over/Under" in bet_name:
                for v in values:
                    val = v.get("value", ""); odd = float(v.get("odd", 0))
                    if odd <= 1.0: continue
                    for line in ("7.5", "8.5", "9.5", "10.5", "11.5", "12.5"):
                        if val == f"Over {line}":  odds[f"cor_ov_{line}"] = odd
                        elif val == f"Under {line}": odds[f"cor_un_{line}"] = odd

        # ── Cards ─────────────────────────────────────────────────────────────
        for bet_name, values in raw.items():
            if "Card" in bet_name and "Over/Under" in bet_name:
                for v in values:
                    val = v.get("value", ""); odd = float(v.get("odd", 0))
                    if odd <= 1.0: continue
                    for line in ("1.5", "2.5", "3.5", "4.5", "5.5"):
                        if val == f"Over {line}":  odds[f"crd_ov_{line}"] = odd
                        elif val == f"Under {line}": odds[f"crd_un_{line}"] = odd

        n = len(odds)
        print(f"{n} markets fetched")
        return odds

    # ── Automatic lineup fetching ──────────────────────────────────────────────
    def fetch_auto_lineups(self, fixture_id: int,
                           home: str, away: str) -> tuple:
        """
        Fetch confirmed lineups from API-Football.
        Only available ~60 minutes before kickoff.
        Returns (players_home, players_away) each as list of dicts.
        """
        if not fixture_id or self.afl._remaining() < 1:
            return [], []

        print(f"  [API] Checking lineups …", end=" ", flush=True)
        raw = self.afl.fetch_live_lineups(fixture_id)
        if not raw:
            print("not released yet")
            return [], []

        players_h, players_a = [], []
        for team_data in raw:
            team_name = team_data.get("team", {}).get("name", "")
            is_home   = home.lower() in team_name.lower() or                         team_name.lower() in home.lower()
            target    = players_h if is_home else players_a

            for p in team_data.get("startXI", []):
                pi  = p.get("player", {})
                pos = pi.get("pos", "MF")
                # Map API position codes to our GK/DF/MF/FW
                pos_map = {"G": "GK", "D": "DF", "M": "MF", "F": "FW"}
                pos     = pos_map.get(pos[0] if pos else "M", "MF")
                target.append({
                    "name"    : pi.get("name", ""),
                    "position": pos,
                    "number"  : pi.get("number", 0),
                    "id"      : pi.get("id", 0),
                })

        h_n = len(players_h); a_n = len(players_a)
        if h_n or a_n:
            print(f"✓  {home}: {h_n} players  |  {away}: {a_n} players")
        else:
            print("not released yet")
        return players_h, players_a

    # ── Player xG/90 from API stats ────────────────────────────────────────────
    def enrich_players_with_xg(self, players: list,
                                team_name: str, season: int = 2024) -> list:
        """
        Match lineup players to their historical xG/90 from the API stats.
        Replaces position-weight estimates with actual player data.
        Returns players list with added 'xg_per90' key.
        """
        if not players or self.afl._remaining() < 1:
            return players

        team_row = self.db.fetchone(
            "SELECT team_id FROM teams WHERE name LIKE %s LIMIT 1",
            (f"%{team_name.split()[0]}%",)
        )
        if not team_row:
            return players

        print(f"  [API] Fetching player xG stats for {team_name} …",
              end=" ", flush=True)
        raw = self.afl.fetch_player_stats(team_row["team_id"], season)
        if not raw:
            print("unavailable")
            return players

        # Build lookup: player name → xG per 90
        xg_lookup = {}
        for item in raw:
            p     = item.get("player", {})
            stats = item.get("statistics", [{}])[0]
            goals = stats.get("goals", {})
            games = stats.get("games", {})
            xg_val   = goals.get("assists")  # API sometimes puts xG here
            shots    = goals.get("total", 0) or 0
            minutes  = games.get("minutes", 0) or 0
            # xG per 90 proxy: shots on target * conversion rate
            sot      = goals.get("saves", 0) or 0   # reuse field
            actual_xg = goals.get("rating")          # sometimes xG rating
            # Best approximation from free tier: goals/90
            goals_n  = goals.get("total", 0) or 0
            xg90     = (goals_n / max(minutes, 1)) * 90 if minutes > 0 else 0
            name     = p.get("name", "")
            if name:
                xg_lookup[name.lower()] = float(xg90)

        enriched = []
        for p in players:
            p_copy = p.copy()
            pname  = p.get("name", "").lower()
            # Fuzzy match player name
            xg90 = xg_lookup.get(pname)
            if xg90 is None:
                # Try last name match
                last = pname.split()[-1] if pname else ""
                for k, v in xg_lookup.items():
                    if last and last in k:
                        xg90 = v
                        break
            p_copy["xg_per90"] = xg90 if xg90 is not None else None
            enriched.append(p_copy)

        matched = sum(1 for p in enriched if p.get("xg_per90") is not None)
        print(f"{matched}/{len(enriched)} players matched")
        return enriched

    # ── Standings context ──────────────────────────────────────────────────────
    def standings_context(self, home: str, away: str,
                          season: int = 2024) -> dict:
        """
        Fetch current PL standings to provide league position context.
        Returns dict with position, points, form string for each team.
        """
        if self.afl._remaining() < 1:
            return {}

        print(f"  [API] Fetching standings …", end=" ", flush=True)
        table = self.afl.fetch_standings(season)
        if not table:
            print("unavailable")
            return {}

        result = {}
        for entry in table:
            team_name = entry.get("team", {}).get("name", "")
            if home.split()[0].lower() in team_name.lower() or                team_name.lower() in home.lower():
                result["home"] = {
                    "team"  : team_name,
                    "pos"   : entry.get("rank", 0),
                    "pts"   : entry.get("points", 0),
                    "played": entry.get("all", {}).get("played", 0),
                    "gd"    : entry.get("goalsDiff", 0),
                    "form"  : entry.get("form", ""),
                }
            elif away.split()[0].lower() in team_name.lower() or                  team_name.lower() in away.lower():
                result["away"] = {
                    "team"  : team_name,
                    "pos"   : entry.get("rank", 0),
                    "pts"   : entry.get("points", 0),
                    "played": entry.get("all", {}).get("played", 0),
                    "gd"    : entry.get("goalsDiff", 0),
                    "form"  : entry.get("form", ""),
                }

        n = len(result)
        if n == 2:
            h = result["home"]; a = result["away"]
            print(f"  {h['team']} (#{h['pos']})  vs  {a['team']} (#{a['pos']})")
        else:
            print(f"{n} teams found in table")
        return result

    def all_team_names(self) -> list:
        rows = self.db.fetchall(
            "SELECT DISTINCT home_team_name FROM matches_basic"
        )
        names = [r["home_team_name"] for r in rows]
        rows2 = self.db.fetchall(
            "SELECT DISTINCT away_team_name FROM matches_basic"
        )
        names += [r["away_team_name"] for r in rows2]
        return sorted(set(names))


# ══════════════════════════════════════════════════════════════════════════════
# TEAM RESOLVER
# ══════════════════════════════════════════════════════════════════════════════
def resolve_team(query: str, known: list) -> str:
    q = query.lower().strip()
    for alias, canonical in ALIASES.items():
        if q == alias:
            matches = [t for t in known if canonical.lower() in t.lower()]
            if matches:
                return matches[0]
    for t in known:
        if t.lower() == q:
            return t
    hits = [t for t in known if q in t.lower()]
    if len(hits) == 1:
        return hits[0]
    if hits:
        close = get_close_matches(q, [t.lower() for t in hits], n=1, cutoff=0.3)
        if close:
            return next(t for t in hits if t.lower() == close[0])
        return hits[0]
    close = get_close_matches(q, [t.lower() for t in known], n=3, cutoff=0.3)
    raise ValueError(
        f"Cannot find '{query}'. "
        f"Closest: {get_close_matches(q, [t.lower() for t in known], n=3)}"
    )


# ══════════════════════════════════════════════════════════════════════════════
# TIME-DECAY WEIGHTS
# ══════════════════════════════════════════════════════════════════════════════
def time_weights(dates: pd.Series) -> np.ndarray:
    ref      = dates.max()
    days_ago = (ref - dates).dt.days.values.astype(float)
    lam      = np.log(2) / DECAY_HALF_LIFE
    w        = np.exp(-lam * days_ago)
    w[days_ago <= RECENT_DAYS] *= 1.5
    return w / w.mean()


# ══════════════════════════════════════════════════════════════════════════════
# BAYESIAN MODEL  —  MAP + Laplace (fast, <5 sec)
# ══════════════════════════════════════════════════════════════════════════════
class FixtureModel:
    """
    Extended Hierarchical Poisson model using:
      - Goals (basic signal)
      - xG (more stable than goals)
      - Shots on target (defensive solidity proxy)
      - H2H factor (historical matchup bias)
      - Injury adjustment (lambda scaling)
      - Momentum (last-5 trajectory)

    λ_home = exp(intercept + home_adv
                 + 0.4*attack_goals[h] + 0.6*attack_xg[h]
                 - 0.4*defense_goals[a] - 0.6*defense_xg[a])
             × h2h_factor × injury_factor × momentum_factor

    Inference: MAP (L-BFGS-B) + diagonal Laplace approximation
    """

    def __init__(self, df: pd.DataFrame, home: str, away: str,
                 h2h_df: pd.DataFrame = None,
                 home_injuries: list = None,
                 away_injuries: list = None,
                 weather: dict = None):
        self.df      = df
        self.home    = home
        self.away    = away
        self.h2h_df  = h2h_df if h2h_df is not None else pd.DataFrame()
        self.home_inj= home_injuries or []
        self.away_inj= away_injuries or []
        self.weather = weather or {}
        self.weather_goal_factor   = float(
            self.weather.get("goal_factor", 1.0))
        self.weather_corner_factor = float(
            self.weather.get("corner_factor", 1.0))

        self.teams = sorted(set(df["home_team_name"]) | set(df["away_team_name"]))
        self.t2i   = {t: i for i, t in enumerate(self.teams)}
        self.nt    = len(self.teams)

        self.hi = np.array([self.t2i[t] for t in df["home_team_name"]])
        self.ai = np.array([self.t2i[t] for t in df["away_team_name"]])
        self.hg = df["home_goals"].values.astype(float)
        self.ag = df["away_goals"].values.astype(float)
        self.w  = time_weights(df["match_date"])

        # xG signal (fall back to goals if missing)
        hxg = df["home_xg"].fillna(df["home_goals"]).values.astype(float)
        axg = df["away_xg"].fillna(df["away_goals"]).values.astype(float)
        self.hxg = hxg
        self.axg = axg

        # shots on target ratio (proxy for defensive quality)
        hsot = df["home_sot"].fillna(3.0).values.astype(float)
        asot = df["away_sot"].fillna(2.5).values.astype(float)
        self.hsot = np.clip(hsot / (hsot + asot + 1e-9), 0.2, 0.8)

        self.n_params    = 2 + 2 * self.nt + 1   # +1 for Dixon-Coles rho
        self.post_samples= None
        self.map_theta   = None

        # Pre-compute adjustment factors
        self.h2h_factor   = self._compute_h2h_factor()
        self.inj_h_factor = self._compute_injury_factor(self.home_inj, is_home=True)
        self.inj_a_factor = self._compute_injury_factor(self.away_inj, is_home=False)
        self.mom_h        = self._compute_momentum(home, is_home=True)
        self.mom_a        = self._compute_momentum(away, is_home=False)
        self.fatigue_h    = self._compute_fatigue(home)
        self.fatigue_a    = self._compute_fatigue(away)

    def _compute_h2h_factor(self) -> float:
        """
        Slight multiplicative adjustment to home lambda based on H2H record.
        Range: 0.92 – 1.08
        """
        if self.h2h_df.empty or len(self.h2h_df) < 3:
            return 1.0
        h2h = self.h2h_df.copy()
        home_wins = ((h2h["home_team"] == self.home) &
                     (h2h["home_goals"] > h2h["away_goals"])).sum()
        away_wins = ((h2h["away_team"] == self.home) &
                     (h2h["away_goals"] > h2h["home_goals"])).sum()
        total = len(h2h)
        home_win_rate = (home_wins + away_wins) / total
        # Neutral = 0.45 win rate; above that → positive factor
        factor = 1.0 + (home_win_rate - 0.45) * 0.15
        return float(np.clip(factor, 0.92, 1.08))

    def _compute_injury_factor(self, injuries: list, is_home: bool) -> float:
        """
        Attenuate lambda based on key player absences.
        Rough impact: goalkeeper = -0.05, striker = -0.08, midfielder = -0.04
        """
        impact = 0.0
        pos_map = {"GK": 0.04, "FW": 0.08, "MF": 0.04, "DF": 0.03}
        for inj in injuries:
            pos = inj.get("position", "MF")
            impact += pos_map.get(pos, 0.04)
        factor = max(0.75, 1.0 - impact)
        return float(factor)

    def _compute_momentum(self, team: str, is_home: bool, n: int = 5) -> float:
        """
        Momentum factor using home/away split.
        A team's home momentum and away momentum are computed separately
        because many teams perform very differently at home vs away.
        Range: 0.92 – 1.08
        """
        df   = self.df
        if is_home:
            mask = df["home_team_name"] == team
            xg_col = "home_xg"; g_col = "home_goals"
        else:
            mask = df["away_team_name"] == team
            xg_col = "away_xg"; g_col = "away_goals"

        games = df[mask].tail(n)
        if len(games) < 2:
            # Fall back to combined if insufficient split data
            mask2 = (df["home_team_name"] == team) | (df["away_team_name"] == team)
            games = df[mask2].tail(n)
            xg_values = []
            for _, r in games.iterrows():
                if r["home_team_name"] == team:
                    xg_values.append(r["home_xg"] if pd.notna(r["home_xg"]) else r["home_goals"])
                else:
                    xg_values.append(r["away_xg"] if pd.notna(r["away_xg"]) else r["away_goals"])
        else:
            xg_values = []
            for _, r in games.iterrows():
                val = r[xg_col] if pd.notna(r[xg_col]) else r[g_col]
                xg_values.append(float(val))

        if not xg_values:
            return 1.0

        season_avg = float(np.mean(xg_values))
        recent_avg = float(np.mean(xg_values[-3:])) if len(xg_values) >= 3 else season_avg
        factor = 1.0 + (recent_avg - season_avg) * 0.08
        return float(np.clip(factor, 0.92, 1.08))

    def _compute_fatigue(self, team: str) -> float:
        """
        Late-season fatigue factor based on matchday number.
        Teams playing matchday 30+ show average xG drop of ~4% in the data.
        Range: 0.94 – 1.0 (fatigue only reduces, never boosts)

        Also accounts for fixture congestion:
        if last match was < 4 days ago, apply additional -3% factor.
        """
        df   = self.df
        mask = (df["home_team_name"] == team) | (df["away_team_name"] == team)
        recent = df[mask].tail(1)

        fatigue = 1.0

        # Matchday effect — only applies if matchday column available
        if not recent.empty and "matchday" in recent.columns:
            try:
                md = int(recent.iloc[0]["matchday"]) if pd.notna(recent.iloc[0].get("matchday")) else 0
                if md >= 35:
                    fatigue *= 0.94   # last 4 matchdays — maximum fatigue
                elif md >= 30:
                    fatigue *= 0.97   # matchdays 30-34
            except (ValueError, TypeError):
                pass

        # Fixture congestion — last match < 4 days ago
        if not recent.empty:
            try:
                last_date = pd.to_datetime(recent.iloc[0]["match_date"])
                days_since = (df["match_date"].max() - last_date).days
                if 0 < days_since <= 3:
                    fatigue *= 0.97   # midweek turnaround
            except Exception:
                pass

        return float(np.clip(fatigue, 0.88, 1.0))

    def _log_prior(self, theta):
        intercept = theta[0]; home_adv = theta[1]
        attack  = theta[2        : 2 + self.nt]
        defense = theta[2+self.nt: 2 + 2*self.nt]
        rho     = theta[2 + 2*self.nt]   # Dixon-Coles correlation param
        lp  = norm.logpdf(intercept, 0.0, 0.5)
        lp += norm.logpdf(home_adv,  0.3, 0.2)
        lp += norm.logpdf(attack,  0.0, 1.0).sum()
        lp += norm.logpdf(defense, 0.0, 1.0).sum()
        lp += norm.logpdf(rho,    -0.1, 0.2)   # small negative prior — typical in literature
        return lp

    def _log_likelihood(self, theta):
        intercept = theta[0]; home_adv = theta[1]
        attack  = theta[2        : 2 + self.nt]
        defense = theta[2+self.nt: 2 + 2*self.nt]
        rho     = theta[2 + 2*self.nt]   # Dixon-Coles rho

        # Blended signal: 40% goals + 60% xG
        log_lam_h = np.clip(
            intercept + home_adv
            + 0.4*attack[self.hi] + 0.6*(attack[self.hi] * self.hxg / (self.hg + 1e-6))
            - 0.4*defense[self.ai] - 0.6*(defense[self.ai] * self.axg / (self.ag + 1e-6)),
            -4, 4
        )
        log_lam_a = np.clip(
            intercept
            + 0.4*attack[self.ai] + 0.6*(attack[self.ai] * self.axg / (self.ag + 1e-6))
            - 0.4*defense[self.hi] - 0.6*(defense[self.hi] * self.hxg / (self.hg + 1e-6)),
            -4, 4
        )
        lam_h = np.exp(log_lam_h)
        lam_a = np.exp(log_lam_a)

        # Standard Poisson log-pmf (drop constant log(k!))
        ll_h = self.hg * log_lam_h - lam_h
        ll_a = self.ag * log_lam_a - lam_a

        # Dixon-Coles tau correction for low-scoring scorelines
        # tau(x, y, mu1, mu2, rho) adjusts the joint probability of
        # (0,0), (1,0), (0,1), (1,1) which Poisson systematically mis-estimates
        tau = np.ones(len(self.hg))
        eps = 1e-9
        m00 = (self.hg == 0) & (self.ag == 0)
        m10 = (self.hg == 1) & (self.ag == 0)
        m01 = (self.hg == 0) & (self.ag == 1)
        m11 = (self.hg == 1) & (self.ag == 1)
        tau[m00] = np.maximum(1 - lam_h[m00] * lam_a[m00] * rho, eps)
        tau[m10] = np.maximum(1 + lam_a[m10] * rho, eps)
        tau[m01] = np.maximum(1 + lam_h[m01] * rho, eps)
        tau[m11] = np.maximum(1 - rho, eps)

        ll_dc = np.log(np.maximum(tau, eps))
        return np.sum(self.w * (ll_h + ll_a + ll_dc))

    def _neg_lp(self, theta):
        return -(self._log_likelihood(theta) + self._log_prior(theta))

    def fit(self):
        n = len(self.df)
        print(f"  [Model] {n} matches  |  {self.nt} teams  |  {self.n_params} parameters")
        print(f"  [MAP]   Optimising … ", end="", flush=True)
        t0 = time.time()

        theta0 = np.zeros(self.n_params); theta0[1] = 0.3
        theta0[2 + 2*self.nt] = -0.1   # warm-start rho at literature value
        res = minimize(self._neg_lp, theta0, method="L-BFGS-B",
                       options={"maxiter": 2000, "ftol": 1e-10, "gtol": 1e-6})
        if not res.success:
            theta0 = np.random.normal(0, 0.1, self.n_params); theta0[1] = 0.3
            theta0[2 + 2*self.nt] = -0.1
            res = minimize(self._neg_lp, theta0, method="L-BFGS-B",
                           options={"maxiter": 3000, "ftol": 1e-12})

        self.map_theta = res.x
        print(f"done ({time.time()-t0:.1f}s)")

        print(f"  [Laplace] Computing covariance … ", end="", flush=True)
        t1 = time.time()
        eps = 1e-4; n_p = self.n_params
        f0  = self._neg_lp(self.map_theta)
        H_diag = np.zeros(n_p)
        for i in range(n_p):
            ei = np.zeros(n_p); ei[i] = eps
            H_diag[i] = (self._neg_lp(self.map_theta + ei) - 2*f0 +
                         self._neg_lp(self.map_theta - ei)) / eps**2
        post_var = 1.0 / np.maximum(H_diag, 1e-6)
        self.post_cov = np.diag(post_var)
        print(f"done ({time.time()-t1:.1f}s)")

        print(f"  [Sample] Drawing {N_POSTERIOR:,} posterior samples … ", end="", flush=True)
        self.post_samples = np.random.multivariate_normal(
            self.map_theta, self.post_cov, N_POSTERIOR
        )
        print(f"done  ({time.time()-t0:.1f}s total)\n")
        return self

    def lambdas(self):
        hi = self.t2i[self.home]; ai = self.t2i[self.away]
        p  = self.post_samples
        intercept = p[:, 0]; home_adv = p[:, 1]
        attack  = p[:, 2        : 2 + self.nt]
        defense = p[:, 2+self.nt: 2 + 2*self.nt]

        lh = np.clip(intercept + home_adv + attack[:,hi] - defense[:,ai], -4, 4)
        la = np.clip(intercept            + attack[:,ai] - defense[:,hi], -4, 4)

        mu_h = (np.exp(lh) * self.h2h_factor * self.inj_h_factor
                * self.mom_h * self.fatigue_h * self.weather_goal_factor)
        mu_a = (np.exp(la) * self.inj_a_factor
                * self.mom_a * self.fatigue_a * self.weather_goal_factor)

        # Extract rho posterior (Dixon-Coles correlation)
        self.rho_samples = p[:, 2 + 2*self.nt]
        return mu_h, mu_a

    @property
    def rho(self) -> float:
        """MAP estimate of Dixon-Coles rho parameter."""
        return float(self.map_theta[2 + 2*self.nt])

    def form(self, team: str, n: int = 5) -> dict:
        df   = self.df
        mask = (df["home_team_name"] == team) | (df["away_team_name"] == team)
        games= df[mask].tail(n)
        res, gf, ga, xgf = [], [], [], []
        for _, r in games.iterrows():
            if r["home_team_name"] == team:
                gf.append(r["home_goals"]); ga.append(r["away_goals"])
                xgf.append(r["home_xg"] if pd.notna(r["home_xg"]) else r["home_goals"])
                res.append("W" if r["home_goals"] > r["away_goals"]
                           else "D" if r["home_goals"] == r["away_goals"] else "L")
            else:
                gf.append(r["away_goals"]); ga.append(r["home_goals"])
                xgf.append(r["away_xg"] if pd.notna(r["away_xg"]) else r["away_goals"])
                res.append("W" if r["away_goals"] > r["home_goals"]
                           else "D" if r["away_goals"] == r["home_goals"] else "L")
        return {
            "form": "".join(res), "pts": res.count("W")*3 + res.count("D"),
            "gf": sum(gf), "ga": sum(ga), "gd": sum(gf)-sum(ga),
            "avg_xg": round(float(np.mean(xgf)), 2) if xgf else 0.0,
        }

    def strengths(self):
        atk = self.map_theta[2        : 2 + self.nt]
        dfn = self.map_theta[2+self.nt: 2 + 2*self.nt]
        return {t: (atk[i], dfn[i]) for t, i in self.t2i.items()}


# ══════════════════════════════════════════════════════════════════════════════
# MONTE CARLO SIMULATOR  —  all markets
# ══════════════════════════════════════════════════════════════════════════════
class Sim:
    def __init__(self, mu_h, mu_a, rho: float = -0.1):
        idx      = np.random.choice(len(mu_h), N_SIM, replace=True)
        self.mh  = mu_h[idx]; self.ma = mu_a[idx]
        self.sh  = np.random.poisson(self.mh)
        self.sa  = np.random.poisson(self.ma)
        self.tt  = self.sh + self.sa
        self.rho = rho   # Dixon-Coles correlation

    def result(self):
        return {"home_win": float((self.sh>self.sa).mean()),
                "draw"    : float((self.sh==self.sa).mean()),
                "away_win": float((self.sh<self.sa).mean())}

    def ou(self, lines=(0.5,1.5,2.5,3.5,4.5,5.5)):
        r = {}
        for l in lines:
            r[f"over_{l}"]  = float((self.tt > l).mean())
            r[f"under_{l}"] = float((self.tt <= l).mean())
        return r

    def btts(self):
        y = float(((self.sh>0)&(self.sa>0)).mean())
        return {"btts_yes": y, "btts_no": 1-y}

    def correct_score(self, mx=6):
        """
        Correct score probabilities with Dixon-Coles tau correction.
        The tau factor adjusts (0,0), (1,0), (0,1), (1,1) scorelines
        which Poisson systematically gets wrong — this is the main
        improvement for 1X2 and correct score accuracy.
        """
        mh_mean = float(self.mh.mean())
        ma_mean = float(self.ma.mean())
        rho     = self.rho

        def tau(x, y, lh, la, r):
            if   x == 0 and y == 0: return max(1 - lh * la * r, 1e-9)
            elif x == 1 and y == 0: return max(1 + la * r,       1e-9)
            elif x == 0 and y == 1: return max(1 + lh * r,       1e-9)
            elif x == 1 and y == 1: return max(1 - r,            1e-9)
            return 1.0

        rows = []
        for h in range(mx + 1):
            for a in range(mx + 1):
                raw_p = float(((self.sh == h) & (self.sa == a)).mean())
                t     = tau(h, a, mh_mean, ma_mean, rho)
                adj_p = raw_p * t
                rows.append({"score": f"{h}-{a}", "prob": max(adj_p, 0.0)})

        # Renormalise so probabilities sum to 1
        total = sum(r["prob"] for r in rows)
        if total > 0:
            for r in rows:
                r["prob"] /= total

        covered = sum(r["prob"] for r in rows)
        rows.append({"score": "other", "prob": max(0.0, 1 - covered)})
        return (pd.DataFrame(rows)
                  .sort_values("prob", ascending=False)
                  .reset_index(drop=True))

    def halftime(self):
        hh = np.random.poisson(self.mh*0.45)
        ha = np.random.poisson(self.ma*0.45)
        return {"ht_hw": float((hh>ha).mean()),
                "ht_d" : float((hh==ha).mean()),
                "ht_aw": float((hh<ha).mean()),
                "ht_ou_05": float(((hh+ha)>0.5).mean()),
                "ht_ou_15": float(((hh+ha)>1.5).mean())}

    def second_half(self):
        sh2 = np.random.poisson(self.mh*0.55)
        sa2 = np.random.poisson(self.ma*0.55)
        return {"sh_hw": float((sh2>sa2).mean()),
                "sh_d" : float((sh2==sa2).mean()),
                "sh_aw": float((sh2<sa2).mean()),
                "sh_ou_05": float(((sh2+sa2)>0.5).mean()),
                "sh_ou_15": float(((sh2+sa2)>1.5).mean())}

    def half_most_goals(self):
        hh = np.random.poisson(self.mh*0.45); ha = np.random.poisson(self.ma*0.45)
        sh2 = np.random.poisson(self.mh*0.55); sa2 = np.random.poisson(self.ma*0.55)
        ht_tot = hh+ha; sh_tot = sh2+sa2
        return {"first_half" : float((ht_tot > sh_tot).mean()),
                "second_half": float((sh_tot > ht_tot).mean()),
                "equal"      : float((ht_tot == sh_tot).mean())}

    def dc(self):
        r = self.result()
        return {"1X": r["home_win"]+r["draw"],
                "X2": r["away_win"]+r["draw"],
                "12": r["home_win"]+r["away_win"]}

    def draw_no_bet(self):
        r = self.result()
        base = r["home_win"] + r["away_win"] + 1e-9
        return {"dnb_home": r["home_win"]/base,
                "dnb_away": r["away_win"]/base}

    def win_to_nil(self):
        home_wtn = float(((self.sh>self.sa)&(self.sa==0)).mean())
        away_wtn = float(((self.sa>self.sh)&(self.sh==0)).mean())
        return {"home_wtn": home_wtn, "away_wtn": away_wtn}

    def three_way_handicap(self, hcp, half="FT"):
        if half == "FT":
            sh, sa = self.sh, self.sa
        elif half == "HT":
            sh = np.random.poisson(self.mh*0.45)
            sa = np.random.poisson(self.ma*0.45)
        else:  # 2H
            sh = np.random.poisson(self.mh*0.55)
            sa = np.random.poisson(self.ma*0.55)
        adj_diff = (sh + hcp) - sa
        return {"home": float((adj_diff > 0).mean()),
                "draw": float((adj_diff == 0).mean()),
                "away": float((adj_diff < 0).mean())}

    def asian_handicap(self, hcp):
        diff = (self.sh + hcp) - self.sa
        return {"home": float((diff>0).mean()),
                "away": float((diff<0).mean()),
                "push": float((diff==0).mean())}

    def exact_goals(self):
        r = {}
        for n in range(7):
            r[f"exactly_{n}"] = float((self.tt==n).mean())
        r["exactly_7plus"] = float((self.tt>=7).mean())
        return r

    def multiscores(self):
        """Grouped scoreline bundles."""
        cs = self.correct_score()
        def bundle(*scores):
            return sum(cs[cs["score"]==s]["prob"].values[0]
                       if s in cs["score"].values else 0.0 for s in scores)
        return {
            "any_0_0_or_1_0_or_0_1": bundle("0-0","1-0","0-1"),
            "any_1_1_or_2_1_or_1_2": bundle("1-1","2-1","1-2"),
            "any_2_0_or_0_2"        : bundle("2-0","0-2"),
            "any_2_2_or_3_2_or_2_3" : bundle("2-2","3-2","2-3"),
            "any_3_0_or_0_3"        : bundle("3-0","0-3"),
        }

    def score_in_both_halves(self):
        hh = np.random.poisson(self.mh*0.45); ha = np.random.poisson(self.ma*0.45)
        sh2 = np.random.poisson(self.mh*0.55); sa2 = np.random.poisson(self.ma*0.55)
        home_both = float(((hh>0)&(sh2>0)).mean())
        away_both = float(((ha>0)&(sa2>0)).mean())
        return {"home_score_both_halves": home_both,
                "away_score_both_halves": away_both}

    def htft(self):
        """HT/FT combined market — 9 outcomes."""
        hh = np.random.poisson(self.mh*0.45)
        ha = np.random.poisson(self.ma*0.45)
        ht_res = np.where(hh>ha, "H", np.where(hh==ha, "D", "A"))
        ft_res = np.where(self.sh>self.sa, "H", np.where(self.sh==self.sa, "D", "A"))
        combo  = np.char.add(np.char.add(ht_res, "/"), ft_res)
        results = {}
        for label in ["H/H","H/D","H/A","D/H","D/D","D/A","A/H","A/D","A/A"]:
            results[label] = float((combo == label).mean())
        return results

    def time_of_first_goal(self):
        lh = self.mh/90.0; la = self.ma/90.0
        total_rate = lh + la + 1e-9
        no_goal = float((self.tt==0).mean())
        T = np.random.exponential(1.0/total_rate)
        return {
            "0_30" : float((T<=30).mean()) * (1-no_goal),
            "31_60": float(((T>30)&(T<=60)).mean()) * (1-no_goal),
            "61_90": float((T>60).mean()) * (1-no_goal),
            "no_goal": no_goal,
        }

    def xg(self):
        return {"xg_h": float(self.mh.mean()), "xg_a": float(self.ma.mean()),
                "tot" : float((self.mh+self.ma).mean())}

    def corners(self, home_avg=None, away_avg=None, weather_factor=1.0):
        """
        Corners model using team-specific rates when available,
        falling back to PL averages scaled by xG share.
        home_avg / away_avg: team-specific corners per game (from DB).
        """
        xh = float(self.mh.mean()); xa = float(self.ma.mean())
        tot = xh + xa + 1e-9

        if home_avg is not None and away_avg is not None:
            xg_share_h = xh / tot
            xg_share_a = xa / tot
            lh = home_avg * (0.7 + 0.6 * xg_share_h) * weather_factor
            la = away_avg * (0.7 + 0.6 * xg_share_a) * weather_factor
        else:
            lh = PL_CORNERS_H * (xh / tot) * 2 * weather_factor
            la = PL_CORNERS_A * (xa / tot) * 2 * weather_factor

        r = 5.0
        def nb(lam):
            p = r / (r + lam)
            return np.random.negative_binomial(int(r), p, N_SIM)
        ch, ca = nb(lh), nb(la); ct = ch + ca
        out = {
            "mean_h": float(ch.mean()), "mean_a": float(ca.mean()),
            "mean_t": float(ct.mean()),
            "using_team_data": home_avg is not None,
        }
        for l in (7.5, 8.5, 9.5, 10.5, 11.5, 12.5):
            out[f"ov_{l}"] = float((ct > l).mean())
            out[f"un_{l}"] = float((ct <= l).mean())
        ch_ht = np.random.poisson(ch * 0.45); ca_ht = np.random.poisson(ca * 0.45)
        out["mean_ht_corners"] = float((ch_ht + ca_ht).mean())
        return out

    def cards(self, home_avg=None, away_avg=None, referee_factor=1.0):
        """
        Cards model using team-specific yellow card rates when available,
        scaled by referee tendency and game intensity.
        home_avg / away_avg: team-specific cards per game from DB.
        referee_factor: multiplier from referee_card_rate() (1.0 = average).
        """
        intensity = min(float((self.mh + self.ma).mean()) / 2.5, 1.4)

        if home_avg is not None and away_avg is not None:
            lh = home_avg * intensity * referee_factor
            la = away_avg * intensity * referee_factor
        else:
            lh = PL_CARDS_H * intensity * referee_factor
            la = PL_CARDS_A * intensity * 1.10 * referee_factor

        r = 3.0
        def nb(lam):
            p = r / (r + lam)
            return np.random.negative_binomial(int(r), p, N_SIM)
        ch, ca = nb(lh), nb(la); ct = ch + ca

        bp_h = ch * 10 + np.random.binomial(1, 0.05, N_SIM) * 15
        bp_a = ca * 10 + np.random.binomial(1, 0.07, N_SIM) * 15
        bp_t = bp_h + bp_a

        out = {
            "mean_h": float(ch.mean()), "mean_a": float(ca.mean()),
            "mean_t": float(ct.mean()),
            "mean_booking_pts": float(bp_t.mean()),
            "referee_factor": referee_factor,
            "using_team_data": home_avg is not None,
        }
        for l in (1.5, 2.5, 3.5, 4.5, 5.5):
            out[f"ov_{l}"] = float((ct > l).mean())
            out[f"un_{l}"] = float((ct <= l).mean())
        for l in (20.5, 30.5, 40.5, 50.5):
            out[f"bp_ov_{l}"] = float((bp_t > l).mean())
        return out

    def first_goal(self):
        lh = self.mh/90.0; la = self.ma/90.0; tot = lh+la+1e-9
        ng = float((self.tt==0).mean())
        return {"home": float((lh/tot).mean()*(1-ng)),
                "away": float((la/tot).mean()*(1-ng)),
                "no_goal": ng}

    def anytime_goalscorer(self, players_h: list, players_a: list):
        """
        P(player scores at least once) using Poisson race model.

        When player has 'xg_per90' from API stats:
          Uses actual historical goals/90 as the rate — much more accurate.

        When xg_per90 is None (no data):
          Falls back to position-based weights applied to team lambda.

        Position weights (fallback):
          FW=0.35, MF=0.12, DF=0.04, GK=0.01
        """
        pos_weight = {"FW": 0.35, "MF": 0.12, "DF": 0.04, "GK": 0.01}

        def team_probs(players, lam_arr):
            if not players:
                return []
            result = []
            for p in players:
                xg90 = p.get("xg_per90")
                if xg90 is not None and xg90 > 0:
                    # Use actual player xG/90 scaled to match lambda
                    # Player contributes xg90 goals per 90 minutes on average
                    # P(scores >= 1) = 1 - e^(-player_lambda)
                    player_lam = np.full(len(lam_arr), xg90)
                    source = "actual"
                else:
                    # Position-weight fallback
                    w = pos_weight.get(p.get("position", "MF"), 0.12)
                    weights_sum = sum(
                        pos_weight.get(q.get("position", "MF"), 0.12)
                        for q in players
                    )
                    w_norm     = w / max(weights_sum, 1e-9)
                    player_lam = lam_arr * w_norm
                    source     = "pos"

                p_score = float((1 - np.exp(-player_lam)).mean())
                result.append({
                    "name"    : p.get("name", ""),
                    "position": p.get("position", "MF"),
                    "prob"    : p_score,
                    "source"  : source,
                    "xg90"    : xg90 if xg90 is not None else 0.0,
                })
            return sorted(result, key=lambda x: x["prob"], reverse=True)

        return {
            "home": team_probs(players_h, self.mh),
            "away": team_probs(players_a, self.ma),
        }


# ══════════════════════════════════════════════════════════════════════════════
# ODDS & KELLY
# ══════════════════════════════════════════════════════════════════════════════
def implied(odds):   return 1/odds if odds>1 else 0.0
def edge(p, odds):   return p - implied(odds)
def ev(p, odds):     return p*(odds-1) - (1-p)

def kelly_stake(p, odds, bankroll):
    b = odds-1.0
    if b<=0 or p<=0: return 0.0
    f = max(0.0, (b*p-(1-p))/b) * QUARTER_KELLY
    return round(bankroll * min(f, MAX_KELLY_PCT), 0)

def bet_quality(market: str, model_prob: float, edge_val: float) -> str:
    """
    Classify a bet using DUAL conditions — both edge AND model probability
    must meet their respective thresholds.

    Returns one of:
      'premium'    — 🔥  edge >= PREMIUM_EDGE  AND  prob >= PREMIUM_PROB
      'acceptable' — ✅  edge >= MIN_EDGE       AND  prob >= market floor
      'low_prob'   — ⚠️  edge ok but model probability too low (longshot risk)
      'low_edge'   — ⚠️  probability ok but edge too small
      'weak'       — ⚠️  fails both conditions
    """
    min_prob = MIN_PROB_BY_MARKET.get(market, MIN_PROB_DEFAULT)
    prob_ok  = model_prob >= min_prob
    edge_ok  = edge_val   >= MIN_EDGE

    if model_prob >= PREMIUM_PROB and edge_val >= PREMIUM_EDGE:
        return "premium"
    if prob_ok and edge_ok:
        return "acceptable"
    if edge_ok and not prob_ok:
        return "low_prob"      # edge looks good but confidence too low
    if prob_ok and not edge_ok:
        return "low_edge"      # model confident but edge too thin
    return "weak"

def quality_icon(quality: str) -> str:
    return {
        "premium"   : "🔥",
        "acceptable": "✅",
        "low_prob"  : "⚠️",
        "low_edge"  : "⚠️",
        "weak"      : "⚠️",
    }.get(quality, "⚠️")

def should_bet(quality: str) -> bool:
    """Only premium and acceptable qualities warrant placing a bet."""
    return quality in ("premium", "acceptable")

ODDS_PROMPTS_GOALS = [
    ("home_win",   "  1X2    Home Win       "),
    ("draw",       "  1X2    Draw           "),
    ("away_win",   "  1X2    Away Win       "),
    ("over_2.5",   "  O/U    Over 2.5       "),
    ("under_2.5",  "  O/U    Under 2.5      "),
    ("over_1.5",   "  O/U    Over 1.5       "),
    ("over_3.5",   "  O/U    Over 3.5       "),
    ("btts_yes",   "  BTTS   Yes            "),
    ("btts_no",    "  BTTS   No             "),
    ("dc_1x",      "  DC     1X (Home/Draw) "),
    ("dc_x2",      "  DC     X2 (Draw/Away) "),
    ("dnb_home",   "  DNB    Home           "),
    ("dnb_away",   "  DNB    Away           "),
    ("wtn_home",   "  WtN    Home Win to Nil"),
    ("wtn_away",   "  WtN    Away Win to Nil"),
    ("ah_-0.5",    "  AH     Home -0.5      "),
    ("ah_+0.5",    "  AH     Away +0.5      "),
    ("ah_-1.5",    "  AH     Home -1.5      "),
    ("ah_+1.5",    "  AH     Away +1.5      "),
]

ODDS_PROMPTS_CORNERS = [
    ("cor_ov_7.5",  "  COR    Over 7.5       "),
    ("cor_ov_8.5",  "  COR    Over 8.5       "),
    ("cor_ov_9.5",  "  COR    Over 9.5       "),
    ("cor_ov_10.5", "  COR    Over 10.5      "),
    ("cor_un_7.5",  "  COR    Under 7.5      "),
    ("cor_un_8.5",  "  COR    Under 8.5      "),
    ("cor_un_9.5",  "  COR    Under 9.5      "),
    ("cor_un_10.5", "  COR    Under 10.5     "),
]

ODDS_PROMPTS_CARDS = [
    ("crd_ov_1.5",  "  YC     Over 1.5       "),
    ("crd_ov_2.5",  "  YC     Over 2.5       "),
    ("crd_ov_3.5",  "  YC     Over 3.5       "),
    ("crd_ov_4.5",  "  YC     Over 4.5       "),
    ("crd_un_1.5",  "  YC     Under 1.5      "),
    ("crd_un_2.5",  "  YC     Under 2.5      "),
    ("crd_un_3.5",  "  YC     Under 3.5      "),
    ("bp_ov_20.5",  "  BP     Over 20.5 pts  "),
    ("bp_ov_30.5",  "  BP     Over 30.5 pts  "),
    ("bp_ov_40.5",  "  BP     Over 40.5 pts  "),
]

def _collect_odds_section(prompts: list, section_label: str) -> dict:
    odds = {}
    raw = input(f"\n  Enter {section_label} odds? (y/n): ").strip().lower()
    if raw != "y":
        return odds
    for key, label in prompts:
        raw2 = input(label + ": ").strip()
        if raw2:
            try:
                v = float(raw2)
                if v > 1.0:
                    odds[key] = v
            except ValueError:
                pass
    return odds

def get_odds():
    print()
    print("╔" + "═"*56 + "╗")
    print("║" + "  BOOKMAKER ODDS (decimal e.g. 1.85, Enter=skip)  ".center(56) + "║")
    print("║" + "  Each section asks if you want to enter those odds  ".center(56) + "║")
    print("╚" + "═"*56 + "╝")
    odds = {}
    odds.update(_collect_odds_section(ODDS_PROMPTS_GOALS,   "Goals/Result"))
    odds.update(_collect_odds_section(ODDS_PROMPTS_CORNERS, "Corners"))
    odds.update(_collect_odds_section(ODDS_PROMPTS_CARDS,   "Cards/Bookings"))
    return odds


# ══════════════════════════════════════════════════════════════════════════════
# LINEUP INPUT
# ══════════════════════════════════════════════════════════════════════════════
def input_lineup(team_name: str) -> list:
    """
    Ask user to enter starting XI with positions.
    Format per player: Name,POS  (e.g. Saka,FW or White,DF)
    Returns list of dicts.
    """
    print(f"\n  ── Starting XI for {team_name} ──")
    print("  Format: Player Name,POS  (POS = GK/DF/MF/FW)")
    print("  Enter 11 players, one per line. Press Enter to skip lineups.\n")
    players = []
    for i in range(1, 12):
        raw = input(f"    {i:>2}. ").strip()
        if not raw:
            break
        parts = raw.rsplit(",", 1)
        name = parts[0].strip()
        pos  = parts[1].strip().upper() if len(parts) == 2 else "MF"
        if pos not in ("GK","DF","MF","FW"):
            pos = "MF"
        players.append({"name": name, "position": pos})
    return players


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL TRACKER
# ══════════════════════════════════════════════════════════════════════════════
class ExcelTracker:
    """
    Creates and appends to predictions_tracker.xlsx.
    One row per prediction appended automatically after each prediction.
    Manual columns (actual result, bet placed, etc.) left blank for you to fill.
    """

    PRED_HEADERS = [
        # Auto-filled
        "ID", "Date", "Home Team", "Away Team",
        "Home xG", "Away xG", "Total xG",
        "Pred 1X2", "Home Win%", "Draw%", "Away Win%",
        "Over 2.5%", "Under 2.5%", "Over 1.5%", "Over 3.5%",
        "BTTS Yes%", "BTTS No%",
        "Top Correct Score", "Top CS Prob%",
        "HT Home Win%", "HT Draw%", "HT Away Win%",
        "Half Most Goals", "HT/FT Most Likely",
        "Win To Nil Home%", "Win To Nil Away%",
        "Draw No Bet Home%", "Draw No Bet Away%",
        "Exp Corners Total", "Exp Cards Total", "Exp Booking Pts",
        "Best Market", "Edge%", "EV", "Kelly Stake (RWF)", "Bankroll (RWF)",
        "Odds Entered",
        "H2H Factor", "Inj Factor Home", "Inj Factor Away",
        "Momentum Home", "Momentum Away",
        "Injury Notes Home", "Injury Notes Away",
        # Manual columns
        "ACTUAL Score", "ACTUAL 1X2 (H/D/A)", "ACTUAL Goals",
        "ACTUAL BTTS (Y/N)", "ACTUAL HT Score",
        "ACTUAL Corners", "ACTUAL Cards",
        "Bet Placed? (Y/N)", "Market Bet On",
        "Stake Placed (RWF)", "Odds Taken", "Result (W/L/P)",
        "Profit/Loss (RWF)",
        # Formula columns
        "1X2 Correct?", "O/U 2.5 Correct?", "BTTS Correct?",
        "Corners Correct?", "Cards Correct?",
        "Closing Odds (fill after)", "CLV (Closing Line Value)",
        "Bankroll After (RWF)", "Running ROI%",
    ]

    PERF_HEADERS = [
        "Metric", "Value"
    ]

    # Column indices (1-based) for formula references
    COL_PRED_1X2      = 8   # H
    COL_HOME_WIN_PCT  = 9   # I
    COL_ACTUAL_1X2    = 45  # AS (ACTUAL 1X2)
    COL_ACTUAL_GOALS  = 46  # AT
    COL_OVER25_PCT    = 12  # L
    COL_BET_PLACED    = 51  # AY
    COL_STAKE         = 53  # BA
    COL_RESULT        = 55  # BC
    COL_PL            = 56  # BD  Profit/Loss
    COL_1X2_CORRECT   = 57  # BE
    COL_OU_CORRECT    = 58  # BF
    COL_BTTS_CORRECT  = 59  # BG
    COL_BANKROLL_AFT  = 60  # BH
    COL_ROI           = 61  # BI

    # Styling
    HDR_FILL_AUTO   = PatternFill("solid", fgColor="1F3864")   # dark blue
    HDR_FILL_MANUAL = PatternFill("solid", fgColor="833C00")   # dark orange
    HDR_FILL_CALC   = PatternFill("solid", fgColor="375623")   # dark green
    HDR_FONT        = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    DATA_FONT       = Font(name="Arial", size=9)
    PERF_HDR_FILL   = PatternFill("solid", fgColor="2E4057")
    BORDER          = Border(
        left  = Side(style="thin", color="CCCCCC"),
        right = Side(style="thin", color="CCCCCC"),
        top   = Side(style="thin", color="CCCCCC"),
        bottom= Side(style="thin", color="CCCCCC"),
    )

    def __init__(self, path: Path):
        self.path = path
        if not path.exists():
            self._create()

    def _create(self):
        wb = openpyxl.Workbook()

        # ── Predictions sheet ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Predictions"
        ws.freeze_panes = "E2"

        n_auto   = 44   # columns A–AR  (auto-filled by script)
        n_manual = 13   # columns AS–BE (you fill in)
        n_calc   = 5    # columns BF–BJ (Excel formulas)

        for col_idx, header in enumerate(self.PRED_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = self.HDR_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = self.BORDER

            if col_idx <= n_auto:
                cell.fill = self.HDR_FILL_AUTO
            elif col_idx <= n_auto + n_manual:
                cell.fill = self.HDR_FILL_MANUAL
            else:
                cell.fill = self.HDR_FILL_CALC

        # Column widths
        col_widths = {
            1: 6, 2: 12, 3: 22, 4: 22,
            5: 9, 6: 9, 7: 9,
            8: 12, 9: 9, 10: 9, 11: 9,
            12: 10, 13: 10, 14: 10, 15: 10,
            16: 10, 17: 10,
            18: 14, 19: 11,
            20: 13, 21: 10, 22: 13,
            23: 14, 24: 14,
            25: 14, 26: 14, 27: 14, 28: 14,
            29: 12, 30: 12, 31: 12,
            32: 16, 33: 8, 34: 8, 35: 15, 36: 15,
            37: 16,
            38: 10, 39: 12, 40: 12, 41: 11, 42: 11,
            43: 22, 44: 22,
        }
        for c, w in col_widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w
        # Manual + calc columns
        for c in range(45, len(self.PRED_HEADERS)+2):
            ws.column_dimensions[get_column_letter(c)].width = 14
        ws.row_dimensions[1].height = 36

        # ── Performance sheet ──────────────────────────────────────────────────
        wp = wb.create_sheet("Performance")
        wp.freeze_panes = "A2"

        perf_metrics = [
            ("Total Predictions",   f"=COUNTA(Predictions!B2:B10000)"),
            ("Bets Placed",         f"=COUNTIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\")"),
            ("Bets Won",            f"=COUNTIF(Predictions!{get_column_letter(56)}2:{get_column_letter(56)}10000,\"W\")"),
            ("Bets Lost",           f"=COUNTIF(Predictions!{get_column_letter(56)}2:{get_column_letter(56)}10000,\"L\")"),
            ("Strike Rate %",       f"=IFERROR(B4/B3*100,0)"),
            ("Total Staked (RWF)",  f"=IFERROR(SUMIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!{get_column_letter(54)}2:{get_column_letter(54)}10000),0)"),
            ("Total Return (RWF)",  f"=IFERROR(SUMIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!{get_column_letter(57)}2:{get_column_letter(57)}10000)+B7,0)"),
            ("Net Profit (RWF)",    f"=B8-B7"),
            ("ROI %",               f"=IFERROR(B9/B7*100,0)"),
            ("1X2 Accuracy %",      f"=IFERROR(COUNTIF(Predictions!{get_column_letter(58)}2:{get_column_letter(58)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(58)}2:{get_column_letter(58)}10000)*100,0)"),
            ("O/U 2.5 Accuracy %",  f"=IFERROR(COUNTIF(Predictions!{get_column_letter(59)}2:{get_column_letter(59)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(59)}2:{get_column_letter(59)}10000)*100,0)"),
            ("BTTS Accuracy %",     f"=IFERROR(COUNTIF(Predictions!{get_column_letter(60)}2:{get_column_letter(60)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(60)}2:{get_column_letter(60)}10000)*100,0)"),
            ("Corners Accuracy %",  f"=IFERROR(COUNTIF(Predictions!{get_column_letter(61)}2:{get_column_letter(61)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(61)}2:{get_column_letter(61)}10000)*100,0)"),
            ("Cards Accuracy %",    f"=IFERROR(COUNTIF(Predictions!{get_column_letter(62)}2:{get_column_letter(62)}10000,\"✓\")/COUNTA(Predictions!{get_column_letter(62)}2:{get_column_letter(62)}10000)*100,0)"),
            ("Avg CLV % (bets)",    f"=IFERROR(AVERAGEIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!{get_column_letter(64)}2:{get_column_letter(64)}10000),0)"),
            ("Avg Edge % (bets)",   f"=IFERROR(AVERAGEIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!AH2:AH10000),0)"),
            ("Avg EV (bets)",       f"=IFERROR(AVERAGEIF(Predictions!{get_column_letter(52)}2:{get_column_letter(52)}10000,\"Y\",Predictions!AI2:AI10000),0)"),
            ("Current Bankroll",    f"=IFERROR(INDEX(Predictions!{get_column_letter(65)}2:{get_column_letter(65)}10000,MATCH(9E+307,Predictions!{get_column_letter(65)}2:{get_column_letter(65)}10000)),0)"),
        ]

        wp["A1"] = "Metric"; wp["B1"] = "Value"
        for cell in [wp["A1"], wp["B1"]]:
            cell.fill = self.PERF_HDR_FILL
            cell.font = self.HDR_FONT
            cell.alignment = Alignment(horizontal="center")

        for r, (metric, formula) in enumerate(perf_metrics, start=2):
            wp.cell(row=r, column=1, value=metric).font = Font(bold=True, name="Arial", size=10)
            cell = wp.cell(row=r, column=2, value=formula)
            cell.font = Font(name="Arial", size=10)
            cell.number_format = "#,##0.00"

        wp.column_dimensions["A"].width = 26
        wp.column_dimensions["B"].width = 20

        # ── Legend sheet ───────────────────────────────────────────────────────
        wl = wb.create_sheet("Legend")
        legend_rows = [
            ("COLUMN COLOR GUIDE", ""),
            ("Dark Blue headers", "Auto-filled by prediction script"),
            ("Dark Orange headers", "You fill in manually after the match"),
            ("Dark Green headers", "Calculated automatically by Excel formulas"),
            ("", ""),
            ("MANUAL COLUMN GUIDE", ""),
            ("ACTUAL Score", "e.g. 2-1"),
            ("ACTUAL 1X2", "H = Home win, D = Draw, A = Away win"),
            ("ACTUAL Goals", "Total goals scored"),
            ("ACTUAL BTTS", "Y if both teams scored, N if not"),
            ("ACTUAL HT Score", "e.g. 1-0"),
            ("Bet Placed?", "Y or N"),
            ("Market Bet On", "e.g. over_2.5, home_win, btts_yes"),
            ("Stake Placed (RWF)", "Exact amount you staked"),
            ("Odds Taken", "Decimal odds you got"),
            ("Result (W/L/P)", "W=Win, L=Loss, P=Push/Void"),
            ("Profit/Loss (RWF)", "Positive if won, negative if lost"),
            ("", ""),
            ("POSITION CODES (for lineup entry)", ""),
            ("GK", "Goalkeeper"),
            ("DF", "Defender"),
            ("MF", "Midfielder"),
            ("FW", "Forward"),
        ]
        wl["A1"] = "PREDICTIONS TRACKER — GUIDE"
        wl["A1"].font = Font(bold=True, size=14, name="Arial")
        for r, (col, desc) in enumerate(legend_rows, start=3):
            wl.cell(row=r, column=1, value=col).font  = Font(bold=bool(col and "GUIDE" in col or "CODES" in col), name="Arial")
            wl.cell(row=r, column=2, value=desc).font = Font(name="Arial")
        wl.column_dimensions["A"].width = 35
        wl.column_dimensions["B"].width = 50

        wb.save(self.path)
        print(f"[Tracker] Created {self.path.name} with Predictions, Performance, Legend sheets.")

    def append_prediction(self, data: dict):
        """Append one prediction row to the Predictions sheet."""
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Predictions"]
        next_row = ws.max_row + 1
        pred_id  = next_row - 1

        # Determine predicted 1X2
        probs_1x2 = {
            "Home Win": data.get("home_win_pct", 0),
            "Draw"    : data.get("draw_pct", 0),
            "Away Win": data.get("away_win_pct", 0),
        }
        pred_1x2 = max(probs_1x2, key=probs_1x2.get)
        half_most = max(
            {"1st Half": data.get("first_half_pct",0),
             "2nd Half": data.get("second_half_pct",0),
             "Equal"   : data.get("equal_pct",0)},
            key=lambda k: {"1st Half": data.get("first_half_pct",0),
                           "2nd Half": data.get("second_half_pct",0),
                           "Equal"   : data.get("equal_pct",0)}[k]
        )

        # Build the HTFT most likely label
        htft = data.get("htft", {})
        htft_best = max(htft, key=htft.get) if htft else "D/H"

        auto_values = [
            pred_id,
            data.get("date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            data.get("home_team", ""),
            data.get("away_team", ""),
            round(data.get("xg_h", 0), 2),
            round(data.get("xg_a", 0), 2),
            round(data.get("xg_h", 0) + data.get("xg_a", 0), 2),
            pred_1x2,
            round(data.get("home_win_pct", 0)*100, 1),
            round(data.get("draw_pct", 0)*100, 1),
            round(data.get("away_win_pct", 0)*100, 1),
            round(data.get("over25_pct", 0)*100, 1),
            round(data.get("under25_pct", 0)*100, 1),
            round(data.get("over15_pct", 0)*100, 1),
            round(data.get("over35_pct", 0)*100, 1),
            round(data.get("btts_yes_pct", 0)*100, 1),
            round(data.get("btts_no_pct", 0)*100, 1),
            data.get("top_correct_score", ""),
            round(data.get("top_cs_pct", 0)*100, 1),
            round(data.get("ht_hw_pct", 0)*100, 1),
            round(data.get("ht_d_pct", 0)*100, 1),
            round(data.get("ht_aw_pct", 0)*100, 1),
            half_most,
            htft_best,
            round(data.get("home_wtn_pct", 0)*100, 1),
            round(data.get("away_wtn_pct", 0)*100, 1),
            round(data.get("dnb_home_pct", 0)*100, 1),
            round(data.get("dnb_away_pct", 0)*100, 1),
            round(data.get("exp_corners", 0), 1),
            round(data.get("exp_cards", 0), 1),
            round(data.get("exp_booking_pts", 0), 1),
            data.get("best_market", ""),
            round(data.get("edge_pct", 0)*100, 2),
            round(data.get("ev", 0), 3),
            data.get("kelly_stake_rwf", 0),
            data.get("bankroll_rwf", 0),
            data.get("odds_entered", ""),
            round(data.get("h2h_factor", 1.0), 3),
            round(data.get("inj_h_factor", 1.0), 3),
            round(data.get("inj_a_factor", 1.0), 3),
            round(data.get("mom_h", 1.0), 3),
            round(data.get("mom_a", 1.0), 3),
            data.get("injury_notes_home", ""),
            data.get("injury_notes_away", ""),
        ]

        # Write auto values
        for col_idx, val in enumerate(auto_values, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font   = self.DATA_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=False)

        # Skip manual columns (45–57) — leave blank
        n_auto   = len(auto_values)
        n_manual = 13
        for col_idx in range(n_auto+1, n_auto+n_manual+1):
            cell = ws.cell(row=next_row, column=col_idx, value="")
            cell.fill   = PatternFill("solid", fgColor="FFF2CC")
            cell.border = self.BORDER

        # Column layout (1-based):
        # Auto cols 1-44 (same as before)
        # Manual cols 45-57:
        #   45=ACTUAL Score, 46=ACTUAL 1X2, 47=ACTUAL Goals, 48=ACTUAL BTTS
        #   49=ACTUAL HT Score, 50=ACTUAL Corners, 51=ACTUAL Cards
        #   52=Bet Placed?, 53=Market Bet On, 54=Stake Placed
        #   55=Odds Taken, 56=Result, 57=Profit/Loss
        # Formula cols 58-64:
        #   58=1X2 Correct?, 59=O/U Correct?, 60=BTTS Correct?
        #   61=Corners Correct?, 62=Cards Correct?
        #   63=Closing Odds (manual fill), 64=CLV
        #   65=Bankroll After, 66=Running ROI%

        r   = next_row
        ac  = get_column_letter

        # Named column references
        c_pred_1x2    = ac(8)    # H  - Predicted 1X2
        c_over25      = ac(12)   # L  - Over 2.5%
        c_btts_yes    = ac(16)   # P  - BTTS Yes%
        c_exp_corners = ac(29)   # AC - Expected corners
        c_exp_cards   = ac(30)   # AD - Expected cards
        c_best_market = ac(32)   # AF - Best Market
        c_bankroll    = ac(36)   # AJ - Bankroll at prediction time
        c_act_score   = ac(45)   # AS - ACTUAL Score
        c_act_1x2     = ac(46)   # AT - ACTUAL 1X2
        c_act_goals   = ac(47)   # AU - ACTUAL Goals
        c_act_btts    = ac(48)   # AV - ACTUAL BTTS
        c_act_ht      = ac(49)   # AW - ACTUAL HT Score
        c_act_corners = ac(50)   # AX - ACTUAL Corners
        c_act_cards   = ac(51)   # AY - ACTUAL Cards
        c_bet_placed  = ac(52)   # AZ - Bet Placed?
        c_market      = ac(53)   # BA - Market Bet On
        c_stake       = ac(54)   # BB - Stake Placed
        c_odds_taken  = ac(55)   # BC - Odds Taken
        c_result      = ac(56)   # BD - Result W/L/P
        c_pl          = ac(57)   # BE - Profit/Loss
        c_closing     = ac(63)   # BK - Closing Odds (manual)

        formula_cols = [
            # 58: 1X2 Correct?
            (58, f'=IF({c_act_1x2}{r}="","",'
                 f'IF(AND({c_pred_1x2}{r}="Home Win",{c_act_1x2}{r}="H"),"✓",'
                 f'IF(AND({c_pred_1x2}{r}="Draw",{c_act_1x2}{r}="D"),"✓",'
                 f'IF(AND({c_pred_1x2}{r}="Away Win",{c_act_1x2}{r}="A"),"✓","✗"))))'),

            # 59: O/U 2.5 Correct?
            (59, f'=IF({c_act_goals}{r}="","",IF(AND({c_over25}{r}>50,{c_act_goals}{r}>2),"✓",'
                 f'IF(AND({c_over25}{r}<=50,{c_act_goals}{r}<=2),"✓","✗")))'),

            # 60: BTTS Correct?
            (60, f'=IF({c_act_btts}{r}="","",IF(AND({c_btts_yes}{r}>50,{c_act_btts}{r}="Y"),"✓",'
                 f'IF(AND({c_btts_yes}{r}<=50,{c_act_btts}{r}="N"),"✓","✗")))'),

            # 61: Corners Correct? (model over/under vs actual)
            (61, f'=IF({c_act_corners}{r}="","",IF(AND({c_exp_corners}{r}>9.5,{c_act_corners}{r}>9),"✓",'
                 f'IF(AND({c_exp_corners}{r}<=9.5,{c_act_corners}{r}<=9),"✓","✗")))'),

            # 62: Cards Correct?
            (62, f'=IF({c_act_cards}{r}="","",IF(AND({c_exp_cards}{r}>3.5,{c_act_cards}{r}>3),"✓",'
                 f'IF(AND({c_exp_cards}{r}<=3.5,{c_act_cards}{r}<=3),"✓","✗")))'),

            # 63: Closing Odds — MANUAL fill (left blank, formula-green styled as reminder)
            (63, ""),

            # 64: CLV — Closing Line Value = (Odds Taken / Closing Odds - 1) * 100
            # Positive CLV means you beat the closing line → genuine edge
            (64, f'=IFERROR(({c_odds_taken}{r}/{c_closing}{r}-1)*100,"")'),

            # 65: Bankroll After
            (65, f'=IF({c_pl}{r}="",{c_bankroll}{r},'
                 f'{c_bankroll}{r}+IF({c_pl}{r}="",0,{c_pl}{r}))'),

            # 66: Running ROI%
            (66, f'=IFERROR(SUMIF({c_bet_placed}$2:{c_bet_placed}{r},"Y",'
                 f'{c_pl}$2:{c_pl}{r})/SUMIF({c_bet_placed}$2:{c_bet_placed}{r},"Y",'
                 f'{c_stake}$2:{c_stake}{r})*100,0)'),
        ]

        for col_idx, formula in formula_cols:
            cell = ws.cell(row=next_row, column=col_idx, value=formula)
            if col_idx == 63:
                # Closing Odds — manual, styled like manual columns but green border
                cell.fill   = PatternFill("solid", fgColor="FFF2CC")
            else:
                cell.fill   = PatternFill("solid", fgColor="E2EFDA")
                cell.font   = Font(name="Arial", size=9, color="006100")
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        wb.save(self.path)
        print(f"[Tracker] Row {pred_id} appended → {self.path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# TICKET PRINTER
# ══════════════════════════════════════════════════════════════════════════════
W = 74

def bar(p, w=24):
    f = int(min(max(float(p),0),1)*w)
    return "█"*f + "░"*(w-f)

def row(text):   return f"║ {str(text):<{W-2}} ║"
def sep(l="╠",c="═",r="╣"): return f"{l}{c*W}{r}"
def hdr(text):   return f"║ {'【 '+text+' 】':^{W-2}} ║"

def pct(v): return f"{v*100:.1f}%"

def print_ticket(home, away, mdl, s, odds, bankroll, injuries_h, injuries_a,
                 players_h, players_a, h2h_df,
                 home_rates=None, away_rates=None, referee_factor=1.0,
                 weather=None, standings=None, fixture_info=None):

    res  = s.result();     ou   = s.ou();      bt   = s.btts()
    cs   = s.correct_score(); ht = s.halftime(); sht = s.second_half()
    hmg  = s.half_most_goals(); dc = s.dc();   dnb  = s.draw_no_bet()
    wtn  = s.win_to_nil(); htft = s.htft();    xg   = s.xg()
    # Use team-specific rates for corners and cards
    home_cor = home_rates.get("avg_corners") if home_rates else None
    away_cor = away_rates.get("avg_corners") if away_rates else None
    home_crd = home_rates.get("avg_cards")   if home_rates else None
    away_crd = away_rates.get("avg_cards")   if away_rates else None
    w_goal   = float((weather or {}).get("goal_factor",   1.0))
    w_corner = float((weather or {}).get("corner_factor", 1.0))
    cor  = s.corners(home_avg=home_cor, away_avg=away_cor,
                     weather_factor=w_corner)
    crd  = s.cards(home_avg=home_crd, away_avg=away_crd,
                   referee_factor=referee_factor)
    fg   = s.first_goal()
    eg   = s.exact_goals(); ms  = s.multiscores()
    sibh = s.score_in_both_halves()
    tfg  = s.time_of_first_goal()
    str_ = mdl.strengths()
    hf   = mdl.form(home); af = mdl.form(away)

    # AH probabilities for common lines
    ah_m05 = s.asian_handicap(-0.5)
    ah_p05 = s.asian_handicap(+0.5)
    ah_m15 = s.asian_handicap(-1.5)
    ah_p15 = s.asian_handicap(+1.5)

    # Full market probability map — covers ALL odds inputs
    mkt_probs = {
        # 1X2
        "home_win"  : res["home_win"],
        "draw"      : res["draw"],
        "away_win"  : res["away_win"],
        # Over/Under
        "over_2.5"  : ou["over_2.5"],
        "under_2.5" : ou["under_2.5"],
        "over_1.5"  : ou["over_1.5"],
        "over_3.5"  : ou["over_3.5"],
        # BTTS
        "btts_yes"  : bt["btts_yes"],
        "btts_no"   : bt["btts_no"],
        # Double Chance
        "dc_1x"     : dc["1X"],
        "dc_x2"     : dc["X2"],
        # Draw No Bet
        "dnb_home"  : dnb["dnb_home"],
        "dnb_away"  : dnb["dnb_away"],
        # Win to Nil
        "wtn_home"  : wtn["home_wtn"],
        "wtn_away"  : wtn["away_wtn"],
        # Asian Handicap
        "ah_-0.5"   : ah_m05["home"],
        "ah_+0.5"   : ah_p05["away"],
        "ah_-1.5"   : ah_m15["home"],
        "ah_+1.5"   : ah_p15["away"],
        # Corners
        "cor_ov_7.5" : cor.get("ov_7.5", 0),
        "cor_ov_8.5" : cor.get("ov_8.5", 0),
        "cor_ov_9.5" : cor.get("ov_9.5", 0),
        "cor_ov_10.5": cor.get("ov_10.5", 0),
        "cor_un_7.5" : cor.get("un_7.5", 0),
        "cor_un_8.5" : cor.get("un_8.5", 0),
        "cor_un_9.5" : cor.get("un_9.5", 0),
        "cor_un_10.5": cor.get("un_10.5", 0),
        # Cards
        "crd_ov_1.5" : crd.get("ov_1.5", 0),
        "crd_ov_2.5" : crd.get("ov_2.5", 0),
        "crd_ov_3.5" : crd.get("ov_3.5", 0),
        "crd_ov_4.5" : crd.get("ov_4.5", 0),
        "crd_un_1.5" : crd.get("un_1.5", 0),
        "crd_un_2.5" : crd.get("un_2.5", 0),
        "crd_un_3.5" : crd.get("un_3.5", 0),
        # Booking Points
        "bp_ov_20.5" : crd.get("bp_ov_20.5", 0),
        "bp_ov_30.5" : crd.get("bp_ov_30.5", 0),
        "bp_ov_40.5" : crd.get("bp_ov_40.5", 0),
    }

    print(); print(f"╔{'═'*W}╗")
    print(row("")); print(row("  ⚽  BAYESIAN FOOTBALL INTELLIGENCE  v8.0  ⚽"))
    print(row("  Auto Odds · Auto Lineups · Weather · Fixture Lookup"))
    print(row("")); print(sep())
    print(row(f"  🏟   {home}  vs  {away}"))
    # Fixture info line
    if fixture_info and fixture_info.get("fixture_id"):
        fi     = fixture_info
        dt_str = fi.get("date", "")[:16].replace("T", " ") if fi.get("date") else "TBD"
        venue  = fi.get("venue", "Unknown venue")
        ref    = fi.get("referee", "") or "TBD"
        rnd    = fi.get("round", "")
        print(row(f"  📅   {dt_str}   {rnd}"))
        print(row(f"  🏟   {venue}   Referee: {ref}"))
    else:
        print(row(f"  📅   {datetime.now().strftime('%A, %d %B %Y  %H:%M')}"))
    print(row(f"  💰   Bankroll: {bankroll:,.0f} RWF   Sims: {N_SIM:,}"))
    # Weather line
    if weather and weather.get("description"):
        wf_g = weather.get("goal_factor", 1.0)
        wf_c = weather.get("corner_factor", 1.0)
        effect = ""
        if wf_g < 0.99:
            effect += f"  goals ×{wf_g:.2f}"
        if wf_c != 1.0:
            effect += f"  corners ×{wf_c:.2f}"
        print(row(f"  🌤   {weather['description']}{effect}"))
    # Standings line
    if standings and "home" in standings and "away" in standings:
        hs = standings["home"]; as_ = standings["away"]
        print(row(f"  📊  {hs['team'][:20]}: #{hs['pos']} {hs['pts']}pts  |  "
                  f"{as_['team'][:20]}: #{as_['pos']} {as_['pts']}pts"))
    print(sep())

    # Strengths
    print(hdr("TEAM STRENGTH  (posterior MAP)"))
    print(sep("╠","─","╣"))
    print(row(f"  {'Team':<26}  {'Attack':>8}  {'Defense':>8}  {'Net':>8}  {'xG Avg':>7}"))
    print(sep("╠","─","╣"))
    for team, f_ in [(home, hf), (away, af)]:
        atk, dfn = str_[team]
        net = atk-dfn
        print(row(f"  {team[:26]:<26}  {atk:>+8.3f}  {dfn:>+8.3f}  {net:>+8.3f}  {f_['avg_xg']:>7.2f}"))
    print(sep())

    # Form
    print(hdr("RECENT FORM  (last 5)"))
    print(sep("╠","─","╣"))
    for team, f_ in [(home, hf), (away, af)]:
        print(row(f"  {team[:26]:<26}  {f_['form']}  "
                  f"Pts:{f_['pts']}  GF:{f_['gf']}  GA:{f_['ga']}  "
                  f"GD:{f_['gd']:+d}  xG:{f_['avg_xg']:.2f}"))
    print(sep())

    # Adjustment factors
    print(hdr("MODEL ADJUSTMENTS  (v8.0)"))
    print(sep("╠","─","╣"))
    print(row(f"  H2H Factor: {mdl.h2h_factor:.3f}   "
              f"Momentum Home: ×{mdl.mom_h:.3f}   Away: ×{mdl.mom_a:.3f}"))
    print(row(f"  Injury — Home: ×{mdl.inj_h_factor:.3f}   Away: ×{mdl.inj_a_factor:.3f}   "
              f"Fatigue — Home: ×{mdl.fatigue_h:.3f}   Away: ×{mdl.fatigue_a:.3f}"))
    print(row(f"  Dixon-Coles ρ: {mdl.rho:+.3f}   "
              f"({'lowers' if mdl.rho < 0 else 'raises'} low-score probabilities)"))
    print(row(f"  Weather goal ×{mdl.weather_goal_factor:.3f}   "
              f"corner ×{mdl.weather_corner_factor:.3f}"))
    cor_src = "team data" if cor.get("using_team_data") else "PL averages"
    crd_src = "team data" if crd.get("using_team_data") else "PL averages"
    ref_str = f"×{referee_factor:.2f}" if referee_factor != 1.0 else "average"
    print(row(f"  Corners source: {cor_src}   Cards source: {crd_src}   "
              f"Referee: {ref_str}"))
    print(sep())

    # Injuries
    if injuries_h or injuries_a:
        print(hdr("INJURY REPORT"))
        print(sep("╠","─","╣"))
        for label, injs in [(f"  {home}", injuries_h), (f"  {away}", injuries_a)]:
            if injs:
                for inj in injs:
                    print(row(f"  ❌ {label.strip()[:20]:<20}  "
                              f"{inj.get('player_name',''):<20}  "
                              f"{inj.get('injury_type','')}  "
                              f"Return: {inj.get('expected_return','TBD')}"))
        print(sep())

    # xG
    print(hdr("EXPECTED GOALS  (xG)"))
    print(sep("╠","─","╣"))
    print(row(f"  {home[:22]:<22}  xG {xg['xg_h']:.2f}   {bar(xg['xg_h']/4)}"))
    print(row(f"  {away[:22]:<22}  xG {xg['xg_a']:.2f}   {bar(xg['xg_a']/4)}"))
    tone = "High-scoring" if xg["tot"]>3 else "Moderate" if xg["tot"]>2 else "Low-scoring"
    print(row(f"  Total xG: {xg['tot']:.2f}   [{tone}]"))
    print(sep())

    # H2H
    if not h2h_df.empty and len(h2h_df) >= 3:
        print(hdr("HEAD-TO-HEAD  (last meetings)"))
        print(sep("╠","─","╣"))
        for _, hr in h2h_df.head(6).iterrows():
            result_tag = ("H" if hr["home_goals"]>hr["away_goals"]
                          else "D" if hr["home_goals"]==hr["away_goals"] else "A")
            date_str = str(hr.get("match_date",""))[:10]
            print(row(f"  {date_str}  {hr['home_team'][:18]:<18} "
                      f"{hr['home_goals']}-{hr['away_goals']}  "
                      f"{hr['away_team'][:18]:<18}  [{result_tag}]"))
        print(sep())

    # 1X2
    print(hdr("MATCH RESULT  (1X2)"))
    print(sep("╠","─","╣"))
    for lbl, p in [(f"  1  Home  ({home[:18]})", res["home_win"]),
                   (f"  X  Draw               ", res["draw"]),
                   (f"  2  Away  ({away[:18]})", res["away_win"])]:
        arrow = "▶" if p==max(res.values()) else " "
        print(row(f"{arrow} {lbl[:42]:<42} {pct(p):>7}  {bar(p)}"))
    print(sep("╠","─","╣"))
    print(row(f"  Double Chance  1X:{pct(dc['1X'])}  X2:{pct(dc['X2'])}  12:{pct(dc['12'])}"))
    print(row(f"  Draw No Bet    Home:{pct(dnb['dnb_home'])}   Away:{pct(dnb['dnb_away'])}"))
    print(row(f"  Win To Nil     Home:{pct(wtn['home_wtn'])}   Away:{pct(wtn['away_wtn'])}"))
    print(sep())

    # Goals
    print(hdr("GOALS MARKETS"))
    print(sep("╠","─","╣"))
    for l in (0.5,1.5,2.5,3.5,4.5,5.5):
        ov,un = ou[f"over_{l}"],ou[f"under_{l}"]
        fo = "←" if ov>=un else " "; fu = "←" if un>ov else " "
        print(row(f"  O/U {l}   Over:{pct(ov)} {fo}   Under:{pct(un)} {fu}"))
    print(sep("╠","─","╣"))
    print(row(f"  BTTS Yes:{pct(bt['btts_yes'])}   BTTS No:{pct(bt['btts_no'])}"))
    print(sep("╠","─","╣"))
    print(row(f"  Exact Goals: " + "  ".join(f"{n}:{pct(eg[f'exactly_{n}'])}"
                                              for n in range(6))))
    print(row(f"  Exactly 6+: {pct(eg['exactly_7plus'])}"))
    print(sep())

    # Half analysis
    print(hdr("HALF-TIME / SECOND-HALF ANALYSIS"))
    print(sep("╠","─","╣"))
    print(row(f"  HT Result   Home:{pct(ht['ht_hw'])}  Draw:{pct(ht['ht_d'])}  Away:{pct(ht['ht_aw'])}"))
    print(row(f"  2H Result   Home:{pct(sht['sh_hw'])}  Draw:{pct(sht['sh_d'])}  Away:{pct(sht['sh_aw'])}"))
    print(row(f"  HT O/U 0.5: Over {pct(ht['ht_ou_05'])}   HT O/U 1.5: Over {pct(ht['ht_ou_15'])}"))
    print(row(f"  2H O/U 0.5: Over {pct(sht['sh_ou_05'])}   2H O/U 1.5: Over {pct(sht['sh_ou_15'])}"))
    print(row(f"  Half Most Goals: 1st {pct(hmg['first_half'])}  "
              f"2nd {pct(hmg['second_half'])}  Equal {pct(hmg['equal'])}"))
    print(sep())

    # Score in both halves
    print(hdr("SCORE IN BOTH HALVES"))
    print(sep("╠","─","╣"))
    print(row(f"  {home[:28]:<28} scores in both halves: {pct(sibh['home_score_both_halves'])}"))
    print(row(f"  {away[:28]:<28} scores in both halves: {pct(sibh['away_score_both_halves'])}"))
    print(sep())

    # 3-way handicap
    print(hdr("3-WAY HANDICAP  (home team perspective)"))
    print(sep("╠","─","╣"))
    print(row(f"  {'HCP':<8}  {'Half':<5}  {'Home':>7}  {'Draw':>7}  {'Away':>7}"))
    print(sep("╠","─","╣"))
    for hcp in (-2,-1,0,1,2):
        for half in ("FT","HT","2H"):
            d = s.three_way_handicap(hcp, half)
            print(row(f"  {hcp:>+3}      {half:<5}  {pct(d['home']):>7}  "
                      f"{pct(d['draw']):>7}  {pct(d['away']):>7}"))
    print(sep())

    # Asian handicap
    print(hdr("ASIAN HANDICAP  (home team perspective)"))
    print(sep("╠","─","╣"))
    for hcp in (-1.5,-1.0,-0.75,-0.5,-0.25,0.0,0.25,0.5,1.0,1.5):
        d = s.asian_handicap(hcp)
        push = f"  Push:{pct(d['push'])}" if d["push"]>0.01 else ""
        print(row(f"  AH {hcp:+.2f}   Home:{pct(d['home'])}   Away:{pct(d['away'])}{push}"))
    print(sep())

    # HT/FT
    print(hdr("HT / FT COMBINED"))
    print(sep("╠","─","╣"))
    htft_sorted = sorted(htft.items(), key=lambda x: x[1], reverse=True)
    for label, prob in htft_sorted:
        arrow = "▶" if label==htft_sorted[0][0] else " "
        print(row(f"{arrow} {label:<8}  {pct(prob):>7}  {bar(prob, 20)}"))
    print(sep())

    # Correct Score
    print(hdr("CORRECT SCORE  (Top 15)"))
    print(sep("╠","─","╣"))
    for _, r2 in cs[cs["score"]!="other"].head(15).iterrows():
        h_s,a_s = r2["score"].split("-")
        tag = "H" if int(h_s)>int(a_s) else "D" if int(h_s)==int(a_s) else "A"
        print(row(f"  [{tag}]  {r2['score']:>5}   {pct(r2['prob']):>7}   {bar(r2['prob']*5, 20)}"))
    oth = cs[cs["score"]=="other"]["prob"].values
    if len(oth): print(row(f"  [other combined]: {pct(oth[0])}"))
    print(sep())

    # Multiscores
    print(hdr("MULTISCORES  (bundled)"))
    print(sep("╠","─","╣"))
    for label, prob in sorted(ms.items(), key=lambda x: x[1], reverse=True):
        print(row(f"  {label:<40}  {pct(prob):>7}"))
    print(sep())

    # First goal
    print(hdr("FIRST GOAL & TIMING"))
    print(sep("╠","─","╣"))
    print(row(f"  {home[:30]:<30} scores first:  {pct(fg['home'])}"))
    print(row(f"  {away[:30]:<30} scores first:  {pct(fg['away'])}"))
    print(row(f"  No goals (0-0):                                {pct(fg['no_goal'])}"))
    print(sep("╠","─","╣"))
    print(row(f"  First goal time:  0-30 min:{pct(tfg['0_30'])}  "
              f"31-60:{pct(tfg['31_60'])}  61-90:{pct(tfg['61_90'])}"))
    print(sep())

    # Anytime goalscorer
    if players_h or players_a:
        print(hdr("ANYTIME GOALSCORER"))
        print(sep("╠","─","╣"))
        gs = s.anytime_goalscorer(players_h, players_a)
        print(row(f"  {home[:26]:<26}  {'Pos':<4}  {'Prob':>7}  {'xG/90':>7}  Src"))
        print(sep("╠","─","╣"))
        for p in gs["home"][:8]:
            xg_str = f"{p['xg90']:.3f}" if p.get("xg90") else "  —  "
            src    = "★" if p.get("source") == "actual" else "~"
            print(row(f"  {p['name'][:26]:<26}  {p['position']:<4}  "
                      f"{pct(p['prob']):>7}  {xg_str:>7}  {src}"))
        print(sep("╠","─","╣"))
        print(row(f"  {away[:26]:<26}  {'Pos':<4}  {'Prob':>7}  {'xG/90':>7}  Src"))
        print(sep("╠","─","╣"))
        for p in gs["away"][:8]:
            xg_str = f"{p['xg90']:.3f}" if p.get("xg90") else "  —  "
            src    = "★" if p.get("source") == "actual" else "~"
            print(row(f"  {p['name'][:26]:<26}  {p['position']:<4}  "
                      f"{pct(p['prob']):>7}  {xg_str:>7}  {src}"))
        print(sep("╠","─","╣"))
        print(row("  ★ = actual player data   ~ = position-weight estimate"))
        print(sep())

    # Corners
    print(hdr("CORNERS  (Negative Binomial)"))
    print(sep("╠","─","╣"))
    print(row(f"  Expected → Home:{cor['mean_h']:.1f}  Away:{cor['mean_a']:.1f}  "
              f"Total:{cor['mean_t']:.1f}  HT:{cor['mean_ht_corners']:.1f}"))
    print(sep("╠","─","╣"))
    for l in (7.5,8.5,9.5,10.5,11.5,12.5):
        ov,un = cor[f"ov_{l}"],cor[f"un_{l}"]
        fo = "←" if ov>=un else " "
        print(row(f"  Corners O/U {l}   Over:{pct(ov)} {fo}   Under:{pct(un)}"))
    print(sep())

    # Cards + Bookings
    print(hdr("CARDS & BOOKING POINTS  (Negative Binomial)"))
    print(sep("╠","─","╣"))
    print(row(f"  Expected → Home:{crd['mean_h']:.1f}  Away:{crd['mean_a']:.1f}  "
              f"Total:{crd['mean_t']:.1f}  Booking Pts:{crd['mean_booking_pts']:.0f}"))
    print(sep("╠","─","╣"))
    for l in (1.5,2.5,3.5,4.5,5.5):
        ov,un = crd[f"ov_{l}"],crd[f"un_{l}"]
        fo = "←" if ov>=un else " "
        print(row(f"  Cards O/U {l}   Over:{pct(ov)} {fo}   Under:{pct(un)}"))
    print(sep("╠","─","╣"))
    for l in (20.5,30.5,40.5,50.5):
        print(row(f"  Booking Points Over {l}:  {pct(crd[f'bp_ov_{l}'])}"))
    print(sep())

    # Value analysis
    print(hdr("💎  VALUE ANALYSIS  |  EV & QUARTER-KELLY (RWF)"))
    print(sep("╠","─","╣"))

    value_bets = []
    for mkt, mp in mkt_probs.items():
        if mkt in odds:
            e_val    = edge(mp, odds[mkt])
            ev_val   = ev(mp, odds[mkt])
            quality  = bet_quality(mkt, mp, e_val)
            # Include all positive-edge bets in the table (even weak ones)
            # but only recommend bets that pass both conditions
            if e_val > 0:
                value_bets.append({
                    "market" : mkt,
                    "model_p": mp,
                    "implied": implied(odds[mkt]),
                    "edge"   : e_val,
                    "ev"     : ev_val,
                    "odds"   : odds[mkt],
                    "stake"  : kelly_stake(mp, odds[mkt], bankroll)
                             if should_bet(quality) else 0.0,
                    "quality": quality,
                })

    # Sort: bettable bets first (by edge), then observe-only
    value_bets.sort(
        key=lambda x: (should_bet(x["quality"]), x["edge"]),
        reverse=True
    )

    best_market = ""; best_edge = 0.0; best_ev = 0.0; best_stake = 0.0

    if not odds:
        print(row("  No odds entered — EV analysis skipped."))
    elif not value_bets:
        print(row("  ⚠  No positive edge found. Recommendation: SKIP."))
    else:
        print(row(f"  {'Market':<13} {'Model%':>7} {'Implied%':>9} "
                  f"{'Edge':>7} {'EV@1u':>8} {'Stake RWF':>12}  Quality"))
        print(sep("╠","─","╣"))

        bettable = [b for b in value_bets if should_bet(b["quality"])]
        observe  = [b for b in value_bets if not should_bet(b["quality"])]

        for a in bettable:
            icon = quality_icon(a["quality"])
            print(row(f"  {icon} {a['market']:<12} {pct(a['model_p']):>7} "
                      f"{pct(a['implied']):>9} {a['edge']:>+7.1%} "
                      f"{a['ev']:>+8.3f} {a['stake']:>12,.0f}  "
                      f"{'PREMIUM' if a['quality']=='premium' else 'BET'}"))

        if observe:
            print(sep("╠","─","╣"))
            print(row("  ── OBSERVE ONLY (fails probability or edge floor) ──────"))
            for a in observe:
                min_p = MIN_PROB_BY_MARKET.get(a["market"], MIN_PROB_DEFAULT)
                reason = (
                    f"prob {pct(a['model_p'])} < floor {pct(min_p)}"
                    if a["quality"] == "low_prob"
                    else f"edge {a['edge']:+.1%} < {pct(MIN_EDGE)} min"
                    if a["quality"] == "low_edge"
                    else f"fails both conditions"
                )
                print(row(f"  ⚠️  {a['market']:<12} {pct(a['model_p']):>7} "
                          f"{pct(a['implied']):>9} {a['edge']:>+7.1%}  "
                          f"[{reason}]"))

        if bettable:
            best = bettable[0]
            best_market = best["market"]; best_edge = best["edge"]
            best_ev = best["ev"]; best_stake = best["stake"]

            print(sep("╠","═","╣"))
            print(hdr("🎯  RECOMMENDED BET"))
            print(sep("╠","─","╣"))
            pct_of_br = best["stake"] / bankroll * 100 if bankroll else 0
            min_p = MIN_PROB_BY_MARKET.get(best["market"], MIN_PROB_DEFAULT)
            print(row(f"  Market  : {best['market']}"))
            print(row(f"  Odds    : {best['odds']:.2f}"))
            print(row(f"  Model % : {pct(best['model_p'])}  ≥  floor {pct(min_p)}  ✓"))
            print(row(f"  Edge    : {best['edge']:+.1%}  ≥  min {pct(MIN_EDGE)}  ✓"))
            print(row(f"  EV per 1 RWF: {best['ev']:+.4f}"))
            print(row(f"  ▶ STAKE : {best['stake']:,.0f} RWF  ({pct_of_br:.2f}% of bankroll)"))
            print(row(f"  ▶ RETURN: {best['stake'] * best['odds']:,.0f} RWF  if win"))
            if best["quality"] == "premium":
                print(row(f"  🔥 PREMIUM — model prob ≥ {pct(PREMIUM_PROB)} "
                          f"AND edge ≥ {pct(PREMIUM_EDGE)}  (both conditions met)"))
        else:
            print(sep("╠","═","╣"))
            print(hdr("⚠️  NO BETTABLE EDGE FOUND"))
            print(sep("╠","─","╣"))
            print(row("  All positive edges fail the probability floor."))
            print(row("  These are high-edge longshots — variance too high to bet."))
            print(row("  Recommendation: SKIP this game or observe only."))

    print(sep("╠","═","╣"))
    print(hdr("🛡  BANKROLL RULES  (Dual-Condition Filter)"))
    print(sep("╠","─","╣"))
    print(row(f"  • BET only when BOTH conditions are met simultaneously:"))
    print(row(f"      Edge  ≥ {pct(MIN_EDGE)} above implied  AND  Model prob ≥ market floor"))
    print(row(f"  • PREMIUM (🔥): edge ≥ {pct(PREMIUM_EDGE)} AND model prob ≥ {pct(PREMIUM_PROB)}"))
    print(row(f"  • ⚠️ entries show edge but fail probability floor — OBSERVE ONLY."))
    print(row(f"  • Quarter-Kelly (×0.25). Hard cap: {MAX_KELLY_PCT*100:.0f}% of bankroll per bet."))
    print(row(f"  • Drawdown 30% from peak → halve all stakes immediately."))
    print(sep())
    print(row("  ⚠  Statistical model. Gamble responsibly. 18+."))
    print(f"╚{'═'*W}╝\n")

    return {
        "best_market": best_market, "best_edge": best_edge,
        "best_ev": best_ev, "best_stake": best_stake,
    }


# ══════════════════════════════════════════════════════════════════════════════
# UPCOMING FIXTURES DISPLAY
# ══════════════════════════════════════════════════════════════════════════════
def print_upcoming(fixtures: list):
    if not fixtures:
        return
    print(f"\n{'─'*74}")
    print(f"  📅  UPCOMING PREMIER LEAGUE FIXTURES")
    print(f"{'─'*74}")
    for f in fixtures:
        print(f"  {f['date']} {f['time']}   "
              f"{f['home']:<25} vs  {f['away']:<25}  {f['matchday']}")
    print(f"{'─'*74}\n")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print(f"""
╔{'═'*74}╗
║{'':74}║
║{'  ⚽  BAYESIAN FOOTBALL PREDICTION ENGINE  v7.0  ⚽':^74}║
║{'  xG · H2H · Injuries · All Markets · MySQL · Excel Tracker':^74}║
║{'':74}║
╠{'═'*74}╣
║  First run: enter MySQL password + API keys when prompted             ║
║  Every prediction auto-appends to predictions_tracker.xlsx            ║
╚{'═'*74}╝
""")

    # ── First-run setup ────────────────────────────────────────────────────────
    setup_env()

    # ── Database ───────────────────────────────────────────────────────────────
    print("[DB] Connecting to MySQL …", end=" ", flush=True)
    db = DB()
    print("connected ✓")

    # ── API clients ────────────────────────────────────────────────────────────
    afl = APIFootball(db)
    fd  = FootballDataOrg(db)
    dm  = DataManager(db, afl, fd)

    # ── Excel tracker ──────────────────────────────────────────────────────────
    tracker = ExcelTracker(TRACKER_FILE)

    # ── Startup sync & backfill ────────────────────────────────────────────────
    dm.ensure_teams()
    dm.startup_sync()

    # ── Upcoming fixtures ──────────────────────────────────────────────────────
    remaining = afl._remaining()
    if remaining >= 2:
        upcoming = dm.fetch_upcoming()
        if upcoming:
            print_upcoming(upcoming)
        else:
            print("[Fixtures] Could not load upcoming fixtures "
                  "(may need a paid API plan for this feature).\n")
    else:
        print(f"[API] Only {remaining} requests left today — "
              f"skipping upcoming fixtures.\n")

    # ── Team list ──────────────────────────────────────────────────────────────
    all_teams = dm.all_team_names()
    if not all_teams:
        print("❌  No match data found. Check your API keys and MySQL connection.")
        sys.exit(1)

    print(f"✅  {len(all_teams)} teams in database. Ready.")
    print("    'teams' → list all  |  'quit' → exit\n")

    # ── Main loop ──────────────────────────────────────────────────────────────
    while True:
        try:
            raw = input("🎯  Fixture (e.g. Arsenal vs Man City): ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nGoodbye! 🍀"); break

        if not raw: continue
        if raw.lower() in ("quit","exit","q"):
            print("Stay disciplined. 🍀"); break
        if raw.lower() == "teams":
            print()
            for i, t in enumerate(all_teams, 1):
                print(f"  {i:>2}. {t}")
            print(); continue

        # Parse fixture
        delim = None
        for d in (" vs "," VS "," Vs "," v "," V "):
            if d in raw: delim = d; break
        if not delim:
            print("  ❌  Format: 'Team A vs Team B'\n"); continue

        parts = raw.split(delim, 1)
        try:
            home = resolve_team(parts[0].strip(), all_teams)
            away = resolve_team(parts[1].strip(), all_teams)
        except ValueError as e:
            print(f"  ❌  {e}\n"); continue

        if home == away:
            print("  ❌  Teams must be different.\n"); continue

        print(f"\n  ✓  {home}  vs  {away}")

        # Bankroll
        try:
            br = input(f"  💰  Bankroll (RWF, Enter = 20,000): ").strip()
            bankroll = float(br.replace(",","")) if br else 20000.0
        except ValueError:
            bankroll = 20000.0

        # ── Auto fixture lookup ────────────────────────────────────────────────
        fixture_info = dm.find_fixture(home, away)
        fixture_id   = fixture_info.get("fixture_id") if fixture_info else None

        # ── Fixture-first backfill ─────────────────────────────────────────────
        remaining = afl._remaining()
        if remaining >= 3:
            print(f"\n[Backfill] Prioritising stats for {home} & {away} …")
            dm._run_backfill_batch(priority_teams=[home, away])
        else:
            print(f"[API] {remaining} requests left — skipping backfill.")

        # ── Fetch supporting data ──────────────────────────────────────────────
        print()
        injuries_h = dm.fetch_injuries_for_team(home)
        injuries_a = dm.fetch_injuries_for_team(away)
        h2h_df     = dm.fetch_h2h(home, away)

        # ── Standings context ──────────────────────────────────────────────────
        standings = dm.standings_context(home, away) \
                    if afl._remaining() >= 1 else {}

        # ── Team-specific non-goal rates ───────────────────────────────────────
        home_rates = dm.team_non_goal_rates(home)
        away_rates = dm.team_non_goal_rates(away)
        if home_rates:
            print(f"  [Rates] {home}: {home_rates['n_matches']} matches — "
                  f"avg corners {home_rates.get('avg_corners',0):.1f}  "
                  f"avg cards {home_rates.get('avg_cards',0):.1f}")
        else:
            print(f"  [Rates] {home}: using PL averages")
        if away_rates:
            print(f"  [Rates] {away}: {away_rates['n_matches']} matches — "
                  f"avg corners {away_rates.get('avg_corners',0):.1f}  "
                  f"avg cards {away_rates.get('avg_cards',0):.1f}")
        else:
            print(f"  [Rates] {away}: using PL averages")

        # ── Referee ────────────────────────────────────────────────────────────
        # Use auto-detected referee from fixture if available
        auto_ref = (fixture_info or {}).get("referee", "")
        if auto_ref:
            print(f"  [Referee] Auto-detected: {auto_ref}")
            referee_name   = auto_ref
            referee_factor = dm.referee_card_rate(referee_name)
            if referee_factor != 1.0:
                print(f"  [Referee] {referee_name}: ×{referee_factor:.2f} vs PL avg")
        else:
            referee_name = input("  Referee name (Enter to skip): ").strip()
            referee_factor = dm.referee_card_rate(referee_name) \
                             if referee_name else 1.0
            if referee_name and referee_factor != 1.0:
                print(f"  [Referee] {referee_name}: ×{referee_factor:.2f} vs PL avg")

        # ── Weather ────────────────────────────────────────────────────────────
        venue       = (fixture_info or {}).get("venue", "")
        match_dt    = (fixture_info or {}).get("date", "")
        weather     = {}
        if venue and match_dt:
            print(f"  [Weather] Fetching for {venue} …", end=" ", flush=True)
            weather = dm.weather.get_match_weather(venue, match_dt)
            print(weather.get("description", "unavailable"))
        else:
            print("  [Weather] No venue info — weather adjustment skipped.")

        # ── Auto lineups (available ~60 min before kickoff) ───────────────────
        players_h, players_a = [], []
        lineups_auto = False
        if fixture_id and afl._remaining() >= 1:
            players_h, players_a = dm.fetch_auto_lineups(
                fixture_id, home, away
            )
            if players_h or players_a:
                lineups_auto = True
                # Enrich with actual player xG/90 stats
                if afl._remaining() >= 1:
                    players_h = dm.enrich_players_with_xg(
                        players_h, home)
                if afl._remaining() >= 1:
                    players_a = dm.enrich_players_with_xg(
                        players_a, away)

        # Manual lineup entry if auto failed or no fixture ID
        if not lineups_auto:
            print("\n  Lineups not yet released (or no fixture found).")
            do_manual = input(
                "  Enter lineups manually for goalscorer markets? (y/n): "
            ).strip().lower() == "y"
            if do_manual:
                players_h = input_lineup(home)
                players_a = input_lineup(away)

        # ── Injury notes for tracker ───────────────────────────────────────────
        inj_notes_h = "; ".join(
            f"{i['player_name']} ({i['injury_type']})" for i in injuries_h
        ) if injuries_h else "None"
        inj_notes_a = "; ".join(
            f"{i['player_name']} ({i['injury_type']})" for i in injuries_a
        ) if injuries_a else "None"

        # ── Load match data ────────────────────────────────────────────────────
        df_fix = dm.load_for_fixture(home, away)
        if df_fix.empty:
            print("  ❌  No match data found for these teams.\n"); continue
        print(f"  [Data] {len(df_fix)} relevant matches loaded.")

        # ── Fit model ─────────────────────────────────────────────────────────
        print(f"  [Fitting] MAP + Laplace …")
        mdl = FixtureModel(df_fix, home, away, h2h_df,
                           injuries_h, injuries_a, weather)
        mdl.fit()

        # ── Simulate ──────────────────────────────────────────────────────────
        mu_h, mu_a = mdl.lambdas()
        print(f"  [Simulate] {N_SIM:,} draws … ", end="", flush=True)
        s = Sim(mu_h, mu_a, rho=mdl.rho)
        print("done ✓")

        # ── Auto odds — fetch from Bet365 if fixture found ─────────────────────
        odds = {}
        if fixture_id and afl._remaining() >= 1:
            odds = dm.fetch_auto_odds(fixture_id)

        if odds:
            print(f"  [Odds] {len(odds)} markets fetched automatically from Bet365.")
            override = input(
                "  Add/override any odds manually? (y/n): "
            ).strip().lower()
            if override == "y":
                manual = get_odds()
                odds.update(manual)   # manual overrides auto
        else:
            print("  [Odds] Auto odds unavailable — enter manually.")
            odds = get_odds()

        # Print ticket + get value summary
        val = print_ticket(home, away, mdl, s, odds, bankroll,
                           injuries_h, injuries_a, players_h, players_a, h2h_df,
                           home_rates=home_rates, away_rates=away_rates,
                           referee_factor=referee_factor,
                           weather=weather, standings=standings,
                           fixture_info=fixture_info)

        # Build tracker data dict — use same team rates as print_ticket
        home_cor = home_rates.get("avg_corners") if home_rates else None
        away_cor = away_rates.get("avg_corners") if away_rates else None
        home_crd = home_rates.get("avg_cards")   if home_rates else None
        away_crd = away_rates.get("avg_cards")   if away_rates else None

        res  = s.result(); ou = s.ou(); bt = s.btts()
        cs   = s.correct_score(); ht = s.halftime()
        hmg  = s.half_most_goals(); htft_d = s.htft()
        wtn  = s.win_to_nil(); dnb = s.draw_no_bet()
        w_corner = float((weather or {}).get("corner_factor", 1.0))
        cor  = s.corners(home_avg=home_cor, away_avg=away_cor,
                         weather_factor=w_corner)
        crd  = s.cards(home_avg=home_crd, away_avg=away_crd,
                       referee_factor=referee_factor)
        xg   = s.xg()

        top_cs    = cs[cs["score"] != "other"].iloc[0]
        htft_best = max(htft_d, key=htft_d.get)

        tracker_data = {
            "date"             : datetime.now().strftime("%Y-%m-%d %H:%M"),
            "home_team"        : home, "away_team": away,
            "xg_h"             : xg["xg_h"], "xg_a": xg["xg_a"],
            "home_win_pct"     : res["home_win"],
            "draw_pct"         : res["draw"],
            "away_win_pct"     : res["away_win"],
            "over25_pct"       : ou["over_2.5"],
            "under25_pct"      : ou["under_2.5"],
            "over15_pct"       : ou["over_1.5"],
            "over35_pct"       : ou["over_3.5"],
            "btts_yes_pct"     : bt["btts_yes"],
            "btts_no_pct"      : bt["btts_no"],
            "top_correct_score": top_cs["score"],
            "top_cs_pct"       : top_cs["prob"],
            "ht_hw_pct"        : ht["ht_hw"],
            "ht_d_pct"         : ht["ht_d"],
            "ht_aw_pct"        : ht["ht_aw"],
            "first_half_pct"   : hmg["first_half"],
            "second_half_pct"  : hmg["second_half"],
            "equal_pct"        : hmg["equal"],
            "htft"             : htft_d,
            "home_wtn_pct"     : wtn["home_wtn"],
            "away_wtn_pct"     : wtn["away_wtn"],
            "dnb_home_pct"     : dnb["dnb_home"],
            "dnb_away_pct"     : dnb["dnb_away"],
            "exp_corners"      : cor["mean_t"],
            "exp_cards"        : crd["mean_t"],
            "exp_booking_pts"  : crd["mean_booking_pts"],
            "best_market"      : val["best_market"],
            "edge_pct"         : val["best_edge"],
            "ev"               : val["best_ev"],
            "kelly_stake_rwf"  : val["best_stake"],
            "bankroll_rwf"     : bankroll,
            "odds_entered"     : json.dumps(odds),
            "h2h_factor"       : mdl.h2h_factor,
            "inj_h_factor"     : mdl.inj_h_factor,
            "inj_a_factor"     : mdl.inj_a_factor,
            "mom_h"            : mdl.mom_h,
            "mom_a"            : mdl.mom_a,
            "fatigue_h"        : mdl.fatigue_h,
            "fatigue_a"        : mdl.fatigue_a,
            "dc_rho"           : mdl.rho,
            "referee_factor"   : referee_factor,
            "corners_src"      : "team" if cor.get("using_team_data") else "PL_avg",
            "cards_src"        : "team" if crd.get("using_team_data") else "PL_avg",
            "weather_goal"     : mdl.weather_goal_factor,
            "weather_corner"   : mdl.weather_corner_factor,
            "weather_desc"     : (weather or {}).get("description", ""),
            "venue"            : (fixture_info or {}).get("venue", ""),
            "referee_auto"     : (fixture_info or {}).get("referee", ""),
            "fixture_id"       : fixture_id or 0,
            "injury_notes_home": inj_notes_h,
            "injury_notes_away": inj_notes_a,
        }

        # Save to Excel tracker
        tracker.append_prediction(tracker_data)

        # Save to DB predictions log
        db.execute(
            """INSERT INTO predictions_log
               (prediction_date, home_team, away_team,
                home_xg, away_xg, home_win_pct, draw_pct, away_win_pct,
                over25_pct, under25_pct, over15_pct, over35_pct,
                btts_yes_pct, btts_no_pct,
                top_correct_score, top_cs_pct,
                ht_home_win_pct, ht_draw_pct, ht_away_win_pct,
                expected_corners, expected_cards,
                best_market, edge_pct, ev,
                suggested_stake_rwf, bankroll_rwf,
                odds_entered, injury_notes_home, injury_notes_away)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                       %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (tracker_data["date"], home, away,
             xg["xg_h"], xg["xg_a"],
             res["home_win"], res["draw"], res["away_win"],
             ou["over_2.5"], ou["under_2.5"], ou["over_1.5"], ou["over_3.5"],
             bt["btts_yes"], bt["btts_no"],
             top_cs["score"], float(top_cs["prob"]),
             ht["ht_hw"], ht["ht_d"], ht["ht_aw"],
             cor["mean_t"], crd["mean_t"],
             val["best_market"], val["best_edge"], val["best_ev"],
             val["best_stake"], bankroll,
             json.dumps(odds), inj_notes_h, inj_notes_a)
        )

        try:
            again = input("  🔄  Another fixture? (y/n): ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            print("\nGoodbye! 🍀"); break
        if again != "y":
            print("\n  The edge is in the process. Stay disciplined. 🍀\n")
            db.close(); break

    db.close()


if __name__ == "__main__":
    main()